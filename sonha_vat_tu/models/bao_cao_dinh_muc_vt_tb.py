# -*- coding: utf-8 -*-
import base64
import io
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

from odoo import _, api, fields, models
from odoo.exceptions import UserError

from .bao_cao_ghi_chu import REPORT_DMTB

REPORT_MONTH_COUNT = 3
QTY_FIELDS = tuple('qty_t%d' % idx for idx in range(REPORT_MONTH_COUNT))
NVL_QTY_FIELDS = tuple('qty_nvl_t%d' % idx for idx in range(REPORT_MONTH_COUNT))

MDM_TYPE_NVL = 'Nguyên vật liệu'
MDM_TYPE_BTP = 'Bán thành phẩm'
DMTB_ROW_BTP = 'btp'
DMTB_ROW_TOTAL = 'total'
DMTB_BTP_LABEL = 'Bán thành phẩm'


class BaoCaoDinhMucVtTbWizard(models.TransientModel):
    _name = 'bao.cao.dinh.muc.vt.tb.wizard'
    _description = 'Wizard báo cáo định mức vật tư trung bình'

    period_ids = fields.Many2many(
        'ke.hoach.vat.tu',
        'bao_cao_dmtb_wizard_period_rel',
        'wizard_id',
        'period_id',
        string='Kế hoạch',
        help='Chọn nhiều kỳ cùng tháng — có thể nhiều file cùng đơn vị SX '
             '(vd. innox/nhựa tách file); báo cáo gom theo đơn vị sản xuất.',
    )
    nhom_id = fields.Many2one(
        'dmtb.nhom',
        string='Nhóm',
        required=True,
        domain=[('active', '=', True)],
    )
    nguon_sl_sp = fields.Selection(
        [
            ('khkd', 'Kế hoạch kinh doanh'),
            ('khsx', 'Kế hoạch sản xuất'),
        ],
        string='Nguồn',
        required=True,
        default='khkd',
    )
    sl_qty_column_label = fields.Char(string='Nhãn cột SL', readonly=True)
    period_codes = fields.Char(string='Mã các kỳ', readonly=True)
    period_month = fields.Char(string='Tháng kế hoạch', readonly=True)
    column_spec_json = fields.Text(string='Cột tháng (JSON)', readonly=True)
    line_ids = fields.One2many(
        'bao.cao.dinh.muc.vt.tb.line', 'wizard_id', string='Chi tiết')

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        period_id = self.env.context.get('default_period_id')
        if not period_id:
            active_id = self.env.context.get('active_id')
            active_model = self.env.context.get('active_model')
            if active_model == 'ke.hoach.vat.tu' and active_id:
                period_id = active_id
        if period_id and 'period_ids' in fields_list:
            res['period_ids'] = [(6, 0, [period_id])]
        if 'nhom_id' in fields_list and not res.get('nhom_id'):
            default_nhom = self.env['dmtb.nhom'].search(
                [('active', '=', True)], order='name, id', limit=1,
            )
            if default_nhom:
                res['nhom_id'] = default_nhom.id
        return res

    def _plan_model_name(self):
        self.ensure_one()
        return (
            'ke.hoach.kinh.doanh.line' if self.nguon_sl_sp == 'khkd'
            else 'ke.hoach.san.xuat'
        )

    def _nhom_label(self):
        self.ensure_one()
        return self.nhom_id.name or ''

    def _sl_label(self):
        self.ensure_one()
        name = (self.nhom_id.name or '').strip()
        return _('SL %s') % name if name else _('SL sản phẩm')

    def _selected_periods(self):
        self.ensure_one()
        if not self.period_ids:
            raise UserError(_('Vui lòng chọn ít nhất một kỳ kế hoạch.'))
        if not self.nhom_id:
            raise UserError(_('Vui lòng chọn nhóm báo cáo.'))
        if not self.nhom_id.nganh_hang_ids:
            raise UserError(_(
                'Nhóm "%(name)s" chưa cấu hình ngành hàng.',
                name=self.nhom_id.name,
            ))
        periods = self.period_ids.sorted(
            key=lambda p: (
                (p.company_sx_id.company_code or p.company_sx_id.name or '').upper(),
                p.period_month or '',
                p.code or '',
            )
        )
        months = {(p.period_month or '').strip() for p in periods if p.period_month}
        if len(months) > 1:
            raise UserError(_(
                'Các kỳ đã chọn phải cùng tháng kế hoạch (hiện có: %(months)s).',
                months=', '.join(sorted(months)),
            ))
        return periods

    @staticmethod
    def _report_month_keys(period):
        horizon = period._get_horizon_months()
        if len(horizon) < REPORT_MONTH_COUNT:
            raise UserError(_(
                'Kỳ "%(code)s" không xác định được %(count)s tháng báo cáo.',
                code=period.code or period.display_name,
                count=REPORT_MONTH_COUNT,
            ))
        return horizon[:REPORT_MONTH_COUNT]

    @staticmethod
    def _empty_month_metrics():
        return [{'sl_sp': 0.0, 'sl_nvl': 0.0} for _ in range(REPORT_MONTH_COUNT)]

    @staticmethod
    def _classify_ma_kind(mdm_hh_type):
        """Phân loại mã trên kế hoạch: bồn (TP) / BTP / NVL mua thẳng."""
        label = (mdm_hh_type or '').strip()
        if label == MDM_TYPE_BTP:
            return 'btp'
        if label == MDM_TYPE_NVL:
            return 'nvl'
        return 'bo'

    @api.model
    def _ma_meta_maps(self, ma_codes):
        meta_map = self.env['ma.hang'].get_mdm_sap_meta_map(ma_codes)
        kind_map = {}
        nganh_map = {}
        for ma, meta in meta_map.items():
            kind_map[ma] = self._classify_ma_kind(meta.get('mdm_hh_type'))
            nganh_map[ma] = meta.get('nganh_hang_id') or False
        return kind_map, nganh_map, meta_map

    @staticmethod
    def _line_nganh_id(line, nganh_map):
        nh = getattr(line, 'nganh_hang', False)
        if nh:
            return nh.id
        ma = (getattr(line, 'ma_sap', None) or '').strip()
        return nganh_map.get(ma) or False

    def _aggregate_period(self, period, nganh_ids):
        """Tách SL SP / SL NVL theo ngành nhóm + dòng Bán thành phẩm."""
        plan_model = self._plan_model_name()
        period_field = (
            'kinh_doanh_id.period_sx_id'
            if plan_model == 'ke.hoach.kinh.doanh.line'
            else 'period_id'
        )
        nganh_set = set(nganh_ids)
        plan_lines = self.env[plan_model].search([
            (period_field, '=', period.id),
            ('nganh_hang', 'in', list(nganh_ids)),
        ])
        chi_tiet = self.env['tinh.toan.vat.tu.chi.tiet'].search([
            ('period_id', '=', period.id),
        ])

        ma_codes = set()
        for line in plan_lines:
            ma = (line.ma_sap or '').strip()
            if ma:
                ma_codes.add(ma)
        for row in chi_tiet:
            ma = (row.ma or '').strip()
            if ma:
                ma_codes.add(ma)

        kind_map, nganh_map, _meta = self._ma_meta_maps(ma_codes)

        for line in plan_lines:
            ma = (line.ma_sap or '').strip()
            if not ma:
                continue
            nh_id = self._line_nganh_id(line, nganh_map)
            if nh_id:
                nganh_map[ma] = nh_id

        metrics = {
            str(nh_id): self._empty_month_metrics() for nh_id in nganh_ids
        }
        metrics[DMTB_ROW_BTP] = self._empty_month_metrics()

        for line in plan_lines:
            ma = (line.ma_sap or '').strip()
            if not ma:
                continue
            kind = kind_map.get(ma, 'bo')
            if kind != 'bo':
                continue
            nh_id = self._line_nganh_id(line, nganh_map)
            if nh_id not in nganh_set:
                continue
            bucket = str(nh_id)
            for idx, qty_field in enumerate(QTY_FIELDS):
                metrics[bucket][idx]['sl_sp'] += getattr(line, qty_field) or 0.0

        for row in chi_tiet:
            ma = (row.ma or '').strip()
            if not ma:
                continue
            kind = kind_map.get(ma, 'bo')
            if kind in ('btp', 'nvl'):
                bucket = DMTB_ROW_BTP
            else:
                nh_id = nganh_map.get(ma) or False
                if nh_id not in nganh_set:
                    continue
                bucket = str(nh_id)
            for idx, qty_field in enumerate(NVL_QTY_FIELDS):
                metrics[bucket][idx]['sl_nvl'] += getattr(row, qty_field) or 0.0

        month_keys = self._report_month_keys(period)
        return month_keys, metrics

    def _load_ghi_chu_map(self, periods, nhom_id, nguon_sl_sp):
        GhiChu = self.env['bao.cao.ghi.chu'].sudo()
        period_key = GhiChu.period_key_from_periods(periods)
        if not period_key:
            return {}
        prefix = '%s|%s|' % (nhom_id or 0, nguon_sl_sp or '')
        rows = GhiChu.search([
            ('report_type', '=', REPORT_DMTB),
            ('period_key', '=', period_key),
            ('scope_key', '=like', prefix + '%'),
        ])
        out = {}
        for rec in rows:
            if not rec.scope_key.startswith(prefix):
                continue
            tail = rec.scope_key[len(prefix):]
            parts = tail.split('|', 1)
            if len(parts) != 2:
                continue
            try:
                sx_id = int(parts[0])
            except ValueError:
                continue
            row_key = parts[1]
            if rec.noi_dung:
                out[(sx_id, row_key)] = rec.noi_dung
        return out

    def _row_specs(self, nhom):
        specs = []
        for seq, nh in enumerate(nhom.nganh_hang_ids, start=1):
            specs.append({
                'row_key': str(nh.id),
                'nganh_hang_id': nh.id,
                'nganh_hang_label': nh.ten or nh.name or '',
                'is_btp_row': False,
                'sequence': seq,
            })
        specs.append({
            'row_key': DMTB_ROW_BTP,
            'nganh_hang_id': False,
            'nganh_hang_label': DMTB_BTP_LABEL,
            'is_btp_row': True,
            'sequence': len(specs) + 1,
        })
        return specs

    @staticmethod
    def _metrics_has_data(cells):
        return any((c.get('sl_sp') or 0.0) or (c.get('sl_nvl') or 0.0) for c in cells)

    @staticmethod
    def _sum_company_total_metrics(merged, row_specs):
        """Tổng công ty: SL SP = cộng các dòng ngành; SL NVL = cộng cả Bán TP."""
        total = BaoCaoDinhMucVtTbWizard._empty_month_metrics()
        for spec in row_specs:
            cells = merged.get(spec['row_key']) or BaoCaoDinhMucVtTbWizard._empty_month_metrics()
            for idx, cell in enumerate(cells):
                if not spec['is_btp_row']:
                    total[idx]['sl_sp'] += cell.get('sl_sp') or 0.0
                total[idx]['sl_nvl'] += cell.get('sl_nvl') or 0.0
        return total

    def _populate_lines(self):
        self.ensure_one()
        periods = self._selected_periods()
        nhom = self.nhom_id
        nganh_ids = nhom.nganh_hang_ids.ids
        nhom_label = self._nhom_label()
        row_specs = self._row_specs(nhom)
        ghi_chu_map = self._load_ghi_chu_map(periods, nhom.id, self.nguon_sl_sp)

        column_keys = self._report_month_keys(periods[0])
        column_spec = [{'month_key': mk, 'label': mk} for mk in column_keys]
        self.column_spec_json = json.dumps(column_spec, ensure_ascii=False)
        self.period_codes = ', '.join(p.code or '' for p in periods if p.code)
        self.period_month = periods[0].period_month or ''
        self.sl_qty_column_label = self._sl_label()

        Line = self.env['bao.cao.dinh.muc.vt.tb.line']
        self.line_ids.unlink()

        groups = {}
        for period in periods:
            sx = period.company_sx_id
            if not sx:
                raise UserError(_(
                    'Kỳ "%(code)s" chưa có đơn vị sản xuất.',
                    code=period.code or period.display_name,
                ))
            groups.setdefault(sx.id, {'sx': sx, 'periods': []})
            groups[sx.id]['periods'].append(period)

        lines = []
        sorted_groups = sorted(
            groups.values(),
            key=lambda g: (
                (g['sx'].company_code or g['sx'].name or '').upper(),
            ),
        )
        for group in sorted_groups:
            sx = group['sx']
            merged = {
                spec['row_key']: self._empty_month_metrics()
                for spec in row_specs
            }

            has_any = False
            for period in group['periods']:
                month_keys, period_metrics = self._aggregate_period(period, nganh_ids)
                if month_keys != column_keys:
                    raise UserError(_(
                        'Kỳ "%(code)s" có dải tháng khác kỳ đầu tiên.',
                        code=period.code or period.display_name,
                    ))
                if not self._metrics_has_data(
                    [cell for bucket in period_metrics.values() for cell in bucket]
                ):
                    continue
                has_any = True
                for row_key, cells in period_metrics.items():
                    if row_key not in merged:
                        continue
                    for idx, cell in enumerate(cells):
                        merged[row_key][idx]['sl_sp'] += cell.get('sl_sp') or 0.0
                        merged[row_key][idx]['sl_nvl'] += cell.get('sl_nvl') or 0.0

            if not has_any:
                raise UserError(_(
                    'Đơn vị SX "%(sx)s" không có SL sản phẩm hoặc SL NVL '
                    'trong 3 tháng đầu cho nhóm "%(nhom)s".',
                    sx=sx.company_code or sx.name,
                    nhom=nhom_label,
                ))

            total_cells = self._sum_company_total_metrics(merged, row_specs)
            lines.append({
                'wizard_id': self.id,
                'period_id': group['periods'][0].id,
                'nhom_id': nhom.id,
                'company_sx_id': sx.id,
                'company_code': sx.company_code or sx.name or '',
                'nganh_hang_id': False,
                'nganh_hang_label': '',
                'is_btp_row': False,
                'is_total_row': True,
                'row_key': DMTB_ROW_TOTAL,
                'sequence': 0,
                'metrics_json': json.dumps(total_cells, ensure_ascii=False),
                'ghi_chu': ghi_chu_map.get((sx.id, DMTB_ROW_TOTAL), ''),
            })

            for spec in row_specs:
                row_key = spec['row_key']
                cells = merged.get(row_key) or self._empty_month_metrics()
                lines.append({
                    'wizard_id': self.id,
                    'period_id': group['periods'][0].id,
                    'nhom_id': nhom.id,
                    'company_sx_id': sx.id,
                    'company_code': sx.company_code or sx.name or '',
                    'nganh_hang_id': spec['nganh_hang_id'],
                    'nganh_hang_label': spec['nganh_hang_label'],
                    'is_btp_row': spec['is_btp_row'],
                    'is_total_row': False,
                    'row_key': row_key,
                    'sequence': spec['sequence'],
                    'metrics_json': json.dumps(cells, ensure_ascii=False),
                    'ghi_chu': ghi_chu_map.get((sx.id, row_key), ''),
                })

        if lines:
            Line.create(lines)

    def action_open_report(self):
        self.ensure_one()
        self._populate_lines()
        title = _('Định mức vật tư trung bình')
        if self.nhom_id:
            title = _('%s — %s') % (title, self._nhom_label())
        if self.period_codes:
            title = _('%s (%s)') % (title, self.period_codes)
        return {
            'type': 'ir.actions.act_window',
            'name': title,
            'res_model': 'bao.cao.dinh.muc.vt.tb.line',
            'view_mode': 'tree',
            'domain': [('wizard_id', '=', self.id)],
            'context': {
                'bao_cao_dmtb_wizard_id': self.id,
                'bao_cao_dmtb_columns': self.column_spec_json or '[]',
                'bao_cao_dmtb_sl_label': self.sl_qty_column_label or 'SL sản phẩm',
            },
        }

    def action_export_excel(self):
        self.ensure_one()
        if not self.line_ids:
            self._populate_lines()
        sl_label = self.sl_qty_column_label or 'SL sản phẩm'
        lines = self.line_ids
        if not lines:
            raise UserError(_('Không có dữ liệu để xuất Excel.'))

        try:
            column_spec = json.loads(self.column_spec_json or '[]')
        except (TypeError, ValueError):
            raise UserError(_('Không xác định được cột tháng hiển thị.'))
        if not column_spec:
            raise UserError(_('Không xác định được cột tháng hiển thị.'))

        wb = Workbook()
        ws = wb.active
        ws.title = 'Dinh muc VT TB'

        nguon_label = dict(self._fields['nguon_sl_sp'].selection).get(
            self.nguon_sl_sp, self.nguon_sl_sp
        )
        ws.cell(row=1, column=1, value='Báo cáo định mức vật tư trung bình')
        ws.cell(row=2, column=1, value='Kỳ')
        ws.cell(row=2, column=2, value=self.period_codes)
        ws.cell(row=3, column=1, value='Tháng kế hoạch')
        ws.cell(row=3, column=2, value=self.period_month or '')
        ws.cell(row=4, column=1, value='Nhóm')
        ws.cell(row=4, column=2, value=self._nhom_label())
        ws.cell(row=5, column=1, value='Nguồn')
        ws.cell(row=5, column=2, value=nguon_label)

        header_row1 = 7
        header_row2 = 8
        data_row = 9
        yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        col = 1
        for fixed in ('Công ty', 'Ngành hàng'):
            ws.merge_cells(
                start_row=header_row1, start_column=col,
                end_row=header_row2, end_column=col,
            )
            ws.cell(row=header_row1, column=col, value=fixed)
            col += 1

        for col_def in column_spec:
            month_key = col_def.get('label') or col_def.get('month_key') or ''
            group_start = col
            ws.merge_cells(
                start_row=header_row1, start_column=group_start,
                end_row=header_row1, end_column=group_start + 2,
            )
            ws.cell(row=header_row1, column=group_start, value=month_key)
            for sub in (sl_label, 'SL NVL (kg)', 'Vật tư bình quân'):
                cell = ws.cell(row=header_row2, column=col, value=sub)
                if sub == 'Vật tư bình quân':
                    cell.fill = yellow
                col += 1

        ws.cell(row=header_row1, column=col, value='Ghi chú')
        ws.merge_cells(
            start_row=header_row1, start_column=col,
            end_row=header_row2, end_column=col,
        )
        ghi_chu_col = col
        max_col = col

        row_idx = data_row
        sorted_lines = lines.sorted(
            key=lambda l: (
                (l.company_code or '').upper(),
                l.sequence or 0,
                l.id,
            )
        )
        company_start_row = None
        company_code_prev = None
        for line in sorted_lines:
            if line.company_code != company_code_prev:
                if company_start_row and row_idx > company_start_row:
                    ws.merge_cells(
                        start_row=company_start_row, start_column=1,
                        end_row=row_idx - 1, end_column=1,
                    )
                    ws.cell(row=company_start_row, column=1).alignment = Alignment(
                        horizontal='center', vertical='center',
                    )
                company_start_row = row_idx
                company_code_prev = line.company_code
            ws.cell(row=row_idx, column=1, value=line.company_code)
            ws.cell(row=row_idx, column=2, value=line.nganh_hang_label or '')
            col_idx = 3
            for cell in line._metrics_list():
                sp = cell.get('sl_sp') or 0.0
                nvl = cell.get('sl_nvl') or 0.0
                show_dmbq = sp and not line.is_btp_row
                dmbq = (nvl / sp) if show_dmbq else 0.0
                sp_val = sp if (sp and not line.is_btp_row) else None
                ws.cell(row=row_idx, column=col_idx, value=sp_val)
                col_idx += 1
                ws.cell(row=row_idx, column=col_idx, value=nvl or None)
                col_idx += 1
                cell_xl = ws.cell(row=row_idx, column=col_idx, value=dmbq or None)
                cell_xl.fill = yellow
                col_idx += 1
            if not line.is_total_row:
                ws.cell(row=row_idx, column=ghi_chu_col, value=line.ghi_chu or '')
            row_idx += 1
        if company_start_row and row_idx > company_start_row:
            ws.merge_cells(
                start_row=company_start_row, start_column=1,
                end_row=row_idx - 1, end_column=1,
            )
            ws.cell(row=company_start_row, column=1).alignment = Alignment(
                horizontal='center', vertical='center',
            )

        self.env['ke.hoach.vat.tu']._apply_b5_export_style(
            ws, header_row1, header_row2, max_col,
            [(None, 'text')] * max_col,
            meta_rows=5,
        )
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions[get_column_letter(ghi_chu_col)].width = 40

        output = io.BytesIO()
        wb.save(output)
        fname = 'DinhMucVT_TB.xlsx'
        attachment = self.env['ir.attachment'].sudo().create({
            'name': fname,
            'type': 'binary',
            'datas': base64.b64encode(output.getvalue()),
            'mimetype': (
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            ),
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }


class BaoCaoDinhMucVtTbLine(models.TransientModel):
    _name = 'bao.cao.dinh.muc.vt.tb.line'
    _inherit = ['bao.cao.ghi.chu.line.mixin']
    _description = 'Dòng báo cáo định mức vật tư trung bình'
    _order = 'company_code, sequence, id'

    wizard_id = fields.Many2one(
        'bao.cao.dinh.muc.vt.tb.wizard', ondelete='cascade', index=True)
    period_id = fields.Many2one(
        'ke.hoach.vat.tu', string='Kỳ nguồn', readonly=True)
    nhom_id = fields.Many2one(
        'dmtb.nhom', string='Nhóm', readonly=True)
    company_sx_id = fields.Many2one(
        'res.company', string='Đơn vị sản xuất', readonly=True)
    company_code = fields.Char(string='Công ty', index=True)
    nganh_hang_id = fields.Many2one(
        'mdm.nganh.hang', string='Ngành hàng', readonly=True)
    nganh_hang_label = fields.Char(string='Ngành hàng', index=True)
    is_btp_row = fields.Boolean(string='Dòng Bán TP', default=False, index=True)
    is_total_row = fields.Boolean(string='Dòng tổng', default=False, index=True)
    row_key = fields.Char(string='Khóa dòng', index=True)
    sequence = fields.Integer(string='Thứ tự', default=10, index=True)
    metrics_json = fields.Text(string='Số liệu theo cột', readonly=True)
    ghi_chu = fields.Text(string='Ghi chú')

    nguon_sl_sp = fields.Selection(
        related='wizard_id.nguon_sl_sp', string='Nguồn', readonly=True)

    def _metrics_list(self):
        self.ensure_one()
        try:
            data = json.loads(self.metrics_json or '[]')
        except (TypeError, ValueError):
            return []
        return data if isinstance(data, list) else []

    def _sync_ghi_chu_to_master(self):
        GhiChu = self._ghi_chu_master()
        for rec in self:
            wizard = rec.wizard_id
            nhom = rec.nhom_id or wizard.nhom_id
            if not wizard or not nhom or not rec.company_sx_id or rec.is_total_row:
                continue
            period_key = rec._ghi_chu_period_key(wizard)
            scope = GhiChu.scope_key_dmtb(
                nhom.id,
                wizard.nguon_sl_sp,
                rec.company_sx_id.id,
                rec.row_key or (DMTB_ROW_BTP if rec.is_btp_row else rec.nganh_hang_id),
            )
            GhiChu.upsert_note(REPORT_DMTB, period_key, scope, rec.ghi_chu)

    def action_export_excel(self):
        if self:
            wizards = self.mapped('wizard_id')
        else:
            wizard_id = self.env.context.get('bao_cao_dmtb_wizard_id')
            wizards = (
                self.env['bao.cao.dinh.muc.vt.tb.wizard'].browse(wizard_id)
                if wizard_id else self.env['bao.cao.dinh.muc.vt.tb.wizard']
            )
        if len(wizards) != 1:
            raise UserError(_('Vui lòng xuất Excel từ một báo cáo đã mở.'))
        return wizards.action_export_excel()
