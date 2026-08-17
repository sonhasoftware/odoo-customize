# -*- coding: utf-8 -*-
import base64
import io
import warnings

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from markupsafe import Markup

from odoo import api, models, _
from odoo.exceptions import UserError

HEADER_FONT_COLOR = 'FFFFFF'
HEADER_FILL_COLOR = '3F6F8F'
HEADER_BORDER_COLOR = '2F556D'
EXCEL_FONT_NAME = 'Times New Roman'
EXCEL_FONT_SIZE = 10
DV_LAST_ROW = 5000


class VatTuExcelMixin(models.AbstractModel):
    """Đọc/ghi file Excel dùng chung cho các wizard import và các action export."""
    _name = 'vat.tu.excel.mixin'
    _description = 'Helper Excel dùng chung cho vật tư'

    TEMPLATE_SHEET_NAME = False
    DATA_START_ROW = 2
    ERROR_LIMIT = 80

    # ------------------------------------------------------------------
    # Đọc file
    # ------------------------------------------------------------------
    @api.model
    def _cell(self, row, col):
        return row[col] if col < len(row) else None

    @api.model
    def _normalize_ma_nvl(self, value):
        """Excel hay trả mã NVL toàn số về dạng float — đưa lại về chuỗi gốc."""
        if value in (None, ''):
            return ''
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        if isinstance(value, int):
            return str(value)
        text = str(value).strip()
        if text.endswith('.0') and text[:-2].isdigit():
            return text[:-2]
        return text

    @api.model
    def _parse_number(self, value, default=None, label=None):
        """Đọc số từ ô Excel. `default=None` nghĩa là ô rỗng bị coi là lỗi."""
        label = label or _('số lượng')
        if value in (None, ''):
            if default is None:
                raise UserError(_('Thiếu %s.') % label)
            return default
        if isinstance(value, bool):
            return float(value)
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip().replace(' ', '').replace(',', '.').rstrip('%')
        try:
            return float(text)
        except ValueError:
            raise UserError(_('Không đọc được %s "%s".') % (label, value))

    def _load_workbook(self):
        self.ensure_one()
        if not self.file_data:
            raise UserError(_('Vui lòng chọn file Excel.'))
        try:
            data = base64.b64decode(self.file_data)
            with warnings.catch_warnings():
                warnings.filterwarnings(
                    'ignore', message='Data Validation extension', category=UserWarning)
                return load_workbook(io.BytesIO(data), data_only=True)
        except Exception as exc:
            raise UserError(_('Không đọc được file Excel: %s') % exc)

    def _get_import_worksheet(self, workbook=None):
        workbook = workbook or self._load_workbook()
        if self.TEMPLATE_SHEET_NAME and self.TEMPLATE_SHEET_NAME in workbook.sheetnames:
            return workbook[self.TEMPLATE_SHEET_NAME]
        for name in workbook.sheetnames:
            if not name.startswith('_'):
                return workbook[name]
        return workbook.active

    def _read_data_rows(self, data_start_row=None):
        ws = self._get_import_worksheet()
        return list(ws.iter_rows(
            min_row=data_start_row or self.DATA_START_ROW, values_only=True))

    # ------------------------------------------------------------------
    # Đơn vị
    # ------------------------------------------------------------------
    @api.model
    def _company_by_code(self):
        return {
            (c.company_code or '').strip(): c
            for c in self.env['res.company'].sudo().search([])
            if (c.company_code or '').strip()
        }

    @api.model
    def _company_codes(self):
        return sorted(self._company_by_code())

    # ------------------------------------------------------------------
    # Ghi template
    # ------------------------------------------------------------------
    def _apply_company_code_validation(self, wb, ws, first_data_row=None, company_codes=None):
        """Thả dropdown mã đơn vị lên cột A, lấy nguồn từ sheet ẩn."""
        company_codes = self._company_codes() if company_codes is None else company_codes
        if not company_codes:
            return
        ref_sheet = wb.create_sheet('_company_codes')
        ref_sheet.sheet_state = 'hidden'
        for row_idx, code in enumerate(company_codes, start=1):
            ref_sheet.cell(row=row_idx, column=1, value=code)
        dv = DataValidation(
            type='list',
            formula1="='_company_codes'!$A$1:$A$%d" % len(company_codes),
            allow_blank=False,
        )
        dv.error = _('Chỉ được chọn mã đơn vị có trong danh mục.')
        dv.errorTitle = _('Đơn vị không hợp lệ')
        ws.add_data_validation(dv)
        dv.add('A%d:A%d' % (first_data_row or self.DATA_START_ROW, DV_LAST_ROW))

    @api.model
    def _style_excel_header(self, ws, max_col, header_row=1, freeze=True):
        side = Side(style='thin', color=HEADER_BORDER_COLOR)
        header_font = Font(
            name=EXCEL_FONT_NAME, size=EXCEL_FONT_SIZE, bold=True, color=HEADER_FONT_COLOR)
        header_fill = PatternFill(fill_type='solid', fgColor=HEADER_FILL_COLOR)
        header_border = Border(left=side, right=side, top=side, bottom=side)
        header_align = Alignment(horizontal='center', vertical='center')
        for cell in ws[header_row][:max_col]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = header_align
        ws.row_dimensions[header_row].height = 22
        if freeze:
            ws.freeze_panes = 'A%d' % (header_row + 1)

    @api.model
    def _style_excel_body(self, ws, max_col, first_row, last_row):
        """Chỉ tô các dòng thực sự có dữ liệu"""
        if not last_row or last_row < first_row:
            return
        body_font = Font(name=EXCEL_FONT_NAME, size=EXCEL_FONT_SIZE)
        body_align = Alignment(vertical='center')
        for row in ws.iter_rows(
            min_row=first_row, max_row=last_row, min_col=1, max_col=max_col,
        ):
            for cell in row:
                cell.font = body_font
                cell.alignment = body_align

    @api.model
    def _set_excel_widths(self, ws, widths):
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _xlsx_download_action(self, wb, filename):
        output = io.BytesIO()
        wb.save(output)
        attachment = self.env['ir.attachment'].sudo().create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(output.getvalue()),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }

    # ------------------------------------------------------------------
    # Kết quả import
    # ------------------------------------------------------------------
    def _post_period_import_file_log(self, period, body_html):
        """Ghi log import file lên chatter kỳ kế hoạch kèm file Excel."""
        self.ensure_one()
        if not period or not self.file_data:
            return
        filename = self.file_name or 'import.xlsx'
        attachment = self.env['ir.attachment'].sudo().create({
            'name': filename,
            'type': 'binary',
            'datas': self.file_data,
            'res_model': 'ke.hoach.vat.tu',
            'res_id': period.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        period.message_post(body=Markup(body_html), attachment_ids=[attachment.id])

    @api.model
    def _raise_import_errors(self, errors, header=None):
        if not errors:
            return
        message = '\n'.join('- %s' % err for err in errors[:self.ERROR_LIMIT])
        if len(errors) > self.ERROR_LIMIT:
            message += _('\n... còn %d lỗi khác.') % (len(errors) - self.ERROR_LIMIT)
        raise UserError('%s\n%s' % (header, message) if header else message)

    @api.model
    def _notify_and_close(self, title, message, success=True, next_action=None):
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': title,
                'message': message or title,
                'type': 'success' if success else 'warning',
                'sticky': False,
                'next': next_action or {'type': 'ir.actions.act_window_close'},
            },
        }
