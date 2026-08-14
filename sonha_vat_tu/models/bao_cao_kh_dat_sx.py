# -*- coding: utf-8 -*-
import base64
import io
import json
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from odoo import _, api, fields, models
from odoo.exceptions import UserError

from .bao_cao_ghi_chu import REPORT_KH_DSX, BaoCaoGhiChu

REPORT_MONTH_COUNT = 4
QTY_FIELDS = tuple('qty_t%d' % idx for idx in range(REPORT_MONTH_COUNT))

KH_DSX_METRIC_GROUPS = (
    ('qty_sx', 'Kế hoạch sản xuất'),
    ('qty_kd', 'Kế hoạch kinh doanh đặt sản xuất'),
    ('qty_cl', 'Chênh lệch KHSX-KH đặt hàng'),
    ('ty_le', 'Tỷ lệ chênh lệch'),
)


class BaoCaoKhDatSxWizard(models.TransientModel):
    _name = 'bao.cao.kh.dat.sx.wizard'
    _description = 'Wizard Biểu 5 — Tổng hợp KH đặt sản xuất (Bảng 2 KHSX)'

    period_ids = fields.Many2many(
        'ke.hoach.vat.tu',
        'bao_cao_kh_dsx_wizard_period_rel',
        'wizard_id',
        'period_id',
        string='Kế hoạch',
        help='Chọn nhiều kỳ cùng tháng — có thể nhiều file cùng đơn vị SX; '
             'báo cáo gom theo đơn vị sản xuất.',
    )
    period_month = fields.Char(string='Tháng kế hoạch', readonly=True)
    ton_kho_month = fields.Char(
        string='Tháng tồn đầu kỳ',
        readonly=True,
        help='Tồn đầu kỳ = ton_cuoi SAP tháng ngay trước tháng bắt đầu kế hoạch.',
    )
    column_spec_json = fields.Text(string='Cột tháng (JSON)', readonly=True)
    line_ids = fields.One2many(
        'bao.cao.kh.dat.sx.line', 'wizard_id', string='Chi tiết')

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        period_id = self.env.context.get('default_period_id')
        if not period_id:
            active_id = self.env.context.get('active_id')
            active_model = self.env.context.get('active_model')
            if active_model == 'ke.hoach.vat.tu' and active_id:
                period_id = active_id
        if period_id and 'period_ids' in fields_list:
            res['period_ids'] = [(6, 0, [period_id])]
        return res

    @staticmethod
    def _prev_month_label(period_month):
        text = (period_month or '').strip()
        if not text:
            return ''
        try:
            dt = datetime.strptime(text, '%m/%Y')
        except ValueError:
            return text
        if dt.month == 1:
            return '12/%d' % (dt.year - 1)
        return '%02d/%d' % (dt.month - 1, dt.year)

    @staticmethod
    def _report_month_keys(period):
        horizon = period._get_horizon_months()
        if len(horizon) < REPORT_MONTH_COUNT:
            raise UserError(_(
                'Kỳ "%(code)s" không xác định được %(count)s tháng báo cáo.',
                code=period.code or period.display_name,
                count=REPORT_MONTH_COUNT,
            ))
        return horizon[:REPORT_MONTH_COUNT]

    def _selected_periods(self):
        self.ensure_one()
        if not self.period_ids:
            raise UserError(_('Vui lòng chọn ít nhất một kỳ kế hoạch.'))
        periods = self.period_ids.sorted(
            key=lambda p: (
                (p.company_sx_id.company_code or p.company_sx_id.name or '').upper(),
                p.period_month or '',
                p.code or '',
            )
        )
        months = {(p.period_month or '').strip() for p in periods if p.period_month}
        if len(months) > 1:
            raise UserError(_(
                'Các kỳ đã chọn phải cùng tháng kế hoạch (hiện có: %(months)s).',
                months=', '.join(sorted(months)),
            ))
        for period in periods:
            if not period.company_sx_id:
                raise UserError(_(
                    'Kỳ "%(code)s" chưa có đơn vị sản xuất.',
                    code=period.code or period.display_name,
                ))
        return periods

    @api.model
    def _sap_branch_sql(self, sx_company_code):
        code = (sx_company_code or '').strip().upper()
        if code == 'BNH':
            return "chi_nhanh LIKE '21%%'"
        if code == 'SSP':
            return "chi_nhanh LIKE '22%%'"
        return "chi_nhanh NOT LIKE '10%%'"

    @api.model
    def _load_sap_ton_cuoi_map(self, ma_codes, month_key, sx_company_code):
        """{ma_sap: ton_cuoi} — cùng logic B4 (md_sap_ton_kho, tháng T-1)."""
        codes = sorted({(c or '').strip() for c in ma_codes if (c or '').strip()})
        if not codes or not month_key:
            return {}
        cr = self.env.cr
        cr.execute("SELECT to_regclass('public.md_sap_ton_kho')")
        if not cr.fetchone()[0]:
            return {}
        cr.execute("SELECT to_regclass('public.safe_sap_numeric')")
        if not cr.fetchone()[0]:
            return {}
        branch_filter = self._sap_branch_sql(sx_company_code)
        cr.execute(
            """
            WITH sap_rows AS (
                SELECT
                    TRIM(mtk.ma_hang) AS ma_hang,
                    mtk.chi_nhanh,
                    mtk.create_date,
                    mtk.id,
                    safe_sap_numeric(mtk.ton_cuoi) AS ton_cuoi
                FROM md_sap_ton_kho mtk
                WHERE TRIM(mtk.ma_hang) = ANY(%(codes)s)
                  AND fn_md_sap_ton_kho_month_key(
                          mtk.from_date, mtk.to_date, mtk.tu_ngay, mtk.den_ngay, mtk.create_date
                      ) = %(month_key)s
                  AND safe_sap_numeric(mtk.ton_cuoi) <> 0
                  AND """
            + branch_filter
            + """
            ),
            latest AS (
                SELECT DISTINCT ON (ma_hang, chi_nhanh)
                    ma_hang, ton_cuoi
                FROM sap_rows
                ORDER BY ma_hang, chi_nhanh, create_date DESC, id DESC
            )
            SELECT ma_hang, SUM(ton_cuoi) AS ton
            FROM latest
            GROUP BY ma_hang
            """,
            {'codes': codes, 'month_key': month_key},
        )
        return {row[0]: row[1] or 0.0 for row in cr.fetchall()}

    def _merge_period_buckets(self, period, ton_kho_month, buckets):
        sx = period.company_sx_id
        sx_code = sx.company_code or sx.name or ''

        kd_lines = self.env['ke.hoach.kinh.doanh.line'].sudo().search([
            ('kinh_doanh_id.period_sx_id', '=', period.id),
        ])
        period_ma = {
            (kd.ma_sap or '').strip()
            for kd in kd_lines if (kd.ma_sap or '').strip()
        }
        ton_map = self._load_sap_ton_cuoi_map(period_ma, ton_kho_month, sx_code)

        sx_lines = period.ke_hoach_san_xuat_ids
        sx_by_key = {
            (line.company_id.id, (line.ma_sap or '').strip()): line
            for line in sx_lines if (line.ma_sap or '').strip()
        }

        for kd in kd_lines:
            ma = (kd.ma_sap or '').strip()
            if not ma:
                continue
            nganh = kd.nganh_hang.ten if kd.nganh_hang else ''
            key = (sx.id, kd.company_id.id, nganh)
            if key not in buckets:
                buckets[key] = {
                    'period_id': period.id,
                    'company_sx_id': sx.id,
                    'company_dat_id': kd.company_id.id,
                    'nganh_hang': nganh,
                    'ton_dau_ky': 0.0,
                    'qty_kd': [0.0] * REPORT_MONTH_COUNT,
                    'qty_sx': [0.0] * REPORT_MONTH_COUNT,
                    '_ma_ton_seen': set(),
                    '_ma_sx_seen': set(),
                }
            bucket = buckets[key]
            for idx in range(REPORT_MONTH_COUNT):
                bucket['qty_kd'][idx] += getattr(kd, QTY_FIELDS[idx]) or 0.0

            if ma not in bucket['_ma_ton_seen']:
                bucket['_ma_ton_seen'].add(ma)
                bucket['ton_dau_ky'] += ton_map.get(ma, 0.0)

            if ma not in bucket['_ma_sx_seen']:
                bucket['_ma_sx_seen'].add(ma)
                sx_line = sx_by_key.get((kd.company_id.id, ma))
                if sx_line:
                    for idx in range(REPORT_MONTH_COUNT):
                        bucket['qty_sx'][idx] += getattr(sx_line, QTY_FIELDS[idx]) or 0.0

    def _buckets_to_line_vals(self, buckets, month_keys, ghi_chu_map=None):
        ghi_chu_map = ghi_chu_map or {}
        GhiChu = self.env['bao.cao.ghi.chu'].sudo()
        sx_codes = {
            rec.id: (rec.company_code or rec.name or '').strip()
            for rec in self.env['res.company'].sudo().browse({
                b['company_sx_id'] for b in buckets.values()
            })
        }
        out = []
        for bucket in sorted(
            buckets.values(),
            key=lambda b: (
                sx_codes.get(b['company_sx_id'], ''),
                b['company_dat_id'],
                b['nganh_hang'] or '',
            ),
        ):
            metrics = []
            for idx, month_key in enumerate(month_keys):
                kd_qty = bucket['qty_kd'][idx]
                sx_qty = bucket['qty_sx'][idx]
                cl = sx_qty - kd_qty
                metrics.append({
                    'month_key': month_key,
                    'qty_sx': sx_qty,
                    'qty_kd': kd_qty,
                    'qty_cl': cl,
                    'ty_le': (cl / kd_qty) if kd_qty else 0.0,
                })
            scope = GhiChu.scope_key_khdsx(
                bucket['company_sx_id'],
                bucket['company_dat_id'],
                bucket['nganh_hang'],
            )
            out.append({
                'period_id': bucket['period_id'],
                'company_sx_id': bucket['company_sx_id'],
                'company_dat_id': bucket['company_dat_id'],
                'nganh_hang': bucket['nganh_hang'],
                'ton_dau_ky': bucket['ton_dau_ky'],
                'metrics_json': json.dumps(metrics, ensure_ascii=False),
                'ghi_chu': ghi_chu_map.get(scope, ''),
                'wizard_id': self.id,
            })
        return out

    def _populate_lines(self):
        self.ensure_one()
        periods = self._selected_periods()
        period_month = (periods[0].period_month or '').strip()
        month_keys = self._report_month_keys(periods[0])
        self.period_month = period_month
        self.ton_kho_month = self._prev_month_label(period_month)
        self.column_spec_json = json.dumps(
            [{'month_key': mk, 'label': mk} for mk in month_keys],
            ensure_ascii=False,
        )
        self.line_ids.unlink()

        buckets = {}
        for period in periods:
            self._merge_period_buckets(
                period, self.ton_kho_month, buckets,
            )
        GhiChu = self.env['bao.cao.ghi.chu'].sudo()
        period_key = GhiChu.period_key_from_periods(periods)
        ghi_chu_map = GhiChu.load_map(REPORT_KH_DSX, period_key)
        line_vals = self._buckets_to_line_vals(buckets, month_keys, ghi_chu_map)

        if line_vals:
            self.env['bao.cao.kh.dat.sx.line'].create(line_vals)

    def action_open_report(self):
        self.ensure_one()
        self._populate_lines()
        if not self.line_ids:
            raise UserError(_(
                'Không có dữ liệu kế hoạch kinh doanh / sản xuất cho các kỳ đã chọn.',
            ))
        title = _('Tổng hợp kế hoạch sản xuất tháng %s') % (self.period_month or '')
        return {
            'type': 'ir.actions.act_window',
            'name': title,
            'res_model': 'bao.cao.kh.dat.sx.line',
            'view_mode': 'tree',
            'domain': [('wizard_id', '=', self.id)],
            'context': {
                'bao_cao_khdsx_wizard_id': self.id,
                'bao_cao_khdsx_columns': self.column_spec_json or '[]',
                'bao_cao_khdsx_ton_month': self.ton_kho_month or '',
                'default_wizard_id': self.id,
            },
        }

    def action_export_excel(self):
        self.ensure_one()
        if not self.line_ids:
            self._populate_lines()
        lines = self.line_ids.sorted(
            key=lambda l: (
                l.company_sx_code or '',
                l.company_dat_code or '',
                l.nganh_hang or '',
            )
        )
        if not lines:
            raise UserError(_('Không có dữ liệu để xuất Excel.'))

        try:
            column_spec = json.loads(self.column_spec_json or '[]')
        except (TypeError, ValueError):
            raise UserError(_('Không xác định được cột tháng hiển thị.'))
        if not column_spec:
            raise UserError(_('Không xác định được cột tháng hiển thị.'))

        wb = Workbook()
        ws = wb.active
        ws.title = 'KH DAT SX'

        ton_label = 'Tồn kho cuối kỳ T%s' % (self.ton_kho_month or '')
        header_row1 = 1
        header_row2 = 2
        data_row = 3
        col = 1

        fixed_headers = [
            'Công ty sản xuất',
            'Công ty đặt hàng',
            'Ngành hàng',
            ton_label,
        ]
        for label in fixed_headers:
            ws.merge_cells(
                start_row=header_row1, start_column=col,
                end_row=header_row2, end_column=col,
            )
            cell = ws.cell(row=header_row1, column=col, value=label)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            col += 1

        for _metric_key, group_label in KH_DSX_METRIC_GROUPS:
            group_start = col
            ws.merge_cells(
                start_row=header_row1, start_column=group_start,
                end_row=header_row1, end_column=group_start + len(column_spec) - 1,
            )
            cell = ws.cell(row=header_row1, column=group_start, value=group_label)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            for col_def in column_spec:
                month_key = col_def.get('label') or col_def.get('month_key') or ''
                sub = ws.cell(row=header_row2, column=col, value='T%s' % month_key)
                sub.font = Font(bold=True)
                sub.alignment = Alignment(horizontal='center')
                col += 1

        row_idx = data_row
        for line in lines:
            metrics = line._parse_metrics()
            ws.cell(row=row_idx, column=1, value=line.company_sx_code or '')
            ws.cell(row=row_idx, column=2, value=line.company_dat_code or '')
            ws.cell(row=row_idx, column=3, value=line.nganh_hang or '')
            ws.cell(row=row_idx, column=4, value=line.ton_dau_ky or 0.0)
            col = 5
            for _metric_key, _label in KH_DSX_METRIC_GROUPS:
                for cell_data in metrics:
                    if _metric_key == 'ty_le':
                        ws.cell(row=row_idx, column=col, value=cell_data.get('ty_le') or 0.0)
                    else:
                        ws.cell(row=row_idx, column=col, value=cell_data.get(_metric_key) or 0.0)
                    col += 1
            row_idx += 1

        for col_idx in range(1, col):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14

        output = io.BytesIO()
        wb.save(output)
        fname = 'KH_DAT_SX_%s.xlsx' % (
            (self.period_month or 'report').replace('/', ''),
        )
        attachment = self.env['ir.attachment'].sudo().create({
            'name': fname,
            'type': 'binary',
            'datas': base64.b64encode(output.getvalue()),
            'mimetype': (
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            ),
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }


class BaoCaoKhDatSxLine(models.TransientModel):
    _name = 'bao.cao.kh.dat.sx.line'
    _inherit = ['bao.cao.ghi.chu.line.mixin']
    _description = 'Dòng Biểu 5 — Tổng hợp KH đặt sản xuất'
    _order = 'company_sx_code, company_dat_code, nganh_hang, id'

    wizard_id = fields.Many2one(
        'bao.cao.kh.dat.sx.wizard', ondelete='cascade', index=True)
    period_id = fields.Many2one('ke.hoach.vat.tu', string='Kỳ', index=True)

    company_sx_id = fields.Many2one('res.company', string='ĐV sản xuất')
    company_dat_id = fields.Many2one('res.company', string='ĐV đặt hàng')
    company_sx_code = fields.Char(
        string='ĐV SX', compute='_compute_company_codes', store=True)
    company_dat_code = fields.Char(
        string='ĐV đặt', compute='_compute_company_codes', store=True)

    nganh_hang = fields.Char(string='Ngành hàng', index=True)
    ton_dau_ky = fields.Float(string='Tồn đầu kỳ', digits=(16, 2))
    metrics_json = fields.Text(string='Metrics JSON')
    ghi_chu = fields.Text(string='Ghi chú')

    ton_kho_month = fields.Char(
        related='wizard_id.ton_kho_month', string='Tháng tồn', readonly=True)
    period_month = fields.Char(
        related='wizard_id.period_month', string='Tháng KH', readonly=True)

    def _sync_ghi_chu_to_master(self):
        GhiChu = self._ghi_chu_master()
        for rec in self:
            wizard = rec.wizard_id
            if not wizard or not rec.company_sx_id:
                continue
            period_key = rec._ghi_chu_period_key(wizard)
            scope = GhiChu.scope_key_khdsx(
                rec.company_sx_id.id,
                rec.company_dat_id.id if rec.company_dat_id else 0,
                rec.nganh_hang,
            )
            GhiChu.upsert_note(REPORT_KH_DSX, period_key, scope, rec.ghi_chu)

    @api.depends('company_sx_id', 'company_dat_id')
    def _compute_company_codes(self):
        for rec in self:
            sx = rec.company_sx_id
            dat = rec.company_dat_id
            rec.company_sx_code = (
                (sx.company_code or sx.name or '').strip() if sx else ''
            )
            rec.company_dat_code = (
                (dat.company_code or dat.name or '').strip() if dat else ''
            )

    def _parse_metrics(self):
        self.ensure_one()
        try:
            parsed = json.loads(self.metrics_json or '[]')
        except (TypeError, ValueError):
            return []
        return parsed if isinstance(parsed, list) else []

    def action_export_excel(self):
        if self:
            wizards = self.mapped('wizard_id')
        else:
            wizard_id = self.env.context.get('bao_cao_khdsx_wizard_id')
            wizards = (
                self.env['bao.cao.kh.dat.sx.wizard'].browse(wizard_id)
                if wizard_id else self.env['bao.cao.kh.dat.sx.wizard']
            )
        if len(wizards) != 1:
            raise UserError(_('Vui lòng xuất Excel từ một báo cáo đã mở.'))
        return wizards.action_export_excel()
