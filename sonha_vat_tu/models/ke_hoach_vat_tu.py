# -*- coding: utf-8 -*-
from collections import defaultdict
import calendar
from datetime import date
import os as _os
import re

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.numbers import FORMAT_TEXT
from markupsafe import Markup, escape

from odoo import api, fields, models, _
from odoo.exceptions import AccessError, UserError, ValidationError
from odoo.tools.safe_eval import safe_eval


_SQL_FUNCTIONS_PATH = _os.path.join(
    _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))),
    'data', 'sql', 'fn_ke_hoach_vat_tu.sql',
)


class KeHoachVatTu(models.Model):
    _name = 'ke.hoach.vat.tu'
    _description = 'kế hoạch vật tư cần'
    _rec_name = 'code'
    _order = 'period_month desc, id desc'
    _inherit = ['mail.thread', 'mail.activity.mixin', 'vat.tu.excel.mixin']

    code = fields.Char(string='Số chứng từ', readonly=True, copy=False, index=True, tracking=True)
    company_id = fields.Many2one(
        'res.company', string='Đơn vị lập kế hoạch',
        index=True, readonly=True, copy=False,
        default=lambda self: self.env.company.id,
        help='Đơn vị của user tạo kỳ; chỉ dùng phân quyền, không hiển thị trên form.',
    )
    company_sx_id = fields.Many2one(
        'res.company', string='Đơn vị sản xuất',
        copy=False, index=True, readonly=True, tracking=True,
        default=lambda self: self.env.company.id,
        help='Nhà máy sản xuất — lấy theo công ty user đang thao tác.',
    )
    period_month = fields.Char(
        string='Tháng bắt đầu', required=True, tracking=True)
    state = fields.Selection([
        ('ke_hoach', 'Kế hoạch sản xuất'),
        ('dinh_muc', 'Định mức kỳ'),
        ('tinh_toan', 'Tính toán vật tư'),
        ('tong_hop', 'Tổng hợp vật tư cần sản xuất'),
        ('dat_hang', 'Kế hoạch đặt vật tư'),
        ('bcu_tong_hop', 'Tổng hợp KH vật tư BCU'),
        ('phe_duyet', 'Phê duyệt kế hoạch vật tư'),
    ], default='ke_hoach', tracking=True, string='Trạng thái')
    note = fields.Text(string='Ghi chú')

    approval_company_id = fields.Many2one(
        'res.company',
        string='Đơn vị phê duyệt',
        compute='_compute_approval_company',
    )
    approval_flow_id = fields.Many2one(
        'luong.duyet',
        string='Luồng duyệt',
        copy=False,
        tracking=True,
    )
    approval_step_ids = fields.One2many(
        'buoc.duyet.ke.hoach.vat.tu',
        'period_id',
        string='Các bước duyệt',
        copy=False,
    )
    approval_state = fields.Selection(
        [
            ('draft', 'Chưa gửi duyệt'),
            ('approved', 'Đã duyệt'),
        ],
        string='Trạng thái phê duyệt',
        default='draft',
        required=True,
        copy=False,
        tracking=True,
    )
    approval_current_sequence = fields.Integer(
        string='Bước duyệt hiện tại',
        default=1,
        copy=False,
    )
    workflow_form_view_id = fields.Many2one(
        'ir.ui.view',
        string='Form view workflow',
        copy=False,
        index=True,
    )
    can_approve = fields.Boolean(compute='_compute_can_approve')

    ngay_du_phong_b4 = fields.Float(
        string='Số ngày dự phòng (B4)',
        default=15.0,
        digits=(16, 2),
        help='Dự phòng B4 = VT cần dùng tháng đầu ÷ 28 × số ngày này (mặc định 15 ≈ 2 tuần).',
        tracking=True,
    )
    ngay_du_tru_b5 = fields.Float(
        string='Số ngày dự trữ (B5)',
        default=20.0,
        digits=(16, 2),
        help='Dự trữ tối thiểu B5 = VT cần dùng tháng đầu ÷ 28 × số ngày này (mặc định 20).',
        tracking=True,
    )
    co_ke_hoach_vat_tu = fields.Boolean(
        string='Đã tạo kế hoạch vật tư',
        default=False,
        copy=False,
        readonly=True,
        help='Bật khi nhấn Tạo kế hoạch vật tư từ màn sản xuất; menu Kế hoạch vật tư chỉ hiện kỳ có cờ này.',
    )

    ke_hoach_san_xuat_ids = fields.One2many('ke.hoach.san.xuat', 'period_id', string='Kế hoạch sản xuất')
    ke_hoach_vat_tu_line_ids = fields.One2many('ke.hoach.vat.tu.line', 'period_id', string='Kế hoạch vật tư')
    dinh_muc_ids = fields.One2many('dinh.muc', 'period_id', string='Định mức tháng')
    tinh_toan_vat_tu_ids = fields.One2many('tinh.toan.vat.tu', 'period_id', string='Tính toán vật tư')
    tong_hop_vat_tu_ids = fields.One2many(
        'tong.hop.vat.tu', 'period_id', string='Tổng hợp vật tư',
        domain=[('don_vi_kd_id', '=', False)],
    )
    kh_dat_vat_tu_ids = fields.One2many('kh.dat.vat.tu', 'period_id', string='Kế hoạch đặt vật tư')
    kh_dat_vat_tu_bcu_ids = fields.One2many(
        'kh.dat.vat.tu.bcu', 'period_id', string='Tổng hợp KH vật tư BCU')
    phe_duyet_kh_vat_tu_ids = fields.One2many(
        'phe.duyet.kh.vat.tu', 'period_id', string='Phê duyệt kế hoạch vật tư')

    ke_hoach_san_xuat_count = fields.Integer(compute='_compute_counts')
    ke_hoach_vat_tu_line_count = fields.Integer(compute='_compute_counts')
    dinh_muc_count = fields.Integer(compute='_compute_counts')
    tinh_toan_vat_tu_count = fields.Integer(compute='_compute_counts')
    tong_hop_vat_tu_count = fields.Integer(compute='_compute_counts')
    kh_dat_vat_tu_count = fields.Integer(compute='_compute_counts')
    kh_dat_vat_tu_bcu_count = fields.Integer(compute='_compute_counts')
    phe_duyet_kh_vat_tu_count = fields.Integer(compute='_compute_counts')

    @api.depends(
        'company_sx_id',
        'kh_dat_vat_tu_ids.company_id',
    )
    def _compute_approval_company(self):
        for rec in self:
            if rec.company_sx_id:
                rec.approval_company_id = rec.company_sx_id
            else:
                companies = rec.kh_dat_vat_tu_ids.mapped('company_id')
                rec.approval_company_id = companies[:1] if len(companies) == 1 else False

    @api.depends(
        'state',
        'approval_state',
        'approval_current_sequence',
        'approval_step_ids.sequence',
        'approval_step_ids.nguoi_duyet_id',
        'approval_step_ids.da_duyet',
    )
    def _compute_can_approve(self):
        current_user = self.env.user
        for rec in self:
            if rec.state != 'phe_duyet' or rec.approval_state == 'approved':
                rec.can_approve = False
                continue
            current_steps = rec.approval_step_ids.filtered(
                lambda step: step.sequence == rec.approval_current_sequence
            )
            rec.can_approve = any(
                step.nguoi_duyet_id == current_user and not step.da_duyet
                for step in current_steps
            )

    def _approval_step_commands_from_flow(self, flow):
        """Sinh lệnh One2many cho các bước duyệt từ luồng cấu hình."""
        if not flow:
            return [(5, 0, 0)], 1

        lines = [(5, 0, 0)]
        configured_steps = flow.sudo().step_ids.sorted(
            key=lambda step: (step.sequence, step.id)
        )
        for step in configured_steps:
            if step.phuong_thuc == 'ql':
                lines.append((0, 0, {
                    'sequence': step.sequence,
                    'phuong_thuc': step.phuong_thuc,
                    'vai_tro_id': step.vai_tro.id or False,
                    'nguoi_duyet_id': self.env.user.id,
                }))
            else:
                for user in step.ten_nguoi_duyet:
                    lines.append((0, 0, {
                        'sequence': step.sequence,
                        'phuong_thuc': step.phuong_thuc,
                        'vai_tro_id': step.vai_tro.id or False,
                        'nguoi_duyet_id': user.id,
                    }))
        first_sequence = min(configured_steps.mapped('sequence')) if configured_steps else 1
        return lines, first_sequence

    def _sync_approval_steps_from_flow(self):
        """Ghi bước duyệt từ luồng (sudo) — tránh mất line khi field readonly trên form."""
        for rec in self:
            if rec.approval_state == 'approved':
                continue
            lines, first_sequence = rec._approval_step_commands_from_flow(rec.approval_flow_id)
            rec.sudo().write({
                'approval_step_ids': lines,
                'approval_current_sequence': first_sequence,
            })

    @api.onchange('approval_flow_id')
    def _onchange_approval_flow_id(self):
        """Giống MDM: chọn luồng duyệt → sinh luôn các bước trên tab Phê duyệt."""
        lines, first_sequence = self._approval_step_commands_from_flow(self.approval_flow_id)
        self.approval_step_ids = lines
        self.approval_current_sequence = first_sequence

    def write(self, vals):
        flow_changed = 'approval_flow_id' in vals
        if flow_changed:
            vals = dict(vals)
            vals.pop('approval_step_ids', None)
        res = super().write(vals)
        if flow_changed:
            self.filtered(
                lambda rec: rec.approval_state != 'approved'
            )._sync_approval_steps_from_flow()
        return res

    _sql_constraints = [
        ('code_uniq', 'unique(code)', 'Mã kỳ phải duy nhất!'),
    ]

    @api.model
    def init(self):
        with open(_SQL_FUNCTIONS_PATH, 'r', encoding='utf-8') as f:
            self.env.cr.execute(f.read())

    @api.model
    def _month_key_to_date(self, month_key):
        if not month_key:
            return False
        try:
            month, year = str(month_key).split('/')
            return date(int(year), int(month), 1)
        except Exception:
            return False

    @api.model
    def _parse_month_key(self, month_key):
        """'MM/YYYY' -> (month, year) hoặc raise ValueError."""
        parts = (month_key or '').strip().split('/')
        if len(parts) != 2:
            raise ValueError(month_key)
        month, year = int(parts[0]), int(parts[1])
        if month < 1 or month > 12:
            raise ValueError(month_key)
        return month, year

    @api.model
    def _format_month_key(self, month, year):
        return f'{month:02d}/{year}'

    @api.model
    def month_start_from_key(self, month_key):
        month, year = self._parse_month_key(month_key)
        return date(year, month, 1)

    @api.model
    def month_end_from_key(self, month_key):
        month, year = self._parse_month_key(month_key)
        last = calendar.monthrange(year, month)[1]
        return date(year, month, last)

    @api.model
    def iter_calendar_months(self, date_from, date_to):
        """Sinh danh sách 'MM/YYYY' từ date_from → date_to (bao gồm 2 đầu)."""
        cur = date(date_from.year, date_from.month, 1)
        end = date(date_to.year, date_to.month, 1)
        if cur > end:
            return []
        out = []
        while cur <= end:
            out.append(self._format_month_key(cur.month, cur.year))
            if cur.month == 12:
                cur = date(cur.year + 1, 1, 1)
            else:
                cur = date(cur.year, cur.month + 1, 1)
        return out

    @api.model
    def validate_report_month_range(self, date_from, date_to):
        if date_from > date_to:
            raise UserError(_('Từ tháng không được lớn hơn Đến tháng.'))

    def _get_current_production_company(self):
        return self.env.company

    @api.model
    def _get_company_code(self, company=None):
        company = company or self.env.company
        code = (getattr(company, 'company_code', None) or '').strip()
        return code or (company.name or 'XX').strip()

    @api.model
    def _get_creator_company_code(self):
        return self._get_company_code()

    @api.model
    def _period_code_prefix(self, period_month, company_code=None):
        company_code = (company_code or self._get_creator_company_code()).strip()
        month, year = (period_month or '').strip().split('/')
        return 'KHVT_%s_%s%s' % (company_code, month, year)

    @api.model
    def _next_period_code(self, period_month, company_code=None):
        prefix = self._period_code_prefix(period_month, company_code) + '_'
        latest = self.sudo().search([('code', '=like', prefix + '%')], order='code desc', limit=1)
        next_no = 1
        if latest.code:
            try:
                next_no = int(latest.code.rsplit('_', 1)[-1]) + 1
            except (TypeError, ValueError):
                next_no = 1
        return '%s%02d' % (prefix, next_no)

    @api.model
    def _period_code_sequence_suffix(self, code):
        if not code:
            return None
        try:
            return int(str(code).rsplit('_', 1)[-1])
        except (TypeError, ValueError):
            return None

    def _generate_period_code(self, period_month=None, company_sx=None, prefer_suffix=None):
        self.ensure_one()
        period_month = (period_month or self.period_month or '').strip()
        company_sx = company_sx or self.company_sx_id
        if not period_month or not company_sx:
            return False
        company_code = self._get_company_code(company_sx)
        prefix = self._period_code_prefix(period_month, company_code)
        if prefer_suffix is not None:
            candidate = '%s_%02d' % (prefix, prefer_suffix)
            domain = [('code', '=', candidate)]
            if self.id:
                domain.append(('id', '!=', self.id))
            if not self.search(domain, limit=1):
                return candidate
        return self._next_period_code(period_month, company_code)

    @api.onchange('period_month', 'company_sx_id')
    def _onchange_period_month_code(self):
        if self.co_ke_hoach_vat_tu or self.state != 'ke_hoach':
            return
        if not self.period_month or not self.company_sx_id:
            return
        pattern = re.compile(r'^(0[1-9]|1[0-2])/\d{4}$')
        if not pattern.match(self.period_month.strip()):
            return
        prefer = self._period_code_sequence_suffix(self.code)
        self.code = self._generate_period_code(prefer_suffix=prefer)

    @api.model
    def _get_view_cache_key(self, view_id=None, view_type='form', **options):
        key = super()._get_view_cache_key(view_id, view_type, **options)
        u = self.env.user
        return key + (
            options.get('action_id'),
            u.has_group('sonha_vat_tu.group_ban_cung_ung_vat_tu'),
            u.has_group('sonha_vat_tu.group_bo_phan_vat_tu'),
            u.has_group('sonha_vat_tu.group_truong_bo_phan_vat_tu'),
        )

    @api.model
    def _get_view(self, view_id=None, view_type='form', **options):
        arch, view = super()._get_view(view_id=view_id, view_type=view_type, **options)
        if view_type not in ('tree', 'form'):
            return arch, view

        user = self.env.user
        is_ban_cung_ung = user.has_group('sonha_vat_tu.group_ban_cung_ung_vat_tu')
        is_bo_phan = (
            user.has_group('sonha_vat_tu.group_bo_phan_vat_tu')
            or user.has_group('sonha_vat_tu.group_truong_bo_phan_vat_tu')
        )
        action_id = options.get('action_id')


        action_sx = self.env.ref('sonha_vat_tu.action_ke_hoach_san_xuat_period', raise_if_not_found=False)
        action_vt = self.env.ref('sonha_vat_tu.action_ke_hoach_vat_tu_period', raise_if_not_found=False)

        is_kd = False
        is_sx = action_sx and action_id == action_sx.id
        is_vt = action_vt and action_id == action_vt.id
        if not (is_sx or is_vt) and view_type == 'tree':
            form_ref = self.env.context.get('form_view_ref') or ''
            is_sx = 'view_ke_hoach_vat_tu_form_sx' in form_ref
            is_vt = 'view_ke_hoach_vat_tu_form_vt' in form_ref

        lock_create = False
        if is_ban_cung_ung:
            lock_create = True
        elif is_vt:
            lock_create = True

        # Chỉ khóa nút New trên list kỳ — không đụng tree one2many trong form (B5 thêm dòng tay).
        if lock_create and view_type == 'tree':
            for node in arch.xpath('//tree'):
                node.set('create', 'false')
        return arch, view

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if not vals.get('company_id'):
                vals['company_id'] = self.env.company.id
            if not vals.get('company_sx_id'):
                vals['company_sx_id'] = self.env.company.id
            if not vals.get('code') and vals.get('period_month'):
                company_sx = self.env['res.company'].browse(vals['company_sx_id'])
                vals['code'] = self._next_period_code(
                    vals['period_month'],
                    self._get_company_code(company_sx),
                )
        return super().create(vals_list)

    def write(self, vals):
        sx_locked_fields = {
            'period_month', 'company_sx_id', 'ke_hoach_san_xuat_ids',
        }
        if sx_locked_fields & set(vals.keys()):
            locked = self.filtered(lambda rec: rec.co_ke_hoach_vat_tu and rec.state == 'ke_hoach')
            if locked:
                raise UserError(_(
                    'Đã tạo kế hoạch vật tư, không thể sửa kế hoạch sản xuất.'
                ))
        refresh_code = (
            not self.env.context.get('skip_khvt_code_update')
            and ('period_month' in vals or 'company_sx_id' in vals)
        )
        res = super().write(vals)
        if not self.env.context.get('skip_khvt_code_update'):
            for rec in self.filtered(
                lambda r: r.state == 'ke_hoach'
                and not r.co_ke_hoach_vat_tu
                and r.period_month
                and r.company_sx_id
            ):
                if not rec.code or refresh_code:
                    new_code = rec._generate_period_code(
                        prefer_suffix=rec._period_code_sequence_suffix(rec.code),
                    )
                    if new_code and new_code != rec.code:
                        rec.with_context(skip_khvt_code_update=True).write({
                            'code': new_code,
                        })
        return res

    # B1–B5: được xóa; từ B6 (bcu_tong_hop) trở đi: không xóa.
    _PERIOD_DELETABLE_STATES = frozenset({
        'ke_hoach', 'dinh_muc', 'tinh_toan', 'tong_hop', 'dat_hang',
    })

    def unlink(self):
        locked = self.filtered(
            lambda rec: rec.state not in self._PERIOD_DELETABLE_STATES
            or rec.approval_state == 'approved'
        )
        if locked:
            raise UserError(_(
                'Không thể xóa kỳ đã tới bước 6 (Tổng hợp KH BCU) trở đi hoặc đã phê duyệt.'
            ))
        kd_headers = self.env['ke.hoach.kinh.doanh'].sudo().search([
            ('period_sx_id', 'in', self.ids),
        ])
        if kd_headers:
            kd_headers.write({'locked': False, 'period_sx_id': False})
        return super().unlink()

    @api.constrains('period_month')
    def _check_period_month(self):
        pattern = re.compile(r'^(0[1-9]|1[0-2])/\d{4}$')
        for rec in self:
            if rec.period_month and not pattern.match(rec.period_month):
                raise ValidationError('Tháng bắt đầu phải đúng định dạng MM/YYYY, ví dụ 04/2026.')

    @api.constrains('ngay_du_phong_b4', 'ngay_du_tru_b5')
    def _check_tham_so_ngay(self):
        for rec in self:
            if rec.ngay_du_phong_b4 is not False and rec.ngay_du_phong_b4 <= 0:
                raise ValidationError(
                    _('Số ngày B4 phải lớn hơn 0 (đang là %s).') % rec.ngay_du_phong_b4)
            if rec.ngay_du_tru_b5 is not False and rec.ngay_du_tru_b5 <= 0:
                raise ValidationError(
                    _('Số ngày B5 phải lớn hơn 0 (đang là %s).') % rec.ngay_du_tru_b5)

    @api.depends(
        'state',
        'ke_hoach_san_xuat_ids', 'ke_hoach_vat_tu_line_ids',
        'dinh_muc_ids', 'tinh_toan_vat_tu_ids',
        'tong_hop_vat_tu_ids', 'kh_dat_vat_tu_ids',
        'kh_dat_vat_tu_bcu_ids', 'phe_duyet_kh_vat_tu_ids',
    )
    def _compute_counts(self):
        """Đếm bằng len() trên chính One2many đã prefetch, thay vì 8 search_count
        cho mỗi kỳ."""
        for rec in self:
            rec.ke_hoach_san_xuat_count = len(rec.ke_hoach_san_xuat_ids)
            rec.ke_hoach_vat_tu_line_count = len(rec.ke_hoach_vat_tu_line_ids)
            rec.dinh_muc_count = len(rec.dinh_muc_ids)
            rec.tinh_toan_vat_tu_count = len(rec.tinh_toan_vat_tu_ids)
            rec.tong_hop_vat_tu_count = len(rec.tong_hop_vat_tu_ids)
            rec.kh_dat_vat_tu_count = len(rec.kh_dat_vat_tu_ids)
            rec.kh_dat_vat_tu_bcu_count = len(rec.kh_dat_vat_tu_bcu_ids)
            rec.phe_duyet_kh_vat_tu_count = len(rec.phe_duyet_kh_vat_tu_ids)

    def _get_horizon_months(self):
        self.ensure_one()
        if not self.period_month:
            return []
        try:
            m, y = map(int, self.period_month.split('/'))
            res = []
            for i in range(4):
                tm = m + i
                ty = y
                while tm > 12:
                    tm -= 12
                    ty += 1
                res.append(f"{tm:02d}/{ty}")
            return res
        except Exception:
            return []

    def _vat_tu_di_duong_template_rows(self):
        """Sinh dòng template import vật tư đi đường từ B3: 1 dòng / mã NVL (gom theo ĐV SX)."""
        self.ensure_one()
        months = self._get_horizon_months()
        if len(months) < 4:
            months = (months + [''] * 4)[:4]
        month_keys = months[:4]

        if not self.company_sx_id:
            return []

        by_ma = {}
        for line in self.tinh_toan_vat_tu_ids:
            ma_nvl = (line.ma_vat_tu or '').strip()
            if not ma_nvl:
                continue
            if ma_nvl not in by_ma:
                by_ma[ma_nvl] = line.ten_vat_tu or ''

        if not by_ma:
            return []

        ma_nvls = sorted(by_ma)
        existing = {}
        for rec in self.env['vat.tu.di.duong'].sudo().search([
            ('company_id', '=', self.company_sx_id.id),
            ('loai', '=', 'don_vi'),
            ('ma_nvl', 'in', ma_nvls),
            ('month_key', 'in', month_keys),
        ]):
            existing[(rec.ma_nvl, rec.month_key)] = rec.so_luong or 0.0

        return [
            {
                'ma_nvl': ma_nvl,
                'ten_nvl': by_ma[ma_nvl],
                'qtys': [existing.get((ma_nvl, mk), 0.0) for mk in month_keys],
            }
            for ma_nvl in ma_nvls
        ]

    # ------------------------------------------------------------------
    # Actions — gọi thẳng SQL Procedure
    # ------------------------------------------------------------------

    def action_generate_b2(self):
        self.ensure_one()
        if not self.env['ke.hoach.vat.tu.line'].search_count([('period_id', '=', self.id)]):
            raise UserError(_(
                'Chưa có kế hoạch vật tư. Vui lòng tạo kế hoạch vật tư từ kế hoạch sản xuất trước khi sinh định mức.'
            ))

        self.env.cr.execute('CALL public.fn_sinh_dinh_muc(%s)', (self.id,))
        self.env.invalidate_all()
        self.write({'state': 'dinh_muc'})
        self.invalidate_recordset([
            'dinh_muc_ids', 'dinh_muc_count', 'state',
        ])
        return self.action_open_step_b2()

    def _production_company_for_auto_seed(self):
        self.ensure_one()
        production_companies = self.ke_hoach_san_xuat_ids.mapped('company_sx_id').filtered(lambda c: c)
        if len(production_companies) == 1:
            return production_companies
        if self.env.company.company_code in ('BNH', 'SSP'):
            return self.env.company
        return self.env['res.company'].browse()

    def action_lay_ke_hoach_kinh_doanh(self):
        self.ensure_one()
        if self.state != 'ke_hoach':
            raise UserError(_('Kỳ kế hoạch đã sang bước sau, không thể lấy kế hoạch kinh doanh.'))
        if not self._has_plan_edit_rights():
            raise UserError(_('Bạn không có quyền lấy kế hoạch kinh doanh.'))
        if self.co_ke_hoach_vat_tu:
            raise UserError(_(
                'Đã tạo kế hoạch vật tư, không thể lấy lại kế hoạch kinh doanh.'
            ))
        if not self.period_month:
            raise UserError(_('Tháng bắt đầu không được để trống.'))
        if not self.company_sx_id:
            raise UserError(_('Đơn vị sản xuất không được để trống.'))
        self._pull_kinh_doanh_into_san_xuat()
        self.invalidate_recordset(['ke_hoach_san_xuat_ids', 'ke_hoach_san_xuat_count'])
        return True

    def _kinh_doanh_headers_for_pull(self):
        """KHKD của user hiện tại: cùng tháng + ĐV SX, chưa gắn kỳ khác."""
        self.ensure_one()
        base_domain = [
            ('create_uid', '=', self.env.user.id),
            ('period_month', '=', self.period_month),
            ('company_sx_id', '=', self.company_sx_id.id),
        ]
        # KHKD đã gom vào kỳ SX bị xóa: period_sx_id=null nhưng locked vẫn True → mở khóa lại.
        orphans = self.env['ke.hoach.kinh.doanh'].sudo().search(
            base_domain + [('locked', '=', True), ('period_sx_id', '=', False)]
        )
        if orphans:
            orphans.write({'locked': False})
        return self.env['ke.hoach.kinh.doanh'].search(
            base_domain + [
                '|', ('locked', '=', False), ('period_sx_id', '=', self.id),
            ],
            order='company_id, id',
        )

    def _pull_kinh_doanh_into_san_xuat(self):
        self.ensure_one()
        Production = self.env['ke.hoach.san.xuat'].sudo()
        sync_ctx = {
            'is_importing': True,
            'allow_unassigned_production_company': True,
            'skip_kd_sx_sync': True,
            'tracking_disable': True,
        }

        headers = self._kinh_doanh_headers_for_pull().sudo()
        if not headers:
            raise UserError(_(
                'Không có kế hoạch kinh doanh nào của bạn cho tháng %s và đơn vị SX %s.'
            ) % (
                self.period_month,
                self.company_sx_id.company_code or self.company_sx_id.name,
            ))

        pull_stats = []
        vals_list = []
        seq = 10
        for kd in headers:
            ky_count = 0
            for line in kd.line_ids.sudo().sorted(key=lambda l: (l.sequence, l.id)):
                if not line.ma_sap:
                    continue
                ky_count += 1
                vals_list.append({
                    'period_id': self.id,
                    'company_id': line.company_id.id,
                    'company_sx_id': self.company_sx_id.id,
                    'ma_hang': line.ma_hang,
                    'ma_sap': line.ma_sap,
                    'note': line.note,
                    'sequence': seq,
                    'qty_t0': line.qty_t0 or 0.0,
                    'qty_t1': line.qty_t1 or 0.0,
                    'qty_t2': line.qty_t2 or 0.0,
                    'qty_t3': line.qty_t3 or 0.0,
                })
                seq += 10
            if ky_count:
                pull_stats.append((kd.code, ky_count))

        if not vals_list:
            raise UserError(_(
                'Không có dòng kế hoạch kinh doanh. Kiểm tra các KHKD đã import chưa.'
            ))

        existing_sx = self.ke_hoach_san_xuat_ids
        if existing_sx:
            existing_sx.with_context(**sync_ctx).unlink()

        Production.with_context(**sync_ctx).create(vals_list)

        newly_used = headers.filtered(lambda h: not h.locked)
        if newly_used:
            newly_used.write({'locked': True, 'period_sx_id': self.id})
        else:
            headers.write({'period_sx_id': self.id})

        items = ''.join(
            Markup(
                '<li>Đã lấy %d dòng kế hoạch kinh doanh từ %s</li>'
            ) % (count, escape(code))
            for code, count in pull_stats
        )
        self.with_context(vat_tu_chatter_scope='sx').message_post(
            body=Markup('<ul>%s</ul>') % Markup(items),
        )

    def _kinh_doanh_qty_map(self):
        """Gom số lượng KD theo (ĐV đặt hàng, Mã) từ các KHKD đã lấy vào kỳ."""
        self.ensure_one()
        agg = {}
        kd_lines = self.env['ke.hoach.kinh.doanh.line'].sudo().search([
            ('kinh_doanh_id.period_sx_id', '=', self.id),
            ('ma_sap', '!=', False),
        ])
        for line in kd_lines:
            key = (line.company_id.id, line.ma_sap)
            if key not in agg:
                agg[key] = {
                    'qty_t0': 0.0, 'qty_t1': 0.0, 'qty_t2': 0.0, 'qty_t3': 0.0,
                }
            for qty_field in ('qty_t0', 'qty_t1', 'qty_t2', 'qty_t3'):
                agg[key][qty_field] += getattr(line, qty_field) or 0.0
        return agg

    def _prepare_material_plan_values_from_production(self, production_company):
        self.ensure_one()
        kd_lines = self.env['ke.hoach.kinh.doanh.line'].sudo().search([
            ('kinh_doanh_id.period_sx_id', '=', self.id),
            ('ma_sap', '!=', False),
        ], order='kinh_doanh_id, sequence, id')
        sx_lines = self.ke_hoach_san_xuat_ids.sorted(key=lambda r: (r.sequence, r.id))

        kd_queues = defaultdict(list)
        for line in kd_lines:
            kd_queues[(line.company_id.id, line.ma_sap)].append(line)

        sap_codes = sorted({
            (line.ma_sap or '').strip() for line in sx_lines if (line.ma_sap or '').strip()
        })
        meta_map = self.env['ma.hang'].get_mdm_sap_meta_map(sap_codes) if sap_codes else {}
        NganhHang = self.env['mdm.nganh.hang'].sudo()
        nganh_names = {}
        nh_ids = {m['nganh_hang_id'] for m in meta_map.values() if m.get('nganh_hang_id')}
        if nh_ids:
            for nh in NganhHang.browse(list(nh_ids)):
                nganh_names[nh.id] = nh.ten or ''

        vals_list = []
        seq = 10
        for sx_line in sx_lines:
            key = (sx_line.company_id.id, sx_line.ma_sap)
            queue = kd_queues.get(key, [])
            kd_line = queue.pop(0) if queue else None
            ma_sap = sx_line.ma_sap
            meta = meta_map.get((ma_sap or '').strip(), {})
            nganh_hang = sx_line.nganh_hang.ten if sx_line.nganh_hang else ''
            if not nganh_hang and meta.get('nganh_hang_id'):
                nganh_hang = nganh_names.get(meta['nganh_hang_id'], '')
            vals_list.append({
                'period_id': self.id,
                'company_id': sx_line.company_id.id,
                'company_sx_id': production_company.id,
                'nganh_hang': nganh_hang,
                'ma_hang': sx_line.ma_hang,
                'ma_sap': ma_sap,
                'sequence': seq,
                'qty_kd_t0': (kd_line.qty_t0 or 0.0) if kd_line else 0.0,
                'qty_kd_t1': (kd_line.qty_t1 or 0.0) if kd_line else 0.0,
                'qty_kd_t2': (kd_line.qty_t2 or 0.0) if kd_line else 0.0,
                'qty_kd_t3': (kd_line.qty_t3 or 0.0) if kd_line else 0.0,
                'qty_sx_t0': sx_line.qty_t0,
                'qty_sx_t1': sx_line.qty_t1,
                'qty_sx_t2': sx_line.qty_t2,
                'qty_sx_t3': sx_line.qty_t3,
                'qty_t0': sx_line.qty_t0,
                'qty_t1': sx_line.qty_t1,
                'qty_t2': sx_line.qty_t2,
                'qty_t3': sx_line.qty_t3,
                'note': sx_line.note,
            })
            seq += 10
        return vals_list

    def action_create_material_plan(self):
        self.ensure_one()
        if self.state != 'ke_hoach':
            raise UserError(_('Kế hoạch đã sang bước sau, không thể tạo lại kế hoạch vật tư.'))
        if not self.ke_hoach_san_xuat_ids:
            raise UserError(_('Chưa có kế hoạch sản xuất để tạo kế hoạch vật tư.'))
        if self.ke_hoach_vat_tu_line_ids:
            raise UserError(_('Kế hoạch vật tư đã có dữ liệu. Vui lòng xóa dữ liệu cũ nếu cần tạo lại.'))

        production_company = self._get_current_production_company()
        unassigned_sx = self.ke_hoach_san_xuat_ids.filtered(lambda line: not line.company_sx_id)
        if unassigned_sx:
            unassigned_sx.with_context(
                is_importing=True,
                allow_unassigned_production_company=True,
            ).write({'company_sx_id': production_company.id})

        self.write({
            'company_sx_id': production_company.id,
            'co_ke_hoach_vat_tu': True,
        })

        vals_list = self._prepare_material_plan_values_from_production(production_company)
        Line = self.env['ke.hoach.vat.tu.line'].sudo()
        import_ctx = {'skip_period_lock': True, 'is_importing': True, 'tracking_disable': True}
        if vals_list:
            Line.with_context(**import_ctx).create(vals_list)
        count = len(vals_list)
        self.with_context(vat_tu_chatter_scope='vt').message_post(
            body=_('Đã tạo %s dòng kế hoạch vật tư.') % count
        )
        return self.action_open_workflow_vt()

    def action_compute_b3(self):
        self.ensure_one()
        self.env.cr.execute('CALL public.fn_tinh_toan_vat_tu(%s)', (self.id,))
        self.write({'state': 'tinh_toan'})
        self.invalidate_recordset([
            'tinh_toan_vat_tu_ids', 'tinh_toan_vat_tu_count', 'state',
        ])
        return self.action_open_step_b3()

    def action_open_import_vat_tu_di_duong_wizard(self):
        self.ensure_one()
        if self.state != 'tinh_toan':
            raise UserError(_(
                'Chỉ import vật tư đi đường khi đã ở bước Tính toán vật tư.'
            ))
        return {
            'name': _('Import vật tư đi đường'),
            'type': 'ir.actions.act_window',
            'res_model': 'import.vat.tu.di.duong.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'default_period_id': self.id,
                'default_loai': 'don_vi',
                'vat_tu_di_duong_loai': 'don_vi',
            },
        }

    def action_compute_b4(self):
        self.ensure_one()
        self.env.cr.execute(
            'CALL public.fn_tong_hop_vat_tu(%s, %s)',
            (self.id, self.ngay_du_phong_b4 or 15.0)
        )
        self.state = 'tong_hop'
        return self.action_open_step_b4()

    def action_open_import_bcu_wizard(self):
        self.ensure_one()
        if self.state not in ('bcu_tong_hop', 'phe_duyet'):
            raise UserError(_(
                'Chỉ import hàng đi đường BCU khi đã ở bước Tổng hợp KH vật tư BCU.'
            ))
        if not self.kh_dat_vat_tu_bcu_ids:
            raise UserError(_(
                'Chưa có dữ liệu Tổng hợp KH vật tư BCU. Vui lòng chạy bước này trước khi import.'
            ))
        view = self.env.ref('sonha_vat_tu.view_import_tong_hop_bcu_wizard_form')
        return {
            'name': _('Import hàng đi đường BCU'),
            'type': 'ir.actions.act_window',
            'res_model': 'import.tong.hop.bcu.wizard',
            'view_mode': 'form',
            'views': [(view.id, 'form')],
            'view_id': view.id,
            'target': 'new',
            'context': {
                'default_period_id': self.id,
                'period_id_readonly': True,
                'vat_tu_di_duong_loai': 'bcu',
            },
        }

    @api.model
    def _is_bcu_user(self):
        return self.env.user.has_group('sonha_vat_tu.group_ban_cung_ung_vat_tu')

    def _check_bcu_workflow_access(self):
        if not self._is_bcu_user():
            raise AccessError(_('Chỉ Ban cung ứng được thao tác từ bước Tổng hợp KH vật tư BCU trở đi.'))

    def action_submit_to_bcu(self):
        """BCU nhận kế hoạch từ B5 — sinh B6."""
        self.ensure_one()
        self._check_bcu_workflow_access()
        if not self.kh_dat_vat_tu_ids:
            raise UserError(_('Chưa có kế hoạch đặt vật tư (B5).'))
        self.env.cr.execute(
            'CALL public.fn_ke_hoach_dat_vat_tu_bcu(%s)',
            (self.id,),
        )
        b6_lines = self.env['kh.dat.vat.tu.bcu'].search([('period_id', '=', self.id)])
        if b6_lines:
            b6_lines._compute_sl_du_tru_toi_thieu_bcu()
            b6_lines._compute_sl_dat_mua_de_xuat()
            self.env['kh.dat.vat.tu.bcu']._apply_chot_from_bcu_di_duong(b6_lines)
            b6_lines._compute_b6_derived()
            b6_lines.flush_recordset([
                'sl_du_tru_toi_thieu_bcu',
                'sl_dat_mua_de_xuat', 'sl_dat_mua_chot', 'sl_can_mua_theo_moq',
                'sl_ton_kho_cuoi_ky', 'so_ngay_vong_quay_ton',
                'don_gia_ton_kho_cuoi_ky', 'gia_tri_ton_kho_cuoi_ky', 'gia_tri_mua_hang',
            ])
        self.state = 'bcu_tong_hop'
        self.invalidate_recordset([
            'kh_dat_vat_tu_bcu_ids', 'kh_dat_vat_tu_bcu_count', 'state',
        ])
        return self.action_open_step_b6()

    def action_compute_b7(self):
        """BCU chốt B6 → sinh màn phê duyệt B7."""
        self.ensure_one()
        self._check_bcu_workflow_access()
        if not self.kh_dat_vat_tu_bcu_ids:
            raise UserError(_('Chưa có dữ liệu Tổng hợp KH vật tư BCU (B6).'))
        self.env.cr.execute(
            'CALL public.fn_phe_duyet_kh_vat_tu(%s)',
            (self.id,),
        )
        b7_lines = self.env['phe.duyet.kh.vat.tu'].search([
            ('period_id', '=', self.id),
        ])
        b7_lines._apply_leadtime_from_config()
        self.state = 'phe_duyet'
        self.invalidate_recordset([
            'phe_duyet_kh_vat_tu_ids', 'phe_duyet_kh_vat_tu_count', 'state',
        ])
        return self.action_open_step_b7()

    def action_compute_b5(self):
        self.ensure_one()
        self.env.cr.execute(
            'CALL public.fn_ke_hoach_dat_vat_tu(%s, %s)',
            (self.id, self.ngay_du_tru_b5 or 20.0)
        )
        self.state = 'dat_hang'
        return self.action_open_step_b5()

    def action_approve_material_plan(self):
        self.ensure_one()
        if self.state != 'phe_duyet' or not self.phe_duyet_kh_vat_tu_ids:
            raise UserError(_('Chỉ có thể duyệt sau khi đã sinh phê duyệt kế hoạch vật tư.'))
        if self.approval_state == 'approved':
            raise UserError(_('Phê duyệt kế hoạch vật tư đã hoàn tất.'))
        if not self.approval_flow_id:
            raise UserError(_('Vui lòng chọn Luồng duyệt trước khi duyệt.'))
        if not self.approval_step_ids:
            raise UserError(_('Chưa có bước duyệt. Vui lòng chọn lại Luồng duyệt.'))

        current_steps = self.approval_step_ids.filtered(
            lambda step: step.sequence == self.approval_current_sequence
        )
        my_steps = current_steps.filtered(
            lambda step: step.nguoi_duyet_id == self.env.user and not step.da_duyet
        )
        if not my_steps:
            raise UserError(_('Chưa đến lượt bạn duyệt kế hoạch này.'))

        my_steps.sudo().write({
            'da_duyet': True,
            'ngay_duyet': self.env.cr.now(),
        })
        self.message_post(body=_(
            '%s đã duyệt bước %s.'
        ) % (self.env.user.name, self.approval_current_sequence))

        if all(current_steps.mapped('da_duyet')):
            next_sequences = self.approval_step_ids.filtered(
                lambda step: step.sequence > self.approval_current_sequence
            ).mapped('sequence')
            if next_sequences:
                self.approval_current_sequence = min(next_sequences)
            else:
                self.write({
                    'approval_state': 'approved',
                    'approval_current_sequence': 0,
                })
                self.message_post(body=_('Phê duyệt kế hoạch vật tư đã được phê duyệt hoàn tất.'))
        return self.action_open_step_b7()

    def _apply_plan_excel_style(self, ws, header_row, max_col):
        base_font = Font(name='Times New Roman', size=10)
        label_font = Font(name='Times New Roman', size=10, bold=True)
        value_font = Font(name='Times New Roman', size=10)
        header_font = Font(name='Times New Roman', size=10, bold=True, color='FFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='3F6F8F')
        thin_side = Side(style='thin', color='B7C6D0')
        header_side = Side(style='thin', color='2F556D')
        meta_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        header_border = Border(left=header_side, right=header_side, top=header_side, bottom=header_side)
        header_alignment = Alignment(horizontal='center', vertical='center')
        body_alignment = Alignment(vertical='center')

        for row in ws.iter_rows(min_row=1, max_row=max(ws.max_row, header_row), min_col=1, max_col=max_col):
            for cell in row:
                cell.font = base_font
                cell.alignment = body_alignment

        for row_idx in (1, 2, 3):
            label_cell = ws.cell(row=row_idx, column=1)
            value_cell = ws.cell(row=row_idx, column=2)
            label_cell.font = label_font
            value_cell.font = value_font
            label_cell.border = meta_border
            label_cell.alignment = Alignment(horizontal='left', vertical='center')
            value_cell.alignment = Alignment(horizontal='left', vertical='center')
            for col_idx in range(2, max_col + 1):
                meta_cell = ws.cell(row=row_idx, column=col_idx)
                meta_cell.border = meta_border
                meta_cell.font = value_font
                meta_cell.alignment = Alignment(horizontal='left', vertical='center')

        for cell in ws[header_row][:max_col]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = header_alignment

        ws.row_dimensions[header_row].height = 22
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

    @api.model
    def _company_display_code(self, company):
        if not company:
            return ''
        company = company.sudo()
        return company.company_code or company.name or ''

    _PLAN_EXPORT_HEADERS = ['Đơn vị đặt hàng', 'Ngành hàng', 'Tên hàng', 'Mã hàng', 'Mã']

    def _plan_export_row_vals(self, line):
        return [
            self._company_display_code(line.company_id),
            line.nganh_hang.ten if line.nganh_hang else '',
            line.ten_hang or '',
            line.ma_hang or '',
            line.ma_sap or '',
            line.qty_t0 or 0.0,
            line.qty_t1 or 0.0,
            line.qty_t2 or 0.0,
            line.qty_t3 or 0.0,
        ]

    def _get_lines_for_sx_export(self):
        """Xuất dòng SX của kỳ."""
        self.ensure_one()
        return self.ke_hoach_san_xuat_ids

    @api.model
    def _download_kinh_doanh_template(self, kd):
        """Tải template import KHKD."""
        if kd.locked:
            raise UserError(_('Kế hoạch đã lấy vào sản xuất, không thể tải template để import lại.'))
        if not kd.company_sx_id:
            raise UserError(_('Đơn vị sản xuất không được để trống.'))

        helper = self.new({'period_month': kd.period_month})
        wb = Workbook()
        ws = wb.active
        ws.title = 'Ke hoach kinh doanh'
        ws.cell(row=1, column=1, value='Mã')
        ws.cell(row=1, column=2, value=kd.code or '')
        ws.cell(row=2, column=1, value='Tháng bắt đầu')
        helper._excel_text_cell(ws, 2, 2, kd.period_month or '')
        ws.cell(row=3, column=1, value='Đơn vị sản xuất')
        ws.cell(row=3, column=2, value=kd.company_sx_id.company_code or kd.company_sx_id.name or '')
        helper._write_plan_data_sheet(wb, ws, kd.line_ids)
        return self._xlsx_download_action(wb, '%s.xlsx' % kd.code)
    def _write_plan_data_sheet(self, wb, ws, lines):
        months = self._get_horizon_months()
        headers = self._PLAN_EXPORT_HEADERS + ['Tháng %s' % month for month in months]
        header_row = 6
        month_col_start = len(self._PLAN_EXPORT_HEADERS) + 1
        for col_idx, label in enumerate(headers, start=1):
            if col_idx >= month_col_start:
                self._excel_text_cell(ws, header_row, col_idx, label)
            else:
                ws.cell(row=header_row, column=col_idx, value=label)
        for row_offset, line in enumerate(lines, start=1):
            for col_idx, value in enumerate(self._plan_export_row_vals(line), start=1):
                ws.cell(row=header_row + row_offset, column=col_idx, value=value)
        self._apply_plan_excel_style(ws, header_row, len(headers))
        self._apply_company_code_validation(wb, ws, first_data_row=header_row + 1)
        for col_idx in range(1, len(headers) + 1):
            max_len = max(
                len(str(ws.cell(row=row_idx, column=col_idx).value or ''))
                for row_idx in range(1, ws.max_row + 1)
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 28)
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 24
        ws.column_dimensions['E'].width = 24
        for letter in ('F', 'G', 'H', 'I'):
            ws.column_dimensions[letter].width = 16

    @api.model
    def _excel_text_cell(self, ws, row, col, value):
        cell = ws.cell(row=row, column=col, value='' if value is None else str(value))
        cell.number_format = FORMAT_TEXT
        return cell

    def _write_plan_metadata(self, ws):
        ws.cell(row=1, column=1, value='Mã')
        ws.cell(row=1, column=2, value=self.code or '')
        ws.cell(row=2, column=1, value='Tháng bắt đầu')
        self._excel_text_cell(ws, 2, 2, self.period_month or '')

    def _has_plan_edit_rights(self):
        return (
            self.env.user.has_group('sonha_vat_tu.group_bo_phan_vat_tu')
            or self.env.user.has_group('sonha_vat_tu.group_truong_bo_phan_vat_tu')
        )

    def _has_kh_dat_vat_tu_export_rights(self):
        return (
            self._has_plan_edit_rights()
            or self.env.user.has_group('sonha_vat_tu.group_ban_cung_ung_vat_tu')
        )

    def action_download_b1_template(self):
        raise UserError(_(
            'Import kế hoạch kinh doanh thực hiện trên menu Kế hoạch kinh doanh.'
        ))

    def _open_import_plan_wizard(self, import_type, label):
        self.ensure_one()
        if self.state != 'ke_hoach':
            raise UserError(
                _('%s đã khóa vì kỳ kế hoạch đã sang bước sau.') % label.capitalize())
        if self.co_ke_hoach_vat_tu:
            raise UserError(_(
                'Đã tạo kế hoạch vật tư, không thể import lại %s.'
            ) % label)
        if not self._has_plan_edit_rights():
            raise UserError(_('Bạn không có quyền import %s.') % label)
        return {
            'name': _('Import %s') % label,
            'type': 'ir.actions.act_window',
            'res_model': 'import.ke.hoach.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'default_period_id': self.id,
                'default_import_type': import_type,
            },
        }

    def action_open_import_kinh_doanh_wizard(self):
        raise UserError(_(
            'Import kế hoạch kinh doanh thực hiện trên menu Kế hoạch kinh doanh.'
        ))

    def action_open_import_san_xuat_wizard(self):
        return self._open_import_plan_wizard('production', 'kế hoạch sản xuất')

    def action_export_san_xuat(self):
        self.ensure_one()
        if self.state != 'ke_hoach':
            raise UserError(_('Kế hoạch đã sang bước sau, không thể export lại cho sản xuất.'))
        if not self._has_plan_edit_rights():
            raise UserError(_('Bạn không có quyền export kế hoạch cho sản xuất.'))
        lines = self._get_lines_for_sx_export()
        if not lines:
            raise UserError(_('Chưa có kế hoạch kinh doanh / sản xuất để export.'))

        wb = Workbook()
        ws = wb.active
        ws.title = 'Ke hoach san xuat'
        self._write_plan_metadata(ws)
        self._write_plan_data_sheet(wb, ws, lines)
        return self._xlsx_download_action(
            wb,
            'KHSX_%s.xlsx' % (self.code or self.id),
        )

    def action_open_import_kh_dat_vat_tu_wizard(self):
        self.ensure_one()
        if self.state != 'dat_hang':
            raise UserError(_(
                'Chỉ import kế hoạch đặt vật tư khi đã ở bước Kế hoạch đặt vật tư.'
            ))
        if not self.kh_dat_vat_tu_ids:
            raise UserError(_(
                'Chưa có dữ liệu kế hoạch đặt vật tư. Vui lòng chạy bước này trước khi import.'
            ))
        view = self.env.ref('sonha_vat_tu.view_import_kh_dat_vat_tu_wizard_form')
        return {
            'name': _('Import kế hoạch đặt vật tư'),
            'type': 'ir.actions.act_window',
            'res_model': 'import.kh.dat.vat.tu.wizard',
            'view_mode': 'form',
            'views': [(view.id, 'form')],
            'view_id': view.id,
            'target': 'new',
            'context': {
                'default_period_id': self.id,
                'period_id_readonly': True,
            },
        }

    def action_export_kh_dat_vat_tu(self):
        """Xuất Excel bước 5."""
        self.ensure_one()
        if not self._has_kh_dat_vat_tu_export_rights():
            raise UserError(_('Bạn không có quyền xuất kế hoạch đặt vật tư.'))
        if self.state in ('ke_hoach', 'dinh_muc', 'tinh_toan', 'tong_hop'):
            raise UserError(_('Chỉ export được từ bước Kế hoạch đặt vật tư trở đi.'))
        lines = self.kh_dat_vat_tu_ids.sorted(
            key=lambda r: ((r.ma_sap or '').strip(), r.id),
        )
        if not lines:
            raise UserError(_('Chưa có dữ liệu kế hoạch đặt vật tư để xuất.'))

        months = self._get_horizon_months()
        if len(months) < 4:
            months = (months + [''] * 4)[:4]

        wb = Workbook()
        ws = wb.active
        ws.title = 'Ke hoach dat vat tu'

        ws.cell(row=1, column=1, value='Số chứng từ')
        ws.cell(row=1, column=2, value=self.code or '')
        ws.cell(row=2, column=1, value='Tháng bắt đầu')
        ws.cell(row=2, column=2, value=self.period_month or '')
        ws.cell(row=3, column=1, value='Đơn vị sản xuất')
        ws.cell(
            row=3, column=2,
            value=self._get_company_code(self.company_sx_id) if self.company_sx_id else '',
        )

        header_row1 = 4
        header_row2 = 5
        data_row = 6

        fixed_start = [
            ('Mã NVL', 'ma_sap', 'text'),
            ('Tên NVL', 'ten_nvl', 'text'),
            ('Chủng loại', 'chung_loai', 'text'),
            ('ĐVT', 'don_vi_tinh', 'dvt'),
            ('Đơn giá tồn kho', 'don_gia_ton_kho', 'money'),
            ('Tồn NVL đầu kỳ', 'tong_ton_nvl_sl', 'qty'),
            ('Giá trị tồn NVL', 'gia_tri_ton_nvl_dau_ky', 'money'),
        ]
        month_groups = [
            ('Cần dùng', 'tong_sl_vt_can_dung_t', 'Tổng cần dùng', 'tong_vt_can_dung'),
            ('Đi đường', 'tong_hang_di_duong_sl_t', 'Tổng đi đường', 'tong_hang_di_duong'),
        ]
        fixed_end = [
            ('Dự trữ tối thiểu đơn vị', 'sl_du_tru_toi_thieu', 'qty'),
            ('Đề xuất đặt mua', 'sl_dat_mua_de_xuat', 'qty'),
            ('Đặt mua chốt', 'sl_dat_mua_chot', 'qty'),
            ('SL cần mua dựa theo MOQ NCC', 'sl_can_mua_theo_moq', 'qty'),
            ('Đơn giá mua', 'don_gia_mua', 'money'),
            ('Giá trị mua', 'gia_tri_mua_hang', 'money'),
            ('Tồn kho cuối kỳ', 'sl_ton_kho_cuoi_ky', 'qty'),
            ('Ngày vòng quay tồn', 'so_ngay_vong_quay_ton', 'qty2'),
            ('Đơn giá tồn cuối kỳ', 'don_gia_ton_kho_cuoi_ky', 'money'),
            ('Giá trị tồn cuối kỳ', 'gia_tri_ton_kho_cuoi_ky', 'money'),
            ('Ghi chú', 'ghi_chu', 'text'),
        ]

        col_specs = []
        col = 1

        for label, field, kind in fixed_start:
            ws.merge_cells(
                start_row=header_row1, start_column=col,
                end_row=header_row2, end_column=col,
            )
            ws.cell(row=header_row1, column=col, value=label)
            col_specs.append((field, kind))
            col += 1

        for group_label, field_prefix, total_label, total_field in month_groups:
            group_start = col
            ws.merge_cells(
                start_row=header_row1, start_column=group_start,
                end_row=header_row1, end_column=group_start + 3,
            )
            ws.cell(row=header_row1, column=group_start, value=group_label)
            for month_offset, month in enumerate(months):
                ws.cell(
                    row=header_row2, column=col,
                    value='Tháng %s' % month if month else 'T%s' % month_offset,
                )
                col_specs.append((f'{field_prefix}{month_offset}', 'qty'))
                col += 1
            ws.merge_cells(
                start_row=header_row1, start_column=col,
                end_row=header_row2, end_column=col,
            )
            ws.cell(row=header_row1, column=col, value=total_label)
            col_specs.append((total_field, 'qty'))
            col += 1

        for label, field, kind in fixed_end:
            ws.merge_cells(
                start_row=header_row1, start_column=col,
                end_row=header_row2, end_column=col,
            )
            ws.cell(row=header_row1, column=col, value=label)
            col_specs.append((field, kind))
            col += 1

        max_col = col - 1
        row_idx = data_row
        for line in lines:
            for col_idx, (field, kind) in enumerate(col_specs, start=1):
                ws.cell(
                    row=row_idx, column=col_idx,
                    value=self._b5_export_value(line, field, kind),
                )
            row_idx += 1

        self._apply_b5_export_style(ws, header_row1, header_row2, max_col, col_specs)

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 8
        for col_idx in range(5, max_col + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14

        code = (self.code or '').strip()
        if code.upper().startswith('KHVT_'):
            file_suffix = code[5:]
        else:
            file_suffix = code or str(self.id)
        return self._xlsx_download_action(
            wb,
            'Data_KHDVT_%s.xlsx' % file_suffix,
        )

    def action_export_vat_tu_can(self):
        """Xuất Excel bước 3 — layout pivot giống UI (Tháng × Đơn vị KD + Tổng)."""
        self.ensure_one()
        lines = self.tinh_toan_vat_tu_ids
        if not lines:
            raise UserError(_('Chưa có dữ liệu tính toán vật tư để xuất.'))
        if self.state in ('ke_hoach', 'dinh_muc'):
            raise UserError(_('Chỉ export được từ bước Tính toán vật tư trở đi.'))

        months = self._get_horizon_months()
        if len(months) < 4:
            months = (months + [''] * 4)[:4]

        # Đơn vị KD xuất hiện trong B3, sort theo mã
        kd_map = {}
        for line in lines:
            if not line.don_vi_kd_id:
                continue
            cid = line.don_vi_kd_id.id
            code = (line.don_vi_kd_code or '').strip()
            if not code:
                code = (
                    self._company_display_code(line.don_vi_kd_id)
                    or '#%s' % cid
                )
            kd_map[cid] = code
        kd_companies = sorted(kd_map.items(), key=lambda item: str(item[1]))

        # Pivot: 1 dòng / mã NVL
        by_mat = {}
        for line in lines:
            key = (line.ma_vat_tu or '').strip() or ('id:%s' % line.id)
            if key not in by_mat:
                by_mat[key] = {
                    'ma_vat_tu': line.ma_vat_tu or '',
                    'ten_vat_tu': line.ten_vat_tu or '',
                    'don_vi_tinh': (
                        line.don_vi_tinh.display_name if line.don_vi_tinh else ''
                    ),
                    'by_company': {},
                }
            if line.don_vi_kd_id:
                by_mat[key]['by_company'][line.don_vi_kd_id.id] = line
        rows = [by_mat[k] for k in sorted(by_mat.keys())]

        wb = Workbook()
        ws = wb.active
        ws.title = 'Vat tu can'

        ws.cell(row=1, column=1, value='Số chứng từ')
        ws.cell(row=1, column=2, value=self.code or '')
        ws.cell(row=2, column=1, value='Tháng bắt đầu')
        ws.cell(row=2, column=2, value=self.period_month or '')

        header_row1 = 4
        header_row2 = 5
        data_row = 6

        # Meta + specs cho style (qty/text)
        col_specs = []
        col = 1
        for label in ('Mã NVL', 'Tên NVL', 'ĐVT'):
            ws.merge_cells(
                start_row=header_row1, start_column=col,
                end_row=header_row2, end_column=col,
            )
            ws.cell(row=header_row1, column=col, value=label)
            col_specs.append((None, 'text'))
            col += 1

        for month_offset, month in enumerate(months):
            group_start = col
            span = len(kd_companies) + 1  # KD codes + Tổng
            ws.merge_cells(
                start_row=header_row1, start_column=group_start,
                end_row=header_row1, end_column=group_start + span - 1,
            )
            month_label = 'Tháng %s' % month if month else 'T%s' % month_offset
            ws.cell(row=header_row1, column=group_start, value=month_label)
            for _cid, code in kd_companies:
                ws.cell(row=header_row2, column=col, value=code)
                col_specs.append((None, 'qty'))
                col += 1
            ws.cell(row=header_row2, column=col, value='Tổng')
            col_specs.append((None, 'qty'))
            col += 1

        max_col = col - 1
        row_idx = data_row
        for row in rows:
            ws.cell(row=row_idx, column=1, value=row['ma_vat_tu'])
            ws.cell(row=row_idx, column=2, value=row['ten_vat_tu'])
            ws.cell(row=row_idx, column=3, value=row['don_vi_tinh'])
            col_idx = 4
            for month_offset in range(4):
                month_total = 0.0
                for cid, _code in kd_companies:
                    line = row['by_company'].get(cid)
                    qty = getattr(line, 'qty_t%s' % month_offset, 0.0) if line else 0.0
                    qty = qty or 0.0
                    month_total += qty
                    ws.cell(row=row_idx, column=col_idx, value=qty)
                    col_idx += 1
                ws.cell(row=row_idx, column=col_idx, value=month_total)
                col_idx += 1
            row_idx += 1

        self._apply_b5_export_style(ws, header_row1, header_row2, max_col, col_specs)

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 10
        for col_idx in range(4, max_col + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12

        code = (self.code or '').strip()
        if code.upper().startswith('KHVT_'):
            file_suffix = code[5:]
        else:
            file_suffix = code or str(self.id)
        return self._xlsx_download_action(
            wb,
            'Data_VatTuCan_%s.xlsx' % file_suffix,
        )

    @api.model
    def load_periods_for_report(self, period_states):
        """Kỳ ở state trong phạm vi record rule của user."""
        return self.search([
            ('state', 'in', period_states),
        ], order='period_month desc, id desc')

    @api.model
    def resolve_period_plans(self, calendar_month, periods):
        """Tháng lịch → danh sách (kỳ, offset T0–T3)."""
        out = []
        for period in periods:
            horizon = period._get_horizon_months()
            if calendar_month not in horizon:
                continue
            out.append((period, horizon.index(calendar_month)))
        return out

    def action_export_tong_hop_vat_tu(self):
        """Xuất Excel bước 4 — layout 2 tầng header giống form Tổng hợp vật tư."""
        self.ensure_one()
        lines = self.tong_hop_vat_tu_ids.filtered(lambda r: not r.don_vi_kd_id).sorted(
            key=lambda r: ((r.ma_sap or '').strip(), r.id),
        )
        if not lines:
            raise UserError(_('Chưa có dữ liệu tổng hợp vật tư để xuất.'))
        if self.state in ('ke_hoach', 'dinh_muc', 'tinh_toan'):
            raise UserError(_('Chỉ export được từ bước Tổng hợp vật tư cần sản xuất trở đi.'))

        months = self._get_horizon_months()
        if len(months) < 4:
            months = (months + [''] * 4)[:4]

        wb = Workbook()
        ws = wb.active
        ws.title = 'Tong hop vat tu'

        ws.cell(row=1, column=1, value='Số chứng từ')
        ws.cell(row=1, column=2, value=self.code or '')
        ws.cell(row=2, column=1, value='Tháng bắt đầu')
        ws.cell(row=2, column=2, value=self.period_month or '')

        header_row1 = 4
        header_row2 = 5
        data_row = 6

        fixed_start = [
            ('Mã NVL', 'ma_sap', 'text'),
            ('Tên NVL', 'ten_nvl', 'text'),
            ('Chủng loại', 'chung_loai', 'text'),
            ('ĐVT', 'don_vi_tinh', 'dvt'),
            ('Tồn đầu', 'ton_dau', 'qty'),
        ]
        month_groups = [
            ('Hàng đi đường', 've_du_kien_don_vi_t'),
            ('Cần dùng', 'vt_can_dung_t'),
            ('Tồn cuối', 'ton_cuoi_t'),
        ]
        fixed_end = [
            ('Dự phòng', 'so_luong_du_phong', 'qty'),
            ('Thiếu', 'so_luong_thieu', 'qty'),
            ('Cần mua', 'so_luong_can_mua', 'qty'),
            ('Ghi chú', 'ghi_chu', 'text'),
        ]

        col_specs = []
        col = 1

        for label, field, kind in fixed_start:
            ws.merge_cells(
                start_row=header_row1, start_column=col,
                end_row=header_row2, end_column=col,
            )
            ws.cell(row=header_row1, column=col, value=label)
            col_specs.append((field, kind))
            col += 1

        for group_label, field_prefix in month_groups:
            group_start = col
            ws.merge_cells(
                start_row=header_row1, start_column=group_start,
                end_row=header_row1, end_column=group_start + 3,
            )
            ws.cell(row=header_row1, column=group_start, value=group_label)
            for month_offset, month in enumerate(months):
                ws.cell(
                    row=header_row2, column=col,
                    value='Tháng %s' % month if month else 'T%s' % month_offset,
                )
                col_specs.append((f'{field_prefix}{month_offset}', 'qty'))
                col += 1

        for label, field, kind in fixed_end:
            ws.merge_cells(
                start_row=header_row1, start_column=col,
                end_row=header_row2, end_column=col,
            )
            ws.cell(row=header_row1, column=col, value=label)
            col_specs.append((field, kind))
            col += 1

        max_col = col - 1
        row_idx = data_row
        for line in lines:
            for col_idx, (field, kind) in enumerate(col_specs, start=1):
                ws.cell(
                    row=row_idx, column=col_idx,
                    value=self._b5_export_value(line, field, kind),
                )
            row_idx += 1

        self._apply_b5_export_style(ws, header_row1, header_row2, max_col, col_specs, meta_rows=3)

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 12
        for col_idx in range(6, max_col + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12

        code = (self.code or '').strip()
        if code.upper().startswith('KHVT_'):
            file_suffix = code[5:]
        else:
            file_suffix = code or str(self.id)
        return self._xlsx_download_action(
            wb,
            'Data_TongHopVT_%s.xlsx' % file_suffix,
        )

    @staticmethod
    def _b5_export_value(line, field, kind):
        if field == 'don_vi_tinh':
            dvt = line.don_vi_tinh
            return dvt.display_name if dvt else ''
        value = getattr(line, field, False)
        if kind == 'text':
            return value or ''
        if value in (False, None):
            return 0.0 if kind != 'text' else ''
        return value

    def _apply_b5_export_style(self, ws, header_row1, header_row2, max_col, col_specs, meta_rows=2,
                               header_row1_height=22, header_row2_height=22):
        header_font = Font(name='Times New Roman', size=10, bold=True, color='FFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='3F6F8F')
        header_side = Side(style='thin', color='2F556D')
        header_border = Border(
            left=header_side, right=header_side, top=header_side, bottom=header_side,
        )
        label_font = Font(name='Times New Roman', size=10, bold=True)
        base_font = Font(name='Times New Roman', size=10)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left = Alignment(horizontal='left', vertical='center')
        right = Alignment(horizontal='right', vertical='center')
        qty_fmt = '#,##0.000'
        qty2_fmt = '#,##0.00'
        money_fmt = '#,##0'

        for row_idx in range(1, meta_rows + 1):
            ws.cell(row=row_idx, column=1).font = label_font

        for row_idx in (header_row1, header_row2):
            height = header_row1_height if row_idx == header_row1 else header_row2_height
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = header_border
                cell.alignment = center
            ws.row_dimensions[row_idx].height = height

        kind_fmt = {'qty': qty_fmt, 'qty2': qty2_fmt, 'money': money_fmt}
        for row_idx in range(header_row2 + 1, ws.max_row + 1):
            for col_idx, (_label, kind) in enumerate(col_specs, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = base_font
                if kind == 'text':
                    cell.alignment = left
                else:
                    cell.alignment = right
                    fmt = kind_fmt.get(kind)
                    if fmt:
                        cell.number_format = fmt

        ws.freeze_panes = ws.cell(row=header_row2 + 1, column=1).coordinate

    _IMPORT_BCU_ACTION_XMLID = 'sonha_vat_tu.action_import_tong_hop_bcu_server'
    _EXPORT_B5_ACTION_XMLID = 'sonha_vat_tu.action_export_kh_dat_vat_tu_server'
    _EXPORT_B4_ACTION_XMLID = 'sonha_vat_tu.action_export_tong_hop_vat_tu_server'
    _EXPORT_B3_ACTION_XMLID = 'sonha_vat_tu.action_export_vat_tu_can_server'
    _B3_FORM_VIEW_XMLID = 'sonha_vat_tu.view_ke_hoach_vat_tu_form_b3'
    _B4_FORM_VIEW_XMLID = 'sonha_vat_tu.view_ke_hoach_vat_tu_form_b4'
    _B5_FORM_VIEW_XMLID = 'sonha_vat_tu.view_ke_hoach_vat_tu_form_b5'
    _B6_FORM_VIEW_XMLID = 'sonha_vat_tu.view_ke_hoach_vat_tu_form_b6'
    _B7_FORM_VIEW_XMLID = 'sonha_vat_tu.view_ke_hoach_vat_tu_form_b7'

    def _toolbar_remove_action(self, toolbar, action_xmlid):
        action = self.env.ref(action_xmlid, raise_if_not_found=False)
        if not action:
            return
        for key in ('action', 'print'):
            items = toolbar.get(key)
            if items:
                toolbar[key] = [
                    item for item in items
                    if item.get('id') != action.id
                ]

    _ACTION_FORM_VIEW_XMLIDS = (
        ('sonha_vat_tu.action_ke_hoach_san_xuat_period', 'sonha_vat_tu.view_ke_hoach_vat_tu_form_sx'),
        ('sonha_vat_tu.action_ke_hoach_vat_tu_period', 'sonha_vat_tu.view_ke_hoach_vat_tu_form_vt'),
        ('sonha_vat_tu.action_ke_hoach_vat_tu_b2', 'sonha_vat_tu.view_ke_hoach_vat_tu_form_b2'),
        ('sonha_vat_tu.action_ke_hoach_vat_tu_b3', 'sonha_vat_tu.view_ke_hoach_vat_tu_form_b3'),
        ('sonha_vat_tu.action_ke_hoach_vat_tu_b4', 'sonha_vat_tu.view_ke_hoach_vat_tu_form_b4'),
        ('sonha_vat_tu.action_ke_hoach_vat_tu_b5', 'sonha_vat_tu.view_ke_hoach_vat_tu_form_b5'),
        ('sonha_vat_tu.action_ke_hoach_vat_tu_b6', 'sonha_vat_tu.view_ke_hoach_vat_tu_form_b6'),
        ('sonha_vat_tu.action_ke_hoach_vat_tu_b7', 'sonha_vat_tu.view_ke_hoach_vat_tu_form_b7'),
    )

    @api.model
    def _form_view_id_for_action(self, action_id):
        if not action_id:
            return False
        for action_xmlid, view_xmlid in self._ACTION_FORM_VIEW_XMLIDS:
            action = self.env.ref(action_xmlid, raise_if_not_found=False)
            if action and action.id == action_id:
                view = self.env.ref(view_xmlid, raise_if_not_found=False)
                return view.id if view else False
        return False

    @api.model
    def _apply_action_form_view(self, views, options):
        form_view_id = self._form_view_id_for_action((options or {}).get('action_id'))
        if not form_view_id:
            return views
        return [
            (form_view_id if vtype == 'form' else vid, vtype)
            for vid, vtype in views
        ]

    @api.model
    def get_views(self, views, options=None):
        """Ẩn Import BCU / Export B3 / Export B5 khỏi form không đúng bước."""
        options = dict(options or {})
        views = self._apply_action_form_view(list(views), options)
        res = super().get_views(views, options=options)
        form = res.get('views', {}).get('form')
        if not form or not (options or {}).get('toolbar'):
            return res
        toolbar = form.setdefault('toolbar', {})
        form_view_id = form.get('id')

        b3_view = self.env.ref(self._B3_FORM_VIEW_XMLID, raise_if_not_found=False)
        b4_view = self.env.ref(self._B4_FORM_VIEW_XMLID, raise_if_not_found=False)
        b5_view = self.env.ref(self._B5_FORM_VIEW_XMLID, raise_if_not_found=False)
        b6_view = self.env.ref(self._B6_FORM_VIEW_XMLID, raise_if_not_found=False)

        if b4_view and form_view_id != b4_view.id:
            self._toolbar_remove_action(toolbar, self._EXPORT_B4_ACTION_XMLID)
        if b6_view and form_view_id != b6_view.id:
            self._toolbar_remove_action(toolbar, self._IMPORT_BCU_ACTION_XMLID)
        if b3_view and form_view_id != b3_view.id:
            self._toolbar_remove_action(toolbar, self._EXPORT_B3_ACTION_XMLID)
        if b5_view and form_view_id != b5_view.id:
            self._toolbar_remove_action(toolbar, self._EXPORT_B5_ACTION_XMLID)
        return res

    def _form_view_id_from_context(self):
        form_ref = self.env.context.get('form_view_ref') or ''
        if form_ref:
            view = self.env.ref(form_ref, raise_if_not_found=False)
            if view:
                return view.id
        action_id = self.env.context.get('action')
        if not action_id:
            params = self.env.context.get('params') or {}
            if isinstance(params, dict):
                action_id = params.get('action')
        return self._form_view_id_for_action(action_id)

    def get_formview_id(self, access_uid=None):
        """Luôn theo action/menu; không dùng sticky view cũ."""
        self.ensure_one()
        view_id = self._form_view_id_from_context()
        if view_id:
            return view_id
        return super().get_formview_id(access_uid=access_uid)

    def _action_open_step(self, action_xmlid):
        self.ensure_one()
        action = self.env['ir.actions.act_window']._for_xml_id(action_xmlid)
        action['res_id'] = self.id
        form_views = [
            (view_id, view_type)
            for view_id, view_type in action.get('views', [])
            if view_type == 'form'
        ]
        if form_views:
            action['views'] = form_views
            action['view_mode'] = 'form'
        return action

    def _clear_step_data(self, from_state):
        """Xóa dữ liệu bước hiện tại khi Reset về bước trước."""
        self.ensure_one()
        period = self.sudo()
        ctx = {'tracking_disable': True}
        pid = self.id

        if from_state == 'phe_duyet':
            period.phe_duyet_kh_vat_tu_ids.with_context(**ctx).unlink()
            period.approval_step_ids.unlink()
            period.write({
                'approval_state': 'draft',
                'approval_current_sequence': 1,
            })
        elif from_state == 'bcu_tong_hop':
            period.kh_dat_vat_tu_bcu_ids.with_context(**ctx).unlink()
        elif from_state == 'dat_hang':
            period.kh_dat_vat_tu_ids.with_context(
                force_b5_unlink=True, **ctx,
            ).unlink()
        elif from_state == 'tong_hop':
            period.tong_hop_vat_tu_ids.with_context(**ctx).unlink()
        elif from_state == 'tinh_toan':
            self.env['tinh.toan.vat.tu.chi.tiet'].sudo().search([
                ('period_id', '=', pid),
            ]).with_context(**ctx).unlink()
            period.tinh_toan_vat_tu_ids.with_context(**ctx).unlink()
        elif from_state == 'dinh_muc':
            period.dinh_muc_ids.with_context(**ctx).unlink()

        self.invalidate_recordset([
            'kh_dat_vat_tu_ids', 'kh_dat_vat_tu_count',
            'kh_dat_vat_tu_bcu_ids', 'kh_dat_vat_tu_bcu_count',
            'phe_duyet_kh_vat_tu_ids', 'phe_duyet_kh_vat_tu_count',
            'tong_hop_vat_tu_ids', 'tong_hop_vat_tu_count',
            'tinh_toan_vat_tu_ids', 'tinh_toan_vat_tu_count',
            'dinh_muc_ids', 'dinh_muc_count',
            'approval_step_ids', 'approval_state', 'approval_current_sequence',
        ])

    def _action_reset_step(self, target_state, action_xmlid, extra_vals=None):
        self.ensure_one()
        vals = {'state': target_state}
        if extra_vals:
            vals.update(extra_vals)
        self.write(vals)

        action = self.env['ir.actions.act_window']._for_xml_id(action_xmlid)
        action['res_id'] = self.id
        action['target'] = 'main'
        form_views = [
            (view_id, view_type)
            for view_id, view_type in action.get('views', [])
            if view_type == 'form'
        ]
        if form_views:
            action['views'] = form_views
            action['view_mode'] = 'form'
        ctx = action.get('context') or {}
        if isinstance(ctx, str):
            ctx = safe_eval(ctx)
        ctx = dict(ctx)
        ctx['no_breadcrumbs'] = False
        action['context'] = ctx
        return action

    _RESET_PREVIOUS = {
        'dinh_muc': ('ke_hoach', 'sonha_vat_tu.action_ke_hoach_vat_tu_period'),
        'tinh_toan': ('dinh_muc', 'sonha_vat_tu.action_ke_hoach_vat_tu_b2'),
        'tong_hop': ('tinh_toan', 'sonha_vat_tu.action_ke_hoach_vat_tu_b3'),
        'dat_hang': ('tong_hop', 'sonha_vat_tu.action_ke_hoach_vat_tu_b4'),
        'bcu_tong_hop': ('dat_hang', 'sonha_vat_tu.action_ke_hoach_vat_tu_b5'),
        'phe_duyet': ('bcu_tong_hop', 'sonha_vat_tu.action_ke_hoach_vat_tu_b6'),
    }

    def action_reset_previous_step(self):
        self.ensure_one()
        mapping = self._RESET_PREVIOUS.get(self.state)
        if not mapping:
            raise UserError(_('Không thể quay lại bước trước từ trạng thái hiện tại.'))
        if self.state == 'phe_duyet' and self.approval_state == 'approved':
            raise UserError(_(
                'Không thể quay lại khi phê duyệt kế hoạch vật tư đã hoàn tất.'
            ))
        if self.state in ('bcu_tong_hop', 'phe_duyet'):
            self._check_bcu_workflow_access()
        target_state, action_xmlid = mapping
        self._clear_step_data(self.state)
        return self._action_reset_step(target_state, action_xmlid, None)

    def action_open_workflow_sx(self):
        return self._action_open_step('sonha_vat_tu.action_ke_hoach_san_xuat_period')

    def action_open_workflow_vt(self):
        return self._action_open_step('sonha_vat_tu.action_ke_hoach_vat_tu_period')

    def action_open_step_b1(self):
        return self._action_open_step('sonha_vat_tu.action_ke_hoach_vat_tu_period')

    def action_open_step_b2(self):
        return self._action_open_step('sonha_vat_tu.action_ke_hoach_vat_tu_b2')

    def action_open_dinh_muc_period(self):
        self.ensure_one()
        tree_view = self.env.ref('sonha_vat_tu.view_dinh_muc_tree', raise_if_not_found=False)
        return {
            'type': 'ir.actions.act_window',
            'name': _('Định mức kỳ'),
            'res_model': 'dinh.muc',
            'view_mode': 'tree',
            'views': [(tree_view.id, 'tree')] if tree_view else False,
            'domain': [('period_id', '=', self.id)],
            'context': {'default_period_id': self.id},
        }

    def action_open_step_b3(self):
        return self._action_open_step('sonha_vat_tu.action_ke_hoach_vat_tu_b3')

    def action_open_step_b4(self):
        return self._action_open_step('sonha_vat_tu.action_ke_hoach_vat_tu_b4')

    def action_open_step_b5(self):
        return self._action_open_step('sonha_vat_tu.action_ke_hoach_vat_tu_b5')

    def action_open_step_b6(self):
        self._check_bcu_workflow_access()
        return self._action_open_step('sonha_vat_tu.action_ke_hoach_vat_tu_b6')

    def action_open_step_b7(self):
        self._check_bcu_workflow_access()
        return self._action_open_step('sonha_vat_tu.action_ke_hoach_vat_tu_b7')
