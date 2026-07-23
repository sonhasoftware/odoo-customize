# -*- coding: utf-8 -*-
import base64
import io
from datetime import date

from openpyxl import Workbook

from odoo import _, api, fields, models
from odoo.exceptions import UserError

class BaoCaoTinhToanVatTuWizard(models.TransientModel):
    _name = 'bao.cao.tinh.toan.vat.tu.wizard'
    _description = 'Wizard báo cáo tính toán vật tư cần'

    def _month_selection(self):
        return [(str(m), '%02d' % m) for m in range(1, 13)]

    def _year_selection(self):
        y = date.today().year
        return [(str(i), str(i)) for i in range(y - 5, y + 6)]

    from_month = fields.Selection(
        selection=_month_selection, string='Từ tháng', required=True,
        default=lambda self: str(date.today().month))
    from_year = fields.Selection(
        selection=_year_selection, string='Từ năm', required=True,
        default=lambda self: str(date.today().year))
    to_month = fields.Selection(
        selection=_month_selection, string='Đến tháng', required=True)
    to_year = fields.Selection(
        selection=_year_selection, string='Đến năm', required=True,
        default=lambda self: str(date.today().year))
    company_ids = fields.Many2many(
        'res.company', 'bao_cao_b3_wizard_company_rel', 'wizard_id', 'company_id',
        string='Đơn vị đặt hàng',
        help='Đơn vị kinh doanh (SHI, NAN, TM2, B2B…).',
    )
    line_ids = fields.One2many(
        'bao.cao.tinh.toan.vat.tu.line', 'wizard_id', string='Chi tiết')

    thang_tu = fields.Char(string='Tháng từ', readonly=True)
    thang_den = fields.Char(string='Tháng đến', readonly=True)

    _KD_COMPANY_CODES = ('SHI', 'NAN', 'TM2', 'B2B', 'TM1', 'SHD')

    @api.model
    def _get_kd_company_ids(self):
        """Đơn vị KD mặc định — dùng sudo để user không cần quyền đọc res.company."""
        companies = self.env['res.company'].sudo().search([])
        kd = companies.filtered(
            lambda c: (c.company_code or '').strip().upper() in self._KD_COMPANY_CODES
        )
        return kd.ids

    def _date_range(self):
        self.ensure_one()
        period_model = self.env['ke.hoach.vat.tu']
        d_from = period_model.month_start_from_key(
            '%s/%s' % (int(self.from_month), self.from_year))
        d_to = period_model.month_end_from_key(
            '%s/%s' % (int(self.to_month), self.to_year))
        period_model.validate_report_month_range(d_from, d_to)
        return d_from, d_to

    def _populate_lines(self):
        self.ensure_one()
        period_model = self.env['ke.hoach.vat.tu']
        company_ids = self.company_ids.ids or self._get_kd_company_ids()
        d_from, d_to = self._date_range()
        self.thang_tu = '%02d/%s' % (int(self.from_month), self.from_year)
        self.thang_den = '%02d/%s' % (int(self.to_month), self.to_year)
        calendar_months = period_model.iter_calendar_months(d_from, d_to)
        Line = self.env['bao.cao.tinh.toan.vat.tu.line']
        self.line_ids.unlink()
        if not calendar_months or not company_ids:
            return

        periods = period_model.load_periods_for_report(
            ('tinh_toan', 'tong_hop', 'dat_hang'))
        month_plan = {}
        period_ids = set()
        for cm in calendar_months:
            for period, offset in period_model.resolve_period_plans(cm, periods):
                month_plan.setdefault(cm, []).append((period, offset))
                period_ids.add(period.id)

        if not period_ids:
            return

        b3_rows = self.env['tinh.toan.vat.tu'].search([
            ('period_id', 'in', list(period_ids)),
            ('don_vi_kd_id', 'in', company_ids),
        ])
        materials = {}
        qty_map = {}
        for row in b3_rows:
            key = (row.ma_vat_tu or '').strip()
            if not key:
                continue
            materials[key] = {
                'ten_vat_tu': row.ten_vat_tu,
                'don_vi_tinh': row.don_vi_tinh.id if row.don_vi_tinh else False,
            }
            for off in range(4):
                qty_map[(row.period_id.id, row.don_vi_kd_id.id, key, off)] = (
                    getattr(row, 'qty_t%d' % off) or 0.0
                )

        lines = []
        Company = self.env['res.company'].sudo()
        for ma in sorted(materials.keys()):
            meta = materials[ma]
            for cm in calendar_months:
                plans = month_plan.get(cm) or []
                for period, offset in plans:
                    for cid in company_ids:
                        qty = qty_map.get((period.id, cid, ma, offset), 0.0)
                        if not qty:
                            continue
                        company = Company.browse(cid)
                        lines.append({
                            'wizard_id': self.id,
                            'ma_vat_tu': ma,
                            'ten_vat_tu': meta['ten_vat_tu'],
                            'don_vi_tinh': meta['don_vi_tinh'],
                            'don_vi_kd_id': cid,
                            'don_vi_kd_code': company.company_code or company.name,
                            'month_key': cm,
                            'qty': qty,
                            'period_code': period.code,
                            'qty_offset': 'T%d' % offset,
                        })
        if lines:
            Line.create(lines)

    def action_open_report(self):
        self.ensure_one()
        self._populate_lines()
        d_from, d_to = self._date_range()
        calendar_months = self.env['ke.hoach.vat.tu'].iter_calendar_months(d_from, d_to)
        month_keys_str = ','.join(calendar_months)
        return {
            'type': 'ir.actions.act_window',
            'name': _('Báo cáo tính toán vật tư cần'),
            'res_model': 'bao.cao.tinh.toan.vat.tu.line',
            'view_mode': 'tree',
            'domain': [('wizard_id', '=', self.id)],
            'context': {
                'bao_cao_b3_wizard_id': self.id,
                'bao_cao_thang_tu': self.thang_tu,
                'bao_cao_thang_den': self.thang_den,
                'bao_cao_month_keys': month_keys_str,
            },
        }

    def action_export_excel(self):
        self.ensure_one()
        if not self.line_ids:
            self._populate_lines()
        lines = self.line_ids
        d_from, d_to = self._date_range()
        calendar_months = self.env['ke.hoach.vat.tu'].iter_calendar_months(d_from, d_to)

        kd_map = {}
        for line in lines:
            if line.don_vi_kd_id:
                kd_map[line.don_vi_kd_id.id] = (
                    line.don_vi_kd_code
                    or line.don_vi_kd_id.company_code
                    or line.don_vi_kd_id.name
                )
        kd_companies = sorted(kd_map.items(), key=lambda x: str(x[1]))

        by_mat = {}
        for line in lines:
            key = line.ma_vat_tu
            by_mat.setdefault(key, {
                'ma_vat_tu': line.ma_vat_tu,
                'ten_vat_tu': line.ten_vat_tu,
                'don_vi_tinh': line.don_vi_tinh.display_name if line.don_vi_tinh else '',
                'cells': {},
            })
            by_mat[key]['cells'][(line.month_key, line.don_vi_kd_id.id)] = line.qty

        wb = Workbook()
        ws = wb.active
        ws.title = 'Vat tu can'

        ws.cell(row=1, column=1, value='Báo cáo tính toán vật tư cần')
        ws.cell(row=2, column=1, value='Từ tháng')
        ws.cell(row=2, column=2, value=self.thang_tu)
        ws.cell(row=3, column=1, value='Đến tháng')
        ws.cell(row=3, column=2, value=self.thang_den)

        header_row1 = 5
        header_row2 = 6
        data_row = 7
        col = 1

        for label in ('Mã NVL', 'Tên NVL', 'ĐVT'):
            ws.merge_cells(
                start_row=header_row1, start_column=col,
                end_row=header_row2, end_column=col,
            )
            ws.cell(row=header_row1, column=col, value=label)
            col += 1

        for month_key in calendar_months:
            group_start = col
            span = len(kd_companies) + 1
            ws.merge_cells(
                start_row=header_row1, start_column=group_start,
                end_row=header_row1, end_column=group_start + span - 1,
            )
            ws.cell(row=header_row1, column=group_start, value='Tháng %s' % month_key)
            for cid, code in kd_companies:
                ws.cell(row=header_row2, column=col, value=code)
                col += 1
            ws.cell(row=header_row2, column=col, value='Tổng')
            col += 1

        max_col = col - 1
        row_idx = data_row
        for ma in sorted(by_mat.keys()):
            row = by_mat[ma]
            ws.cell(row=row_idx, column=1, value=row['ma_vat_tu'])
            ws.cell(row=row_idx, column=2, value=row['ten_vat_tu'])
            ws.cell(row=row_idx, column=3, value=row['don_vi_tinh'])
            col_idx = 4
            for month_key in calendar_months:
                for cid, _code in kd_companies:
                    ws.cell(
                        row=row_idx, column=col_idx,
                        value=row['cells'].get((month_key, cid), 0.0),
                    )
                    col_idx += 1
                ws.cell(
                    row=row_idx, column=col_idx,
                    value=sum(
                        row['cells'].get((month_key, c), 0.0)
                        for c, _ in kd_companies
                    ),
                )
                col_idx += 1
            row_idx += 1

        self.env['ke.hoach.vat.tu']._apply_b5_export_style(
            ws, header_row1, header_row2, max_col,
            [(None, 'text')] * 3 + [(None, 'qty')] * (max_col - 3),
            meta_rows=3,
        )
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 10

        output = io.BytesIO()
        wb.save(output)
        fname = 'BaoCao_TinhToanVT_%s_%s.xlsx' % (
            self.thang_tu.replace('/', ''),
            self.thang_den.replace('/', ''),
        )
        attachment = self.env['ir.attachment'].sudo().create({
            'name': fname,
            'type': 'binary',
            'datas': base64.b64encode(output.getvalue()),
            'mimetype': (
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            ),
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }


class BaoCaoTinhToanVatTuLine(models.TransientModel):
    _name = 'bao.cao.tinh.toan.vat.tu.line'
    _description = 'Dòng báo cáo tính toán vật tư cần'
    _order = 'ma_vat_tu, month_key, don_vi_kd_code, id'

    wizard_id = fields.Many2one(
        'bao.cao.tinh.toan.vat.tu.wizard', ondelete='cascade', index=True)
    ma_vat_tu = fields.Char(string='Mã NVL', index=True)
    ten_vat_tu = fields.Char(string='Tên NVL')
    don_vi_tinh = fields.Many2one('mdm.dvt', string='ĐVT')
    don_vi_kd_id = fields.Many2one('res.company', string='Đơn vị KD')
    don_vi_kd_code = fields.Char(string='Mã đơn vị KD')
    month_key = fields.Char(string='Tháng', index=True)
    qty = fields.Float(string='Số lượng', digits=(16, 3))
    period_code = fields.Char(string='Kỳ nguồn')
    qty_offset = fields.Char(string='Cột nguồn')

    def action_export_excel(self):
        if self:
            wizards = self.mapped('wizard_id')
        else:
            wizard_id = self.env.context.get('bao_cao_b3_wizard_id')
            wizards = (
                self.env['bao.cao.tinh.toan.vat.tu.wizard'].browse(wizard_id)
                if wizard_id else self.env['bao.cao.tinh.toan.vat.tu.wizard']
            )
        if len(wizards) != 1:
            raise UserError(_('Vui lòng xuất Excel từ một báo cáo đã mở.'))
        return wizards.action_export_excel()
