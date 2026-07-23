# -*- coding: utf-8 -*-
import base64
import io
from datetime import date

from openpyxl import Workbook

from odoo import _, api, fields, models
from odoo.exceptions import UserError

class BaoCaoTongHopVatTuWizard(models.TransientModel):
    _name = 'bao.cao.tong.hop.vat.tu.wizard'
    _description = 'Wizard báo cáo tổng hợp vật tư cần sản xuất'

    def _month_selection(self):
        return [(str(m), '%02d' % m) for m in range(1, 13)]

    def _year_selection(self):
        y = date.today().year
        return [(str(i), str(i)) for i in range(y - 5, y + 6)]

    from_month = fields.Selection(
        selection=_month_selection, string='Từ tháng', required=True,
        default=lambda self: str(date.today().month))
    from_year = fields.Selection(
        selection=_year_selection, string='Từ năm', required=True,
        default=lambda self: str(date.today().year))
    to_month = fields.Selection(
        selection=_month_selection, string='Đến tháng', required=True)
    to_year = fields.Selection(
        selection=_year_selection, string='Đến năm', required=True,
        default=lambda self: str(date.today().year))
    line_ids = fields.One2many(
        'bao.cao.tong.hop.vat.tu.line', 'wizard_id', string='Chi tiết')

    thang_tu = fields.Char(string='Tháng từ', readonly=True)
    thang_den = fields.Char(string='Tháng đến', readonly=True)

    def _date_range(self):
        self.ensure_one()
        period_model = self.env['ke.hoach.vat.tu']
        d_from = period_model.month_start_from_key(
            '%s/%s' % (int(self.from_month), self.from_year))
        d_to = period_model.month_end_from_key(
            '%s/%s' % (int(self.to_month), self.to_year))
        period_model.validate_report_month_range(d_from, d_to)
        return d_from, d_to

    def _populate_lines(self):
        self.ensure_one()
        period_model = self.env['ke.hoach.vat.tu']
        d_from, d_to = self._date_range()
        self.thang_tu = '%02d/%s' % (int(self.from_month), self.from_year)
        self.thang_den = '%02d/%s' % (int(self.to_month), self.to_year)
        calendar_months = period_model.iter_calendar_months(d_from, d_to)
        Line = self.env['bao.cao.tong.hop.vat.tu.line']
        self.line_ids.unlink()
        if not calendar_months:
            return

        periods = period_model.load_periods_for_report(
            ('tong_hop', 'dat_hang'))
        month_plan = {}
        period_ids = set()
        for cm in calendar_months:
            for period, offset in period_model.resolve_period_plans(cm, periods):
                month_plan.setdefault(cm, []).append((period, offset))
                period_ids.add(period.id)

        if not period_ids:
            return

        th_rows = self.env['tong.hop.vat.tu'].search([
            ('period_id', 'in', list(period_ids)),
            ('don_vi_kd_id', '=', False),
        ])
        materials = {}
        metric_map = {}
        for row in th_rows:
            key = (row.ma_sap or '').strip()
            if not key:
                continue
            materials[key] = {
                'ten_nvl': row.ten_nvl,
                'chung_loai': row.chung_loai,
                'don_vi_tinh': row.don_vi_tinh.id if row.don_vi_tinh else False,
                'ton_dau': row.ton_dau or 0.0,
            }
            for off in range(4):
                metric_map[(row.period_id.id, key, off)] = {
                    've_du_kien_don_vi': getattr(row, 've_du_kien_don_vi_t%d' % off) or 0.0,
                    've_du_kien_bcu': getattr(row, 've_du_kien_t%d' % off) or 0.0,
                    'vt_can_dung': getattr(row, 'vt_can_dung_t%d' % off) or 0.0,
                    'ton_cuoi': getattr(row, 'ton_cuoi_t%d' % off) or 0.0,
                    'ton_dau': row.ton_dau or 0.0,
                }

        lines = []
        for ma in sorted(materials.keys()):
            meta = materials[ma]
            for cm in calendar_months:
                plans = month_plan.get(cm) or []
                for period, offset in plans:
                    metrics = metric_map.get((period.id, ma, offset), {})
                    if not any([
                        metrics.get('ve_du_kien_don_vi'),
                        metrics.get('ve_du_kien_bcu'),
                        metrics.get('vt_can_dung'),
                        metrics.get('ton_cuoi'),
                    ]):
                        continue
                    lines.append({
                        'wizard_id': self.id,
                        'ma_sap': ma,
                        'ten_nvl': meta['ten_nvl'],
                        'chung_loai': meta['chung_loai'],
                        'don_vi_tinh': meta['don_vi_tinh'],
                        'ton_dau': metrics.get('ton_dau', meta['ton_dau']),
                        'month_key': cm,
                        've_du_kien_don_vi': metrics.get('ve_du_kien_don_vi', 0.0),
                        've_du_kien_bcu': metrics.get('ve_du_kien_bcu', 0.0),
                        'vt_can_dung': metrics.get('vt_can_dung', 0.0),
                        'ton_cuoi': metrics.get('ton_cuoi', 0.0),
                        'period_code': period.code,
                        'qty_offset': 'T%d' % offset,
                    })
        if lines:
            Line.create(lines)

    def action_open_report(self):
        self.ensure_one()
        self._populate_lines()
        d_from, d_to = self._date_range()
        calendar_months = self.env['ke.hoach.vat.tu'].iter_calendar_months(d_from, d_to)
        month_keys_str = ','.join(calendar_months)
        return {
            'type': 'ir.actions.act_window',
            'name': _('Báo cáo tổng hợp vật tư cần sản xuất'),
            'res_model': 'bao.cao.tong.hop.vat.tu.line',
            'view_mode': 'tree',
            'domain': [('wizard_id', '=', self.id)],
            'context': {
                'bao_cao_b4_wizard_id': self.id,
                'bao_cao_thang_tu': self.thang_tu,
                'bao_cao_thang_den': self.thang_den,
                'bao_cao_month_keys': month_keys_str,
            },
        }

    def action_export_excel(self):
        self.ensure_one()
        if not self.line_ids:
            self._populate_lines()
        lines = self.line_ids
        d_from, d_to = self._date_range()
        calendar_months = self.env['ke.hoach.vat.tu'].iter_calendar_months(d_from, d_to)

        by_mat = {}
        for line in lines:
            key = line.ma_sap
            entry = by_mat.setdefault(key, {
                'ma_sap': line.ma_sap,
                'ten_nvl': line.ten_nvl,
                'chung_loai': line.chung_loai or '',
                'don_vi_tinh': line.don_vi_tinh.display_name if line.don_vi_tinh else '',
                'ton_dau': line.ton_dau or 0.0,
                'months': {},
            })
            entry['months'][line.month_key] = {
                've_du_kien_don_vi': line.ve_du_kien_don_vi,
                've_du_kien_bcu': line.ve_du_kien_bcu,
                'vt_can_dung': line.vt_can_dung,
                'ton_cuoi': line.ton_cuoi,
            }

        wb = Workbook()
        ws = wb.active
        ws.title = 'Tong hop vat tu'

        ws.cell(row=1, column=1, value='Báo cáo tổng hợp vật tư cần sản xuất')
        ws.cell(row=2, column=1, value='Từ tháng')
        ws.cell(row=2, column=2, value=self.thang_tu)
        ws.cell(row=3, column=1, value='Đến tháng')
        ws.cell(row=3, column=2, value=self.thang_den)

        header_row1 = 5
        header_row2 = 6
        data_row = 7
        col = 1
        fixed_start = ['Mã NVL', 'Tên NVL', 'Chủng loại', 'ĐVT', 'Tồn đầu']
        month_groups = [
            ('Hàng đi đường đơn vị', 've_du_kien_don_vi'),
            ('Hàng đi đường BCU', 've_du_kien_bcu'),
            ('Cần dùng', 'vt_can_dung'),
            ('Tồn cuối', 'ton_cuoi'),
        ]

        for label in fixed_start:
            ws.merge_cells(
                start_row=header_row1, start_column=col,
                end_row=header_row2, end_column=col,
            )
            ws.cell(row=header_row1, column=col, value=label)
            col += 1

        for group_label, _field in month_groups:
            group_start = col
            span = len(calendar_months)
            ws.merge_cells(
                start_row=header_row1, start_column=group_start,
                end_row=header_row1, end_column=group_start + span - 1,
            )
            ws.cell(row=header_row1, column=group_start, value=group_label)
            for month_key in calendar_months:
                ws.cell(row=header_row2, column=col, value='Tháng %s' % month_key)
                col += 1

        max_col = col - 1
        row_idx = data_row
        for ma in sorted(by_mat.keys()):
            row = by_mat[ma]
            ws.cell(row=row_idx, column=1, value=row['ma_sap'])
            ws.cell(row=row_idx, column=2, value=row['ten_nvl'])
            ws.cell(row=row_idx, column=3, value=row['chung_loai'])
            ws.cell(row=row_idx, column=4, value=row['don_vi_tinh'])
            ws.cell(row=row_idx, column=5, value=row['ton_dau'])
            col_idx = 6
            for _group_label, field in month_groups:
                for month_key in calendar_months:
                    month_data = row['months'].get(month_key, {})
                    ws.cell(row=row_idx, column=col_idx, value=month_data.get(field, 0.0))
                    col_idx += 1
            row_idx += 1

        self.env['ke.hoach.vat.tu']._apply_b5_export_style(
            ws, header_row1, header_row2, max_col,
            [(None, 'text')] * 5 + [(None, 'qty')] * (max_col - 5),
            meta_rows=3,
            header_row1_height=36,
            header_row2_height=28,
        )
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 12

        output = io.BytesIO()
        wb.save(output)
        fname = 'BaoCao_TongHopVT_%s_%s.xlsx' % (
            self.thang_tu.replace('/', ''),
            self.thang_den.replace('/', ''),
        )
        attachment = self.env['ir.attachment'].sudo().create({
            'name': fname,
            'type': 'binary',
            'datas': base64.b64encode(output.getvalue()),
            'mimetype': (
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            ),
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }


class BaoCaoTongHopVatTuLine(models.TransientModel):
    _name = 'bao.cao.tong.hop.vat.tu.line'
    _description = 'Dòng báo cáo tổng hợp vật tư'
    _order = 'ma_sap, month_key, id'

    wizard_id = fields.Many2one(
        'bao.cao.tong.hop.vat.tu.wizard', ondelete='cascade', index=True)
    ma_sap = fields.Char(string='Mã NVL', index=True)
    ten_nvl = fields.Char(string='Tên NVL')
    chung_loai = fields.Char(string='Chủng loại')
    don_vi_tinh = fields.Many2one('mdm.dvt', string='ĐVT')
    ton_dau = fields.Float(string='Tồn đầu', digits=(16, 3))
    month_key = fields.Char(string='Tháng', index=True)
    ve_du_kien_don_vi = fields.Float(digits=(16, 3))
    ve_du_kien_bcu = fields.Float(string='Hàng đi đường BCU', digits=(16, 3))
    vt_can_dung = fields.Float(string='Cần dùng', digits=(16, 3))
    ton_cuoi = fields.Float(string='Tồn cuối', digits=(16, 3))
    period_code = fields.Char(string='Kỳ nguồn')
    qty_offset = fields.Char(string='Cột nguồn')

    def action_export_excel(self):
        if self:
            wizards = self.mapped('wizard_id')
        else:
            wizard_id = self.env.context.get('bao_cao_b4_wizard_id')
            wizards = (
                self.env['bao.cao.tong.hop.vat.tu.wizard'].browse(wizard_id)
                if wizard_id else self.env['bao.cao.tong.hop.vat.tu.wizard']
            )
        if len(wizards) != 1:
            raise UserError(_('Vui lòng xuất Excel từ một báo cáo đã mở.'))
        return wizards.action_export_excel()
