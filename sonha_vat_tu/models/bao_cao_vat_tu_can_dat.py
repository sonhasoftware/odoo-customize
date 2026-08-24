# -*- coding: utf-8 -*-
import base64
import io
import json
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from odoo import _, api, fields, models
from odoo.exceptions import UserError

from .bao_cao_ghi_chu import REPORT_VTCD
from .vat_tu_nvl_parse import parse_ten_nvl_specs

REPORT_KIND_KIEM_TRA = 'kiem_tra'
REPORT_KIND_TRINH_LD = 'trinh_ld'

NHOM_INNOX = 'innox'
NHOM_NHUA = 'nhua'
NHOM_TMC = 'tmc'

MA_LINH_VUC_NHOM = {
    'IOXC': NHOM_INNOX,
    'NHUA': NHOM_NHUA,
}

ROW_SUBTOTAL_INNOX = 'subtotal_innox'
ROW_SUBTOTAL_NHUA = 'subtotal_nhua'
ROW_SUBTOTAL_TMC = 'subtotal_tmc'
ROW_SUBTOTAL_GRAND = 'subtotal_grand'
ROW_DETAIL = 'detail'

METRICS_FULL = (
    ('sl_dat_mua', 'SL đặt mua'),
    ('moq', 'MOQ'),
    ('sl_dieu_chuyen', 'SL điều chuyển nội bộ'),
    ('sl_ton_kho', 'SL tồn kho'),
    ('sl_can_dung', 'SL cần dùng'),
    ('vong_quay', 'Vòng quay hàng tồn kho'),
)

METRICS_TRINH_LD = (
    ('sl_dat_mua', 'SL đặt mua'),
    ('moq', 'MOQ'),
)

SUBTOTAL_LABELS = {
    ROW_SUBTOTAL_INNOX: 'Tổng Inox',
    ROW_SUBTOTAL_NHUA: 'Tổng Nhựa',
    ROW_SUBTOTAL_TMC: 'Tổng TMC',
    ROW_SUBTOTAL_GRAND: 'Tổng cộng',
}


def _empty_metrics():
    return {
        'sl_dat_mua': 0.0,
        'moq': 0.0,
        'sl_dieu_chuyen': 0.0,
        'sl_ton_kho': 0.0,
        'sl_can_dung': 0.0,
        'vong_quay': 0.0,
    }


def _metrics_from_bcu(rec):
    return {
        'sl_dat_mua': rec.sl_dat_mua_chot or 0.0,
        'moq': rec.sl_can_mua_theo_moq or 0.0,
        'sl_dieu_chuyen': 0.0,
        'sl_ton_kho': rec.sl_ton_kho_cuoi_ky or 0.0,
        'sl_can_dung': rec.tong_vt_can_dung or 0.0,
        'vong_quay': rec.so_ngay_vong_quay_ton or 0.0,
    }


def _sum_metrics(a, b):
    out = _empty_metrics()
    for key in out:
        out[key] = (a.get(key) or 0.0) + (b.get(key) or 0.0)
    return out


class BaoCaoVtCanDatWizard(models.TransientModel):
    _name = 'bao.cao.vt.can.dat.wizard'
    _description = 'Wizard báo cáo vật tư cần đặt (Biểu 2)'

    period_ids = fields.Many2many(
        'ke.hoach.vat.tu',
        'bao_cao_vtcd_wizard_period_rel',
        'wizard_id',
        'period_id',
        string='Kế hoạch',
        help='Chọn nhiều kỳ cùng tháng — có thể nhiều file cùng đơn vị SX; '
             'báo cáo gom theo đơn vị sản xuất.',
    )
    report_kind = fields.Selection(
        [
            (REPORT_KIND_KIEM_TRA, 'Biểu in kiểm tra vật tư cần đặt'),
            (REPORT_KIND_TRINH_LD, 'Biểu in trình lãnh đạo'),
        ],
        string='Loại báo cáo',
        required=True,
        default=REPORT_KIND_KIEM_TRA,
    )
    period_month = fields.Char(string='Tháng kế hoạch', readonly=True)
    ton_kho_month = fields.Char(string='Tháng tồn kho', readonly=True)
    company_spec_json = fields.Text(string='Công ty (JSON)', readonly=True)
    line_ids = fields.One2many(
        'bao.cao.vt.can.dat.line', 'wizard_id', string='Chi tiết')

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        if 'report_kind' in fields_list and not res.get('report_kind'):
            res['report_kind'] = self.env.context.get(
                'default_report_kind', REPORT_KIND_KIEM_TRA,
            )
        period_id = self.env.context.get('default_period_id')
        if not period_id:
            active_id = self.env.context.get('active_id')
            active_model = self.env.context.get('active_model')
            if active_model == 'ke.hoach.vat.tu' and active_id:
                period_id = active_id
        if period_id and 'period_ids' in fields_list:
            res['period_ids'] = [(6, 0, [period_id])]
        return res

    def _metric_defs(self):
        self.ensure_one()
        if self.report_kind == REPORT_KIND_TRINH_LD:
            return METRICS_TRINH_LD
        return METRICS_FULL

    def _selected_periods(self):
        self.ensure_one()
        if not self.period_ids:
            raise UserError(_('Vui lòng chọn ít nhất một kỳ kế hoạch.'))
        periods = self.period_ids.sorted(
            key=lambda p: (
                (p.company_sx_id.company_code or p.company_sx_id.name or '').upper(),
                p.period_month or '',
                p.code or '',
            )
        )
        months = {(p.period_month or '').strip() for p in periods if p.period_month}
        if len(months) > 1:
            raise UserError(_(
                'Các kỳ đã chọn phải cùng tháng kế hoạch (hiện có: %(months)s).',
                months=', '.join(sorted(months)),
            ))
        for period in periods:
            if not period.company_sx_id:
                raise UserError(_(
                    'Kỳ "%(code)s" chưa có đơn vị sản xuất.',
                    code=period.code or period.display_name,
                ))
        return periods

    @staticmethod
    def _ton_kho_month_label(period_month):
        text = (period_month or '').strip()
        if not text:
            return ''
        try:
            dt = datetime.strptime(text, '%m/%Y')
        except ValueError:
            return text
        month = dt.month + 2
        year = dt.year
        if month > 12:
            month -= 12
            year += 1
        return '%02d/%d' % (month, year)

    def _company_specs(self, periods):
        specs = []
        seen_sx = set()
        for period in periods:
            sx = period.company_sx_id
            if sx.id in seen_sx:
                continue
            seen_sx.add(sx.id)
            code = (sx.company_code or sx.name or '').strip()
            specs.append({
                'period_id': period.id,
                'company_id': sx.id,
                'code': code,
                'label': code,
            })
        return specs

    def _nhom_from_ma_linh_vuc(self, ma_linh_vuc):
        code = (ma_linh_vuc or '').strip().upper()
        return MA_LINH_VUC_NHOM.get(code, NHOM_TMC)

    def _build_detail_rows(self, periods, company_specs):
        Bcu = self.env['kh.dat.vat.tu.bcu'].sudo()
        period_ids = periods.ids
        bcu_lines = Bcu.search([('period_id', 'in', period_ids)])

        ma_codes = {
            (rec.ma_sap or '').strip() for rec in bcu_lines if (rec.ma_sap or '').strip()
        }
        linh_vuc_map = self.env['ma.hang'].get_ma_linh_vuc_map(ma_codes)

        code_by_sx = {
            spec['company_id']: spec['code'] for spec in company_specs
        }
        Period = self.env['ke.hoach.vat.tu'].sudo()
        code_by_period = {
            p.id: code_by_sx.get(p.company_sx_id.id)
            for p in Period.browse(periods.ids)
            if p.company_sx_id.id in code_by_sx
        }

        grouped = {}
        for rec in bcu_lines:
            ma = (rec.ma_sap or '').strip()
            if not ma:
                continue
            bucket = grouped.setdefault(ma, {
                'ma_nvl': ma,
                'ten_nvl': rec.ten_nvl or '',
                'nhom': self._nhom_from_ma_linh_vuc(linh_vuc_map.get(ma)),
                'total': _empty_metrics(),
                'companies': {spec['code']: _empty_metrics() for spec in company_specs},
            })
            if rec.ten_nvl and not bucket['ten_nvl']:
                bucket['ten_nvl'] = rec.ten_nvl
            metrics = _metrics_from_bcu(rec)
            bucket['total'] = _sum_metrics(bucket['total'], metrics)
            comp_code = code_by_period.get(rec.period_id.id)
            if comp_code:
                bucket['companies'][comp_code] = _sum_metrics(
                    bucket['companies'][comp_code], metrics,
                )

        details = []
        for ma in sorted(grouped, key=lambda x: x):
            item = grouped[ma]
            details.append({
                'row_type': ROW_DETAIL,
                'ma_nvl': ma,
                'ten_nvl': item['ten_nvl'],
                'nhom': item['nhom'],
                'metrics_json': json.dumps({
                    'total': item['total'],
                    'companies': item['companies'],
                }, ensure_ascii=False),
            })
        return details

    @staticmethod
    def _aggregate_subtotal(details, nhom_keys):
        total = _empty_metrics()
        companies = {}
        for row in details:
            if row.get('nhom') not in nhom_keys:
                continue
            data = json.loads(row['metrics_json'])
            total = _sum_metrics(total, data.get('total') or {})
            for code, vals in (data.get('companies') or {}).items():
                companies[code] = _sum_metrics(companies.get(code, _empty_metrics()), vals)
        return total, companies

    def _populate_lines(self):
        self.ensure_one()
        periods = self._selected_periods()
        company_specs = self._company_specs(periods)
        self.period_month = periods[0].period_month or ''
        self.ton_kho_month = self._ton_kho_month_label(self.period_month)
        self.company_spec_json = json.dumps(company_specs, ensure_ascii=False)

        details = self._build_detail_rows(periods, company_specs)
        if not details:
            raise UserError(_('Không có dữ liệu B6 cho các kỳ đã chọn.'))

        GhiChu = self.env['bao.cao.ghi.chu'].sudo()
        period_key = GhiChu.period_key_from_periods(periods)
        ghi_chu_map = GhiChu.load_map(REPORT_VTCD, period_key)

        Line = self.env['bao.cao.vt.can.dat.line']
        self.line_ids.unlink()

        seq = 0
        lines = []
        subtotal_defs = (
            (ROW_SUBTOTAL_INNOX, (NHOM_INNOX,)),
            (ROW_SUBTOTAL_NHUA, (NHOM_NHUA,)),
            (ROW_SUBTOTAL_TMC, (NHOM_TMC,)),
            (ROW_SUBTOTAL_GRAND, (NHOM_INNOX, NHOM_NHUA, NHOM_TMC)),
        )
        for row_type, nhom_keys in subtotal_defs:
            total, companies = self._aggregate_subtotal(details, nhom_keys)
            if not any(total.values()) and not any(
                any(v.values()) for v in companies.values()
            ):
                continue
            seq += 1
            lines.append({
                'wizard_id': self.id,
                'sequence': seq,
                'row_type': row_type,
                'label': SUBTOTAL_LABELS[row_type],
                'metrics_json': json.dumps({
                    'total': total,
                    'companies': companies,
                }, ensure_ascii=False),
            })

        for detail in details:
            seq += 1
            specs = parse_ten_nvl_specs(detail['ten_nvl'], nhom=detail.get('nhom'))
            scope = GhiChu.scope_key_vtcd(self.report_kind, detail['ma_nvl'])
            lines.append({
                'wizard_id': self.id,
                'sequence': seq,
                'row_type': ROW_DETAIL,
                'ma_nvl': detail['ma_nvl'],
                'ten_nvl': detail['ten_nvl'],
                'chat_lieu': specs['chat_lieu'],
                'do_bong': specs['do_bong'],
                'do_day': specs['do_day'],
                'kho_rong': specs['kho_rong'],
                'metrics_json': detail['metrics_json'],
                'ghi_chu': ghi_chu_map.get(scope, ''),
            })

        if lines:
            Line.create(lines)

    def _report_context(self):
        self.ensure_one()
        return {
            'bao_cao_vtcd_wizard_id': self.id,
            'bao_cao_vtcd_companies': self.company_spec_json or '[]',
            'bao_cao_vtcd_report_kind': self.report_kind,
            'bao_cao_vtcd_period_month': self.period_month or '',
            'bao_cao_vtcd_ton_kho_month': self.ton_kho_month or '',
        }

    @api.model
    def _format_vtcd_title_period(self, period_month):
        """08/2026 -> 8/2026 cho tiêu đề báo cáo."""
        period = (period_month or '').strip()
        if not period or '/' not in period:
            return period
        month_str, year_str = period.split('/', 1)
        try:
            return '%s/%s' % (int(month_str), year_str.strip())
        except ValueError:
            return period

    def action_open_report(self):
        self.ensure_one()
        self._populate_lines()
        title = self.get_vtcd_report_title()
        return {
            'type': 'ir.actions.act_window',
            'name': title,
            'res_model': 'bao.cao.vt.can.dat.line',
            'view_mode': 'tree',
            'domain': [('wizard_id', '=', self.id)],
            'context': self._report_context(),
        }

    def _metric_columns(self):
        return self._metric_defs()

    def _get_report_company_specs(self):
        self.ensure_one()
        try:
            return json.loads(self.company_spec_json or '[]')
        except (TypeError, ValueError):
            return []

    def _get_report_metric_defs(self):
        self.ensure_one()
        return list(self._metric_defs())

    def get_vtcd_report_title(self):
        self.ensure_one()
        period = self._format_vtcd_title_period(self.period_month)
        if period:
            return 'DUYỆT LƯỢNG MUA VẬT TƯ GIA DỤNG KỲ MUA THÁNG %s' % period
        return 'DUYỆT LƯỢNG MUA VẬT TƯ GIA DỤNG'

    def get_vtcd_metric_date_label(self, metric_key):
        self.ensure_one()
        mapping = {
            'sl_dat_mua': self.period_month or '',
            'moq': self.period_month or '',
            'sl_dieu_chuyen': '',
            'sl_ton_kho': self.ton_kho_month or '',
            'sl_can_dung': self.period_month or '',
            'vong_quay': '',
        }
        return mapping.get(metric_key, '')

    @api.model
    def format_vtcd_qty(self, value):
        """Định dạng số gọn — dùng chung view / Excel / PDF."""
        if value in (None, '', False):
            return ''
        try:
            num = float(value)
        except (TypeError, ValueError):
            return str(value)
        if abs(num - round(num)) < 1e-6:
            return str(int(round(num)))
        return ('%.2f' % num).rstrip('0').rstrip('.')

    def _vtcd_qty_cell_value(self, value):
        """Giá trị số cho ô Excel (None nếu trống)."""
        if value in (None, '', False):
            return None
        try:
            num = float(value)
        except (TypeError, ValueError):
            return value
        if abs(num - round(num)) < 1e-6:
            return int(round(num))
        return round(num, 2)

    def _write_vtcd_excel_footer(self, ws, start_row, max_col):
        """Chân trang Excel giống PDF: Người lập | BCU | Ban lãnh đạo."""
        if max_col < 1:
            return
        now = fields.Datetime.context_timestamp(self, datetime.utcnow())
        date_line = 'Ngày %s tháng %s năm %s' % (now.day, now.month, now.year)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        bold = Font(bold=True)

        col1_end = max(max_col // 3, 1)
        col2_end = max((max_col * 2) // 3, col1_end + 1)
        if col2_end >= max_col:
            col2_end = max_col - 1 if max_col > 1 else max_col

        row = start_row + 2
        ws.merge_cells(
            start_row=row, start_column=col2_end + 1,
            end_row=row, end_column=max_col,
        )
        date_cell = ws.cell(row=row, column=col2_end + 1, value=date_line)
        date_cell.alignment = center

        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col1_end)
        c1 = ws.cell(row=row, column=1, value='Người lập')
        c1.font = bold
        c1.alignment = center

        ws.merge_cells(
            start_row=row, start_column=col1_end + 1,
            end_row=row, end_column=col2_end,
        )
        c2 = ws.cell(row=row, column=col1_end + 1, value='BAN CUNG ỨNG – ĐẤU THẦU')
        c2.font = bold
        c2.alignment = center

        ws.merge_cells(
            start_row=row, start_column=col2_end + 1,
            end_row=row, end_column=max_col,
        )
        c3 = ws.cell(row=row, column=col2_end + 1, value='BAN LÃNH ĐẠO PHÊ DUYỆT')
        c3.font = bold
        c3.alignment = center

        row += 1
        ws.row_dimensions[row].height = 54

    def _prepare_vtcd_export(self):
        """Chuẩn bị dữ liệu dùng chung cho Excel và PDF."""
        self.ensure_one()
        if not self.line_ids:
            self._populate_lines()
        lines = self.line_ids.sorted(key=lambda l: (l.sequence, l.id))
        if not lines:
            raise UserError(_('Không có dữ liệu để xuất.'))

        try:
            company_specs = json.loads(self.company_spec_json or '[]')
        except (TypeError, ValueError):
            raise UserError(_('Không xác định được cột công ty.'))

        metrics = self._metric_columns()
        date_labels = {
            'sl_dat_mua': self.period_month or '',
            'moq': self.period_month or '',
            'sl_dieu_chuyen': '',
            'sl_ton_kho': self.ton_kho_month or '',
            'sl_can_dung': self.period_month or '',
            'vong_quay': '',
        }
        date_map = {key: date_labels.get(key, '') for key, _label in metrics}

        report_title = self.get_vtcd_report_title()
        if self.report_kind == REPORT_KIND_TRINH_LD:
            file_stem = 'ChiTietVatTuCanMua'
            sheet_title = 'Chi tiet VT can mua'
        else:
            file_stem = 'ChiTietVatTuCanIn'
            sheet_title = 'Chi tiet VT can in'

        return {
            'lines': lines,
            'company_specs': company_specs,
            'metrics': metrics,
            'report_title': report_title,
            'file_stem': file_stem,
            'sheet_title': sheet_title,
            'period_month': self.period_month or '',
            'date_map': date_map,
        }

    def action_export_excel(self):
        self.ensure_one()
        data = self._prepare_vtcd_export()

        lines = data['lines']
        company_specs = data['company_specs']
        metrics = data['metrics']
        report_title = data['report_title']
        sheet_title = data['sheet_title']
        file_name = data['file_stem'] + '.xlsx'
        date_map_full = data['date_map']
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title

        header_row1 = 3
        header_row2 = 4
        header_row3 = 5
        data_row = 6

        fixed_headers = (
            'Mã NVL', 'Tên NVL', 'Chất liệu', 'Độ bóng', 'Độ dày', 'Khổ rộng',
        )
        col = 1
        for label in fixed_headers:
            ws.merge_cells(
                start_row=header_row1, start_column=col,
                end_row=header_row3, end_column=col,
            )
            cell = ws.cell(row=header_row1, column=col, value=label)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            col += 1

        def write_metric_block(start_col, group_label, date_map):
            span = len(metrics)
            ws.merge_cells(
                start_row=header_row1, start_column=start_col,
                end_row=header_row1, end_column=start_col + span - 1,
            )
            gcell = ws.cell(row=header_row1, column=start_col, value=group_label)
            gcell.font = Font(bold=True)
            gcell.alignment = Alignment(horizontal='center')
            c = start_col
            for key, label in metrics:
                ws.cell(row=header_row2, column=c, value=label).font = Font(bold=True)
                date_label = date_map.get(key, '')
                ws.cell(row=header_row3, column=c, value=date_label)
                c += 1
            return start_col + span

        col = write_metric_block(col, 'Tổng', date_map_full)
        for spec in company_specs:
            col = write_metric_block(col, spec.get('label') or spec.get('code'), date_map_full)

        ws.merge_cells(
            start_row=header_row1, start_column=col,
            end_row=header_row3, end_column=col,
        )
        ws.cell(row=header_row1, column=col, value='Ghi chú').font = Font(bold=True)
        ghi_chu_col = col
        max_col = col

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        title_cell = ws.cell(row=1, column=1, value=report_title)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True,
        )

        row_idx = data_row
        for line in lines:
            if line.row_type != ROW_DETAIL:
                cell = ws.cell(row=row_idx, column=6, value=line.label or '')
                cell.font = Font(bold=True)
            else:
                ws.cell(row=row_idx, column=1, value=line.ma_nvl)
                ws.cell(row=row_idx, column=2, value=line.ten_nvl)
                ws.cell(row=row_idx, column=3, value=line.chat_lieu)
                ws.cell(row=row_idx, column=4, value=line.do_bong)
                ws.cell(row=row_idx, column=5, value=line.do_day)
                ws.cell(row=row_idx, column=6, value=line.kho_rong)

            col_idx = 7
            payload = line._metrics_payload()
            for key, _label in metrics:
                ws.cell(
                    row=row_idx, column=col_idx,
                    value=self._vtcd_qty_cell_value(payload['total'].get(key)),
                )
                col_idx += 1
            for spec in company_specs:
                code = spec.get('code')
                comp = (payload.get('companies') or {}).get(code) or _empty_metrics()
                for key, _label in metrics:
                    ws.cell(
                        row=row_idx, column=col_idx,
                        value=self._vtcd_qty_cell_value(comp.get(key)),
                    )
                    col_idx += 1
            ws.cell(row=row_idx, column=ghi_chu_col, value=line.ghi_chu or '')
            row_idx += 1

        self._write_vtcd_excel_footer(ws, row_idx, max_col)

        fixed_col_widths = {
            1: 12,   # Mã NVL
            2: 36,   # Tên NVL
            3: 10,   # Chất liệu
            4: 10,   # Độ bóng
            5: 10,   # Độ dày
            6: 12,   # Khổ rộng
        }
        metric_col_width = 18
        for col_idx, width in fixed_col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        first_metric_col = len(fixed_col_widths) + 1
        for col_idx in range(first_metric_col, ghi_chu_col):
            ws.column_dimensions[get_column_letter(col_idx)].width = metric_col_width
        ws.column_dimensions[get_column_letter(max_col)].width = 28

        output = io.BytesIO()
        wb.save(output)
        attachment = self.env['ir.attachment'].sudo().create({
            'name': file_name,
            'type': 'binary',
            'datas': base64.b64encode(output.getvalue()),
            'mimetype': (
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            ),
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }

    def action_export_pdf(self):
        """Xuất PDF."""
        self.ensure_one()
        if self.report_kind != REPORT_KIND_TRINH_LD:
            raise UserError(_('Chỉ báo cáo trình lãnh đạo mới xuất PDF.'))
        if not self.line_ids:
            self._populate_lines()
        return self.env.ref(
            'sonha_vat_tu.action_report_bao_cao_vtcd_trinh_ld'
        ).report_action(self)


class BaoCaoVtCanDatLine(models.TransientModel):
    _name = 'bao.cao.vt.can.dat.line'
    _inherit = ['bao.cao.ghi.chu.line.mixin']
    _description = 'Dòng báo cáo vật tư cần đặt'
    _order = 'sequence, id'

    wizard_id = fields.Many2one(
        'bao.cao.vt.can.dat.wizard', ondelete='cascade', index=True)
    sequence = fields.Integer(default=10)
    row_type = fields.Selection(
        [
            (ROW_SUBTOTAL_INNOX, 'Tổng Inox'),
            (ROW_SUBTOTAL_NHUA, 'Tổng Nhựa'),
            (ROW_SUBTOTAL_TMC, 'Tổng TMC'),
            (ROW_SUBTOTAL_GRAND, 'Tổng cộng'),
            (ROW_DETAIL, 'Chi tiết'),
        ],
        string='Loại dòng',
        default=ROW_DETAIL,
        required=True,
    )
    label = fields.Char(string='Nhãn tổng')
    ma_nvl = fields.Char(string='Mã NVL')
    ten_nvl = fields.Char(string='Tên NVL')
    chat_lieu = fields.Char(string='Chất liệu')
    do_bong = fields.Char(string='Độ bóng')
    do_day = fields.Char(string='Độ dày')
    kho_rong = fields.Char(string='Khổ rộng')
    metrics_json = fields.Text(string='Số liệu', readonly=True)
    ghi_chu = fields.Text(string='Ghi chú')

    report_kind = fields.Selection(
        related='wizard_id.report_kind', readonly=True)

    def _sync_ghi_chu_to_master(self):
        GhiChu = self._ghi_chu_master()
        for rec in self.filtered(lambda r: r.row_type == ROW_DETAIL and r.ma_nvl):
            wizard = rec.wizard_id
            if not wizard:
                continue
            period_key = rec._ghi_chu_period_key(wizard)
            scope = GhiChu.scope_key_vtcd(wizard.report_kind, rec.ma_nvl)
            GhiChu.upsert_note(REPORT_VTCD, period_key, scope, rec.ghi_chu)

    def _metrics_payload(self):
        self.ensure_one()
        try:
            data = json.loads(self.metrics_json or '{}')
        except (TypeError, ValueError):
            data = {}
        if not isinstance(data, dict):
            data = {}
        data.setdefault('total', _empty_metrics())
        data.setdefault('companies', {})
        return data

    def action_export_excel(self):
        if self:
            wizards = self.mapped('wizard_id')
        else:
            wizard_id = self.env.context.get('bao_cao_vtcd_wizard_id')
            wizards = (
                self.env['bao.cao.vt.can.dat.wizard'].browse(wizard_id)
                if wizard_id else self.env['bao.cao.vt.can.dat.wizard']
            )
        if len(wizards) != 1:
            raise UserError(_('Vui lòng xuất Excel từ một báo cáo đã mở.'))
        return wizards.action_export_excel()

    def action_print_pdf(self):
        if self:
            wizards = self.mapped('wizard_id')
        else:
            wizard_id = self.env.context.get('bao_cao_vtcd_wizard_id')
            wizards = (
                self.env['bao.cao.vt.can.dat.wizard'].browse(wizard_id)
                if wizard_id else self.env['bao.cao.vt.can.dat.wizard']
            )
        if len(wizards) != 1:
            raise UserError(_('Vui lòng xuất PDF từ một báo cáo đã mở.'))
        wizard = wizards
        if wizard.report_kind != REPORT_KIND_TRINH_LD:
            raise UserError(_('Chỉ báo cáo trình lãnh đạo mới xuất PDF.'))
        return wizard.action_export_pdf()
