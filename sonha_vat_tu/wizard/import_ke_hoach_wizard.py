# -*- coding: utf-8 -*-
import base64
import io
import re
import warnings
from datetime import date

from markupsafe import Markup
from openpyxl import load_workbook

from odoo import fields, models, _
from odoo.exceptions import UserError


class ImportKeHoachWizard(models.TransientModel):
    _name = 'import.ke.hoach.wizard'
    _description = 'Import ke hoach tu Excel'

    period_id = fields.Many2one(
        'ke.hoach.vat.tu', string='Ky',
        default=lambda self: self.env.context.get('active_id'))
    import_type = fields.Selection([
        ('business', 'Kế hoạch kinh doanh'),
        ('production', 'Kế hoạch sản xuất'),
    ], string='Loại import', required=True,
        default=lambda self: self.env.context.get('default_import_type', 'business'))
    file_data = fields.Binary(string='File Excel')
    file_name = fields.Char(string='Ten file')

    MONTH_RE = re.compile(r'(\d{1,2})\s*[/\-]\s*(\d{4})')
    _IMPORT_CTX = {'is_importing': True, 'tracking_disable': True}
    _QTY_FIELDS = ('qty_t0', 'qty_t1', 'qty_t2', 'qty_t3')
    _WRITE_FIELDS_KD = ('ma_hang', 'qty_t0', 'qty_t1', 'qty_t2', 'qty_t3')
    _WRITE_FIELDS_SX = ('ma_hang', 'qty_t0', 'qty_t1', 'qty_t2', 'qty_t3')
    _PLAN_HEADERS = ['Đơn vị', 'Ngành hàng', 'Tên hàng', 'Mã hàng', 'Mã']

    def _parse_month_header(self, label):
        if not label:
            return None
        m = self.MONTH_RE.search(str(label))
        if not m:
            return None
        try:
            month = int(m.group(1))
            year = int(m.group(2))
            date(year, month, 1)
            return f'{month:02d}/{year}'
        except ValueError:
            return None

    def _read_workbook(self):
        if not self.file_data:
            raise UserError(_('Vui lòng chọn file Excel.'))
        try:
            data = base64.b64decode(self.file_data)
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', message='Data Validation extension', category=UserWarning)
                wb = load_workbook(io.BytesIO(data), data_only=True)
        except Exception as exc:
            raise UserError(_('Không đọc được file Excel: %s') % exc)
        rows = [tuple(row) for row in wb.active.iter_rows(values_only=True)]
        if not rows:
            raise UserError(_('File rỗng.'))
        return rows

    def _norm_text(self, value):
        return str(value or '').strip()

    def _validate_metadata(self, rows):
        if len(rows) < 6:
            raise UserError(_('File Excel không đúng mẫu. Vui lòng tải lại template từ kỳ kế hoạch đang mở.'))

        meta = {}
        for row_idx in range(3):
            row = rows[row_idx] if row_idx < len(rows) else ()
            label = self._norm_text(row[0] if len(row) > 0 else '')
            value = self._norm_text(row[1] if len(row) > 1 else '')
            meta[label.lower()] = value

        code = meta.get('mã') or meta.get('ma')
        month = meta.get('tháng bắt đầu') or meta.get('thang bat dau')

        errors = []
        if not code:
            errors.append(_('Mã kỳ không được để trống. Vui lòng kiểm tra ô B1.'))
        elif code != (self.period_id.code or ''):
            errors.append(_('Mã kỳ trong file là "%s", không đúng với kỳ đang mở "%s".') % (code, self.period_id.code or ''))

        if not month:
            errors.append(_('Tháng bắt đầu không được để trống. Vui lòng kiểm tra ô B2.'))
        elif month != (self.period_id.period_month or ''):
            errors.append(_('Tháng bắt đầu trong file là "%s", không đúng với kỳ đang mở "%s".') % (month, self.period_id.period_month or ''))

        self._raise_errors(errors)

    def _prepare_rows(self, rows, import_type):
        header_row_idx = 5
        header = [str(c).strip() if c is not None else '' for c in rows[header_row_idx]]
        for col_idx, expected in enumerate(self._PLAN_HEADERS):
            actual = header[col_idx] if col_idx < len(header) else ''
            if actual.lower() != expected.lower():
                raise UserError(_(
                    'File Excel không đúng mẫu: cột %s phải là "%s" (đang là "%s"). '
                    'Vui lòng tải lại template từ kỳ kế hoạch đang mở.'
                ) % (col_idx + 1, expected, actual or '—'))
        month_cols = []
        horizon_months = self.period_id._get_horizon_months()
        month_offset_map = {month: idx for idx, month in enumerate(horizon_months)}
        month_start_col = 5
        for idx, label in enumerate(header):
            month_key = self._parse_month_header(label)
            if month_key and idx >= month_start_col and month_key in month_offset_map:
                month_cols.append((idx, month_key, month_offset_map[month_key]))
        if not month_cols:
            raise UserError(_('Không tìm thấy cột tháng hợp lệ trong bảng dữ liệu. Vui lòng kiểm tra dòng tiêu đề số 6.'))
        return header, month_cols, header_row_idx + 1

    def _build_company_lookup(self):
        by_code = {}
        by_name = {}
        for company in self.env['res.company'].sudo().search([]):
            code = self._norm_text(company.company_code)
            name = self._norm_text(company.name)
            if code:
                by_code[code.lower()] = company
            if name:
                by_name[name.lower()] = company
        return by_code, by_name

    def _resolve_company_cached(self, raw, row_idx, errors, company_lookup):
        text = self._norm_text(raw)
        if not text:
            errors.append(_('Dòng %d: thiếu Đơn vị.') % row_idx)
            return self.env['res.company']
        by_code, by_name = company_lookup
        company = by_code.get(text.lower()) or by_name.get(text.lower())
        if not company:
            errors.append(_('Dòng %d: không tìm thấy Đơn vị "%s".') % (row_idx, text))
            return self.env['res.company']
        return company

    def _parse_qty_row(self, row, row_idx, month_cols, errors):
        qty_by_offset = {0: 0.0, 1: 0.0, 2: 0.0, 3: 0.0}
        for col_idx, month_key, offset in month_cols:
            raw_qty = row[col_idx]
            if raw_qty in (None, ''):
                qty = 0.0
            else:
                try:
                    qty = float(raw_qty)
                except (TypeError, ValueError):
                    errors.append(_('Dòng %d, tháng %s: "%s" không phải số.') % (row_idx, month_key, raw_qty))
                    qty = 0.0
            qty_by_offset[offset] = qty
        return qty_by_offset

    def _collect_import_sap_codes(self, rows, header, data_start_idx, import_type):
        codes = set()
        ma_col = 4
        for row in rows[data_start_idx:]:
            if not row or not any(c not in (None, '') for c in row):
                continue
            row = list(row) + [None] * (len(header) - len(row))
            ma_sap = str(row[ma_col]).strip() if row[ma_col] not in (None, '') else ''
            if ma_sap:
                codes.add(ma_sap)
        return codes

    @staticmethod
    def _row_changed(existing, vals, fields):
        return any((existing[f] or 0.0) != (vals.get(f) or 0.0) for f in fields if f.startswith('qty_')) or any(
            (getattr(existing, f, '') or '') != (vals.get(f) or '')
            for f in fields if not f.startswith('qty_')
        )

    def _validate_business_row(self, row_idx, row, errors, company_lookup, mdm_codes):
        company_rec = self._resolve_company_cached(row[0], row_idx, errors, company_lookup)
        if not company_rec:
            return None

        ma_hang_raw = str(row[3]).strip() if row[3] not in (None, '') else ''
        ma_sap = str(row[4]).strip() if row[4] not in (None, '') else ''

        if not ma_sap:
            errors.append(_('Dòng %d: thiếu Mã.') % row_idx)
            return None
        if ma_sap not in mdm_codes:
            errors.append(_(
                'Dòng %d: Mã "%s" không có trong MDM (mdm.tong.hop.line).'
            ) % (row_idx, ma_sap))
            return None

        return {
            'company_id': company_rec.id,
            'ma_hang': ma_hang_raw,
            'ma_sap': ma_sap,
        }

    def _validate_production_row(self, row_idx, row, errors, company_lookup, mdm_codes):
        company_rec = self._resolve_company_cached(row[0], row_idx, errors, company_lookup)
        if not company_rec:
            return None

        ma_hang_raw = str(row[3]).strip() if row[3] not in (None, '') else ''
        ma_sap = str(row[4]).strip() if row[4] not in (None, '') else ''

        if not ma_sap:
            errors.append(_('Dòng %d: thiếu Mã.') % row_idx)
            return None
        if ma_sap not in mdm_codes:
            errors.append(_(
                'Dòng %d: Mã "%s" không có trong MDM (mdm.tong.hop.line).'
            ) % (row_idx, ma_sap))
            return None

        return {
            'company_id': company_rec.id,
            'ma_hang': ma_hang_raw,
            'ma_sap': ma_sap,
        }

    def _raise_errors(self, errors):
        if not errors:
            return
        shown = errors[:80]
        msg = '\n'.join('- %s' % error for error in shown)
        if len(errors) > 80:
            msg += _('\n... còn %d lỗi khác.') % (len(errors) - 80)
        raise UserError(_('File Excel có lỗi, chưa ghi dữ liệu:\n%s') % msg)

    def action_import(self):
        self.ensure_one()
        if not self.period_id:
            raise UserError(_('Thiếu kỳ kế hoạch.'))
        if self.period_id.state != 'ke_hoach':
            raise UserError(_('Kỳ kế hoạch đã sang bước sau, không thể import lại kế hoạch.'))
        is_supply = self.env.user.has_group('sonha_vat_tu.group_ban_cung_ung_vat_tu')
        is_department = self.env.user.has_group('sonha_vat_tu.group_bo_phan_vat_tu')
        is_manager = self.env.user.has_group('sonha_vat_tu.group_truong_bo_phan_vat_tu')
        if self.import_type == 'business' and not (is_supply or is_manager):
            raise UserError(_('Bạn không có quyền import kế hoạch kinh doanh.'))
        if self.import_type == 'production' and not (is_department or is_manager):
            raise UserError(_('Bạn không có quyền import kế hoạch sản xuất.'))

        rows = self._read_workbook()
        self._validate_metadata(rows)
        header, month_cols, data_start_idx = self._prepare_rows(rows, self.import_type)

        DuLieu = self.env['du.lieu.tong.hop.vat.tu'].sudo()
        if self.import_type == 'business':
            count = DuLieu.run_period_bulk(
                self.period_id.id,
                ['kd', 'sx'],
                lambda: self._import_business(rows, header, month_cols, data_start_idx),
            )
            label = 'kế hoạch kinh doanh'
        else:
            count = DuLieu.run_period_bulk(
                self.period_id.id,
                ['sx'],
                lambda: self._import_production(rows, header, month_cols, data_start_idx),
            )
            label = 'kế hoạch sản xuất'

        attachment = self.env['ir.attachment'].sudo().create({
            'name': self.file_name or 'ke_hoach_import.xlsx',
            'type': 'binary',
            'datas': self.file_data,
            'res_model': 'ke.hoach.vat.tu',
            'res_id': self.period_id.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        scope = 'kd' if self.import_type == 'business' else 'sx'
        self.period_id.with_context(vat_tu_chatter_scope=scope).message_post(body=Markup(
            '<p><b>Đã import %s dòng %s từ file %s.</b></p>' %
            (count, label, self.file_name or '-')
        ), attachment_ids=[attachment.id])
        view_xmlid = (
            'sonha_vat_tu.view_ke_hoach_vat_tu_form_kd'
            if self.import_type == 'business'
            else 'sonha_vat_tu.view_ke_hoach_vat_tu_form_sx'
        )
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'ke.hoach.vat.tu',
            'res_id': self.period_id.id,
            'view_mode': 'form',
            'views': [(self.env.ref(view_xmlid).sudo().id, 'form')],
            'context': {'vat_tu_chatter_scope': scope},
            'target': 'current',
        }

    def _import_business(self, rows, header, month_cols, data_start_idx):
        Plan = self.env['ke.hoach.kinh.doanh'].sudo()
        vals_list = []
        errors = []
        seen = set()
        company_lookup = self._build_company_lookup()
        mdm_codes = self.env['ma.hang'].get_mdm_sap_codes_set(
            self._collect_import_sap_codes(rows, header, data_start_idx, 'business'),
        )

        for row_idx, row in enumerate(rows[data_start_idx:], start=data_start_idx + 1):
            if not row or not any(c not in (None, '') for c in row):
                continue
            row = list(row) + [None] * (len(header) - len(row))
            base_vals = self._validate_business_row(row_idx, row, errors, company_lookup, mdm_codes)
            if not base_vals:
                continue

            qty_by_offset = self._parse_qty_row(row, row_idx, month_cols, errors)
            row_key = (base_vals['company_id'], base_vals['ma_sap'])
            if row_key in seen:
                errors.append(_(
                    'Dòng %d: trùng Đơn vị + Mã=%s trong file.'
                ) % (row_idx, base_vals['ma_sap']))
                continue
            seen.add(row_key)

            vals_list.append({
                **base_vals,
                'period_id': self.period_id.id,
                'qty_t0': qty_by_offset[0],
                'qty_t1': qty_by_offset[1],
                'qty_t2': qty_by_offset[2],
                'qty_t3': qty_by_offset[3],
            })

        self._raise_errors(errors)

        existing_map = {
            (line.company_id.id, line.ma_sap): line
            for line in Plan.search([('period_id', '=', self.period_id.id)])
        }
        to_create = []
        to_update = []
        for vals in vals_list:
            existing = existing_map.get((vals['company_id'], vals['ma_sap']))
            if existing:
                if self._row_changed(existing, vals, self._WRITE_FIELDS_KD):
                    to_update.append((existing.id, {f: vals[f] for f in self._WRITE_FIELDS_KD}))
            else:
                to_create.append(vals)

        if to_update:
            Plan._sql_bulk_import_update(to_update)
        if to_create:
            Plan.with_context(**self._IMPORT_CTX).create(to_create)
        if vals_list:
            self.period_id._sync_production_from_business()
        return len(to_create) + len(to_update)

    def _import_production(self, rows, header, month_cols, data_start_idx):
        Plan = self.env['ke.hoach.san.xuat'].sudo()
        vals_list = []
        errors = []
        company_sx = self.env.company
        if company_sx.company_code not in ('BNH', 'SSP'):
            raise UserError(_(
                'Công ty hiện tại không phải công ty sản xuất BNH/SSP. '
                'Vui lòng chọn đúng công ty trước khi import kế hoạch sản xuất.'
            ))

        company_lookup = self._build_company_lookup()
        mdm_codes = self.env['ma.hang'].get_mdm_sap_codes_set(
            self._collect_import_sap_codes(rows, header, data_start_idx, 'production'),
        )
        seen = set()

        for row_idx, row in enumerate(rows[data_start_idx:], start=data_start_idx + 1):
            if not row or not any(c not in (None, '') for c in row):
                continue
            row = list(row) + [None] * (len(header) - len(row))
            base_vals = self._validate_production_row(
                row_idx, row, errors, company_lookup, mdm_codes,
            )
            if not base_vals:
                continue

            qty_by_offset = self._parse_qty_row(row, row_idx, month_cols, errors)
            key = (base_vals['company_id'], base_vals['ma_sap'])
            if key in seen:
                errors.append(_(
                    'Dòng %d: trùng Đơn vị + Mã=%s trong file.'
                ) % (row_idx, base_vals['ma_sap']))
                continue
            seen.add(key)

            vals_list.append({
                **base_vals,
                'period_id': self.period_id.id,
                'company_sx_id': company_sx.id,
                'qty_t0': qty_by_offset[0],
                'qty_t1': qty_by_offset[1],
                'qty_t2': qty_by_offset[2],
                'qty_t3': qty_by_offset[3],
            })

        self._raise_errors(errors)

        business_keys = {
            (row['company_id'], row['ma_sap'])
            for row in self.env['ke.hoach.kinh.doanh'].sudo().search_read(
                [('period_id', '=', self.period_id.id)], ['company_id', 'ma_sap'],
            )
            if row.get('ma_sap') and row.get('company_id')
        }
        imported_keys = {(v['company_id'], v['ma_sap']) for v in vals_list}
        missing = sorted(business_keys - imported_keys)
        if missing:
            for company_id, ma_sap in missing[:20]:
                company = self.env['res.company'].browse(company_id)
                errors.append(_(
                    'Thiếu dòng kế hoạch kinh doanh Đơn vị=%s, Mã=%s. '
                    'Nếu không sản xuất, giữ dòng và nhập Số lượng = 0.'
                ) % (company.company_code or company.name, ma_sap))
            if len(missing) > 20:
                errors.append(_('... còn %d dòng kế hoạch kinh doanh bị thiếu.') % (len(missing) - 20))
        self._raise_errors(errors)

        existing_map = {
            (line.company_id.id, line.ma_sap): line
            for line in Plan.search([('period_id', '=', self.period_id.id)])
        }
        imported_keys = {(v['company_id'], v['ma_sap']) for v in vals_list}
        to_delete = Plan.browse([
            line.id for key, line in existing_map.items() if key not in imported_keys
        ])
        if to_delete:
            to_delete.with_context(**self._IMPORT_CTX).unlink()

        to_create = []
        to_update = []
        for vals in vals_list:
            existing = existing_map.get((vals['company_id'], vals['ma_sap']))
            if existing:
                if self._row_changed(existing, vals, self._WRITE_FIELDS_SX):
                    to_update.append((existing.id, {f: vals[f] for f in self._WRITE_FIELDS_SX}))
            else:
                to_create.append(vals)

        if to_update:
            Plan._sql_bulk_import_update(to_update)
        if to_create:
            Plan.with_context(**self._IMPORT_CTX).create(to_create)
        return len(vals_list)
