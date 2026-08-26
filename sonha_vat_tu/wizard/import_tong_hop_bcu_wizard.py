# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font

from odoo import _, api, fields, models
from odoo.exceptions import UserError

TEMPLATE_WIDTHS = [16, 36, 14, 14, 14, 14, 14, 14, 14, 14]


class ImportTongHopBcuWizard(models.TransientModel):
    """Import hàng đi đường BCU — ghi vat_tu_di_duong (loai=bcu) + B6."""
    _name = 'import.tong.hop.bcu.wizard'
    _description = 'Import hàng đi đường BCU (Tổng hợp KH vật tư BCU)'
    _inherit = ['vat.tu.excel.mixin']

    TEMPLATE_SHEET_NAME = 'Hang di duong BCU'
    META_ROW = 1
    HEADER_ROW = 4
    SUBHEADER_ROW = 5
    DATA_START_ROW = 6
    META_FONT_SIZE = 13
    COL_MA_NVL, COL_TEN_NVL = 0, 1

    period_id = fields.Many2one(
        'ke.hoach.vat.tu',
        string='Kỳ kế hoạch',
        required=True,
        default=lambda self: self.env.context.get('default_period_id') or self.env.context.get('active_id'),
    )
    file_data = fields.Binary(string='File Excel')
    file_name = fields.Char(string='Tên file')

    def _month_keys(self):
        self.ensure_one()
        months = self.period_id._get_horizon_months()
        if len(months) != 4:
            raise UserError(_('Kỳ kế hoạch chưa xác định được 4 tháng tính toán.'))
        return months

    def _month_col_sl(self, offset):
        return 2 + offset * 2

    def _month_col_dg(self, offset):
        return 3 + offset * 2

    def _get_b6_lines(self):
        self.ensure_one()
        return self.env['kh.dat.vat.tu.bcu'].sudo().search([
            ('period_id', '=', self.period_id.id),
        ], order='ma_sap, id')

    def _parse_doc_code_from_sheet(self, ws):
        label = str(ws.cell(row=self.META_ROW, column=1).value or '').strip()
        value = str(ws.cell(row=self.META_ROW, column=2).value or '').strip()
        if label == _('Số chứng từ') and value == _('Mã NVL'):
            return 'legacy_column'
        if label == _('Số chứng từ') and value:
            return value
        return False

    def _detect_import_format(self, ws):
        sub_col3 = str(ws.cell(row=self.SUBHEADER_ROW, column=3).value or '').strip()
        if sub_col3 == _('Số lượng'):
            return 'sl_dg'
        header_col3 = str(ws.cell(row=self.HEADER_ROW, column=3).value or '').strip()
        if header_col3.startswith(_('Tháng')):
            return 'legacy_sl_only'
        return 'sl_dg'

    def _resolve_data_start_row(self, ws, fmt):
        if fmt == 'legacy_sl_only':
            return self.HEADER_ROW + 1
        return self.DATA_START_ROW

    def action_download_template(self):
        self.ensure_one()
        if not self.period_id.code:
            raise UserError(_('Kỳ kế hoạch chưa có số chứng từ.'))

        month_keys = self._month_keys()
        wb = Workbook()
        ws = wb.active
        ws.title = self.TEMPLATE_SHEET_NAME
        ws.cell(row=self.META_ROW, column=1, value=_('Số chứng từ'))
        ws.cell(row=self.META_ROW, column=2, value=self.period_id.code)

        ws.cell(row=self.HEADER_ROW, column=1, value=_('Mã NVL'))
        ws.cell(row=self.HEADER_ROW, column=2, value=_('Tên NVL'))
        ws.merge_cells(
            start_row=self.HEADER_ROW, start_column=1,
            end_row=self.SUBHEADER_ROW, end_column=1,
        )
        ws.merge_cells(
            start_row=self.HEADER_ROW, start_column=2,
            end_row=self.SUBHEADER_ROW, end_column=2,
        )

        for offset, month in enumerate(month_keys):
            col_sl = self._month_col_sl(offset) + 1
            ws.merge_cells(
                start_row=self.HEADER_ROW, start_column=col_sl,
                end_row=self.HEADER_ROW, end_column=col_sl + 1,
            )
            ws.cell(row=self.HEADER_ROW, column=col_sl, value=_('Tháng %s') % month)
            ws.cell(row=self.SUBHEADER_ROW, column=col_sl, value=_('Số lượng'))
            ws.cell(row=self.SUBHEADER_ROW, column=col_sl + 1, value=_('Đơn giá'))

        row_idx = self.DATA_START_ROW
        for line in self._get_b6_lines():
            ws.cell(row=row_idx, column=1, value=line.ma_sap or '')
            ws.cell(row=row_idx, column=2, value=line.ten_nvl or '')
            for offset in range(4):
                ws.cell(
                    row=row_idx, column=self._month_col_sl(offset) + 1,
                    value=line['ve_du_kien_bcu_t%d' % offset] or 0,
                )
                ws.cell(
                    row=row_idx, column=self._month_col_dg(offset) + 1,
                    value=line['ve_du_kien_bcu_dg_t%d' % offset] or 0,
                )
            row_idx += 1

        max_col = 2 + 4 * 2
        meta_font = Font(name='Times New Roman', size=self.META_FONT_SIZE, bold=True)
        ws.cell(row=self.META_ROW, column=1).font = meta_font
        ws.cell(row=self.META_ROW, column=2).font = meta_font
        ws.row_dimensions[self.META_ROW].height = 24
        self._style_excel_header(ws, max_col, header_row=self.HEADER_ROW)
        self._style_excel_header(ws, max_col, header_row=self.SUBHEADER_ROW)
        self._style_excel_body(ws, max_col, self.DATA_START_ROW, row_idx - 1)
        self._set_excel_widths(ws, TEMPLATE_WIDTHS[:max_col])

        return self._xlsx_download_action(
            wb, 'Template_hang_di_duong_BCU_%s.xlsx' % (self.period_id.code or 'ky'))

    def _parse_row_sl_dg(self, row, row_number, existing_map, errors):
        ma_nvl = self._normalize_ma_nvl(self._cell(row, self.COL_MA_NVL))
        if not ma_nvl:
            errors.append(_('Dòng %d: thiếu Mã NVL.') % row_number)
            return None

        line_id = existing_map.get(ma_nvl)
        if not line_id:
            errors.append(
                _('Dòng %d: Mã NVL "%s" không có trong Tổng hợp KH vật tư BCU của kỳ này.')
                % (row_number, ma_nvl)
            )
            return None

        sls, dgs, gts = [], [], []
        for offset in range(4):
            try:
                sl = self._parse_number(
                    self._cell(row, self._month_col_sl(offset)), default=0.0)
                dg = self._parse_number(
                    self._cell(row, self._month_col_dg(offset)), default=0.0)
            except UserError as exc:
                errors.append(_('Dòng %d: %s') % (row_number, exc.args[0]))
                return None
            if sl < 0 or dg < 0:
                errors.append(_('Dòng %d: Số lượng/Đơn giá không được âm.') % row_number)
                return None
            sls.append(sl)
            dgs.append(dg)
            gts.append(sl * dg)

        return {
            'line_id': line_id,
            'ma_nvl': ma_nvl,
            'ten_nvl': str(self._cell(row, self.COL_TEN_NVL) or '').strip(),
            'sls': sls,
            'dgs': dgs,
            'gts': gts,
        }

    def _parse_row_legacy_sl(self, row, row_number, existing_map, errors):
        ma_nvl = self._normalize_ma_nvl(self._cell(row, self.COL_MA_NVL))
        if not ma_nvl:
            errors.append(_('Dòng %d: thiếu Mã NVL.') % row_number)
            return None
        line_id = existing_map.get(ma_nvl)
        if not line_id:
            errors.append(
                _('Dòng %d: Mã NVL "%s" không có trong Tổng hợp KH vật tư BCU của kỳ này.')
                % (row_number, ma_nvl)
            )
            return None
        sls, dgs, gts = [], [], []
        for offset in range(4):
            try:
                sl = self._parse_number(self._cell(row, 2 + offset), default=0.0)
            except UserError as exc:
                errors.append(_('Dòng %d: %s') % (row_number, exc.args[0]))
                return None
            if sl < 0:
                errors.append(_('Dòng %d: Số lượng không được âm.') % row_number)
                return None
            sls.append(sl)
            dgs.append(0.0)
            gts.append(0.0)
        return {
            'line_id': line_id,
            'ma_nvl': ma_nvl,
            'ten_nvl': str(self._cell(row, self.COL_TEN_NVL) or '').strip(),
            'sls': sls,
            'dgs': dgs,
            'gts': gts,
        }

    def _upsert_vat_tu_di_duong_bcu(self, parsed_rows):
        """Đồng bộ bảng phẳng vật tư đi đường (loai=bcu) để menu tra cứu."""
        self.ensure_one()
        period = self.period_id
        company = period.company_sx_id
        if not company:
            return

        month_keys = self._month_keys()
        VatTuDiDuong = self.env['vat.tu.di.duong'].sudo()
        existing = VatTuDiDuong.search([
            ('company_id', '=', company.id),
            ('loai', '=', 'bcu'),
            ('ma_nvl', 'in', [row['ma_nvl'] for row in parsed_rows]),
            ('month_key', 'in', month_keys),
        ])
        existing_map = {
            (rec.ma_nvl, rec.month_key): rec.id for rec in existing
        }

        to_create = []
        update_ids, update_qtys, update_dgs = [], [], []
        for row in parsed_rows:
            for offset, month_key in enumerate(month_keys):
                key = (row['ma_nvl'], month_key)
                vals = {
                    'company_id': company.id,
                    'loai': 'bcu',
                    'ma_nvl': row['ma_nvl'],
                    'ten_nvl': row['ten_nvl'] or False,
                    'month_key': month_key,
                    'month_date': period._month_key_to_date(month_key),
                    'so_luong': row['sls'][offset],
                    'don_gia': row['dgs'][offset],
                }
                line_id = existing_map.get(key)
                if line_id:
                    update_ids.append(line_id)
                    update_qtys.append(vals['so_luong'])
                    update_dgs.append(vals['don_gia'])
                else:
                    to_create.append(vals)

        if update_ids:
            self.env.cr.execute("""
                UPDATE vat_tu_di_duong AS v SET
                    so_luong = data.so_luong,
                    don_gia = data.don_gia,
                    write_uid = %s,
                    write_date = NOW() AT TIME ZONE 'UTC'
                FROM (
                    SELECT unnest(%s::int[]) AS id,
                           unnest(%s::numeric[]) AS so_luong,
                           unnest(%s::numeric[]) AS don_gia
                ) AS data
                WHERE v.id = data.id
            """, [self.env.uid, update_ids, update_qtys, update_dgs])
            VatTuDiDuong.browse(update_ids).invalidate_recordset(
                ['so_luong', 'don_gia', 'gia_tri', 'write_uid', 'write_date'])
        if to_create:
            VatTuDiDuong.with_context(
                vat_tu_di_duong_loai='bcu',
                tracking_disable=True,
                vat_tu_import_bulk=True,
            ).create(to_create)

    def action_import(self):
        self.ensure_one()
        if not self.period_id.code:
            raise UserError(_('Kỳ kế hoạch chưa có số chứng từ.'))

        ws = self._get_import_worksheet()
        expected_code = (self.period_id.code or '').strip()
        doc_code = self._parse_doc_code_from_sheet(ws)
        legacy_column = doc_code == 'legacy_column'
        if legacy_column:
            doc_code = False
        elif not doc_code:
            raise UserError(_('File Excel thiếu Số chứng từ ở ô B1.'))
        elif doc_code != expected_code:
            raise UserError(
                _('Số chứng từ "%s" không khớp kỳ "%s".') % (doc_code, expected_code))

        fmt = self._detect_import_format(ws)
        data_start = self._resolve_data_start_row(ws, fmt)
        rows = list(ws.iter_rows(min_row=data_start, values_only=True))
        if not rows:
            raise UserError(_('File Excel không có dòng dữ liệu.'))

        errors = []
        parsed = []
        existing_map = {line.ma_sap: line.id for line in self._get_b6_lines()}

        for row_number, row in enumerate(rows, start=data_start):
            if not any(cell not in (None, '') for cell in row):
                continue
            if fmt == 'legacy_sl_only':
                item = self._parse_row_legacy_sl(row, row_number, existing_map, errors)
            else:
                item = self._parse_row_sl_dg(row, row_number, existing_map, errors)
            if item:
                parsed.append(item)

        self._raise_import_errors(errors)

        updated = len(parsed)
        if updated:
            ids = [p['line_id'] for p in parsed]
            sl0 = [p['sls'][0] for p in parsed]
            sl1 = [p['sls'][1] for p in parsed]
            sl2 = [p['sls'][2] for p in parsed]
            sl3 = [p['sls'][3] for p in parsed]
            dg0 = [p['dgs'][0] for p in parsed]
            dg1 = [p['dgs'][1] for p in parsed]
            dg2 = [p['dgs'][2] for p in parsed]
            dg3 = [p['dgs'][3] for p in parsed]
            gt0 = [p['gts'][0] for p in parsed]
            gt1 = [p['gts'][1] for p in parsed]
            gt2 = [p['gts'][2] for p in parsed]
            gt3 = [p['gts'][3] for p in parsed]
            self.env.cr.execute(
                """
                UPDATE kh_dat_vat_tu_bcu bcu
                   SET ve_du_kien_bcu_t0 = v.sl0,
                       ve_du_kien_bcu_t1 = v.sl1,
                       ve_du_kien_bcu_t2 = v.sl2,
                       ve_du_kien_bcu_t3 = v.sl3,
                       ve_du_kien_bcu_dg_t0 = v.dg0,
                       ve_du_kien_bcu_dg_t1 = v.dg1,
                       ve_du_kien_bcu_dg_t2 = v.dg2,
                       ve_du_kien_bcu_dg_t3 = v.dg3,
                       ve_du_kien_bcu_gt_t0 = v.gt0,
                       ve_du_kien_bcu_gt_t1 = v.gt1,
                       ve_du_kien_bcu_gt_t2 = v.gt2,
                       ve_du_kien_bcu_gt_t3 = v.gt3,
                       tong_ve_du_kien_bcu = COALESCE(v.sl0, 0) + COALESCE(v.sl1, 0)
                                           + COALESCE(v.sl2, 0) + COALESCE(v.sl3, 0),
                       tong_gia_tri_bcu = COALESCE(v.gt0, 0) + COALESCE(v.gt1, 0)
                                        + COALESCE(v.gt2, 0) + COALESCE(v.gt3, 0),
                       write_uid = %s,
                       write_date = NOW() AT TIME ZONE 'UTC'
                  FROM unnest(
                           %s::int[],
                           %s::numeric[], %s::numeric[], %s::numeric[], %s::numeric[],
                           %s::numeric[], %s::numeric[], %s::numeric[], %s::numeric[],
                           %s::numeric[], %s::numeric[], %s::numeric[], %s::numeric[]
                       ) AS v(
                           id,
                           sl0, sl1, sl2, sl3,
                           dg0, dg1, dg2, dg3,
                           gt0, gt1, gt2, gt3
                       )
                 WHERE bcu.id = v.id
                """,
                [
                    self.env.uid,
                    ids, sl0, sl1, sl2, sl3,
                    dg0, dg1, dg2, dg3,
                    gt0, gt1, gt2, gt3,
                ],
            )
            self.env['kh.dat.vat.tu.bcu'].browse(ids).invalidate_recordset([
                've_du_kien_bcu_t0', 've_du_kien_bcu_t1', 've_du_kien_bcu_t2', 've_du_kien_bcu_t3',
                've_du_kien_bcu_dg_t0', 've_du_kien_bcu_dg_t1', 've_du_kien_bcu_dg_t2', 've_du_kien_bcu_dg_t3',
                've_du_kien_bcu_gt_t0', 've_du_kien_bcu_gt_t1', 've_du_kien_bcu_gt_t2', 've_du_kien_bcu_gt_t3',
                'tong_ve_du_kien_bcu', 'tong_gia_tri_bcu', 'write_date', 'write_uid',
            ])
            b6_lines = self.env['kh.dat.vat.tu.bcu'].browse(ids)
            b6_lines._compute_sl_du_tru_toi_thieu_bcu()
            b6_lines._compute_sl_dat_mua_de_xuat()
            self.env['kh.dat.vat.tu.bcu']._apply_chot_from_bcu_di_duong(b6_lines)
            b6_lines._compute_b6_derived()
            b6_lines.flush_recordset([
                'sl_du_tru_toi_thieu_bcu',
                'sl_dat_mua_de_xuat', 'sl_dat_mua_chot', 'sl_can_mua_theo_moq',
                'sl_ton_kho_cuoi_ky', 'so_ngay_vong_quay_ton',
                'don_gia_ton_kho_cuoi_ky', 'gia_tri_ton_kho_cuoi_ky', 'gia_tri_mua_hang',
            ])
            self._upsert_vat_tu_di_duong_bcu(parsed)
            self._post_period_import_file_log(
                self.period_id,
                '<p><b>Đã import file vật tư đi đường BCU %s.</b></p>'
                % (self.file_name or '-'),
            )
            title = _('Hoàn tất import hàng đi đường BCU.')
            success = True
        else:
            title = _('Không có dữ liệu hợp lệ để import.')
            success = False
        return self._notify_and_close(title, '', success=success)

    @api.model
    def action_open_from_menu(self):
        view = self.env.ref('sonha_vat_tu.view_import_tong_hop_bcu_wizard_form')
        return {
            'name': _('Import vật tư đi đường BCU'),
            'type': 'ir.actions.act_window',
            'res_model': 'import.tong.hop.bcu.wizard',
            'view_mode': 'form',
            'views': [(view.id, 'form')],
            'view_id': view.id,
            'target': 'new',
            'context': {'vat_tu_di_duong_loai': 'bcu'},
        }
