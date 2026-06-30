# -*- coding: utf-8 -*-
import base64
import io
import warnings

from openpyxl import load_workbook
from odoo import _, fields, models
from odoo.exceptions import UserError


class ImportBomWizard(models.TransientModel):
    _name = 'import.bom.wizard'
    _description = 'Import BOM từ Excel'

    file_data = fields.Binary(string='File Excel')
    file_name = fields.Char(string='Tên file')

    def _to_float_or_error(self, raw, field_label, row_idx, errors):
        if raw in (None, ''):
            return 0.0
        try:
            return float(raw)
        except (TypeError, ValueError):
            errors.append(_('Dòng %d: cột %s không phải số (%s).') % (row_idx, field_label, raw))
            return None

    def action_import(self):
        self.ensure_one()
        Bom = self.env['bom'].sudo()
        MaHang = self.env['ma.hang'].sudo()
        if not self.file_data:
            raise UserError(_('Vui lòng chọn file Excel.'))

        try:
            data = base64.b64decode(self.file_data)
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', message='Data Validation extension', category=UserWarning)
                wb = load_workbook(io.BytesIO(data), data_only=True)
        except Exception as exc:
            raise UserError(_('Không đọc được file Excel: %s') % exc)

        ws = wb.active
        rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        if not rows:
            raise UserError(_('File rỗng.'))

        errors = []
        vals_list = []
        seen = set()

        for row_idx, row in enumerate(rows[1:], start=2):
            if not row or not any(c not in (None, '') for c in row):
                continue
            # Template: Mã TP | Tên TP | Mã NVL | Tên NVL | SL định mức | SL SPĐM | Độ dày | Khổ 1 | Khổ 2
            row = (list(row) + [None] * 9)[:9]

            ma_tp  = (str(row[0]).strip() if row[0] not in (None, '') else '')
            ten_tp = (str(row[1]).strip() if row[1] not in (None, '') else '')
            ma_nvl = (str(row[2]).strip() if row[2] not in (None, '') else '')
            ten_nvl = (str(row[3]).strip() if row[3] not in (None, '') else '')
            if not ma_tp or not ma_nvl:
                errors.append(_('Dòng %d: thiếu Mã TP hoặc Mã NVL.') % row_idx)
                continue

            key = (ma_tp, ma_nvl)
            if key in seen:
                errors.append(_('Dòng %d: trùng bộ (Mã TP, Mã NVL) trong file.') % row_idx)
                continue
            seen.add(key)

            sl_dinh_muc = self._to_float_or_error(row[4], 'Số lượng định mức', row_idx, errors)
            sl_spdm     = self._to_float_or_error(row[5], 'Số lượng SPĐM', row_idx, errors)
            if sl_spdm is not None and sl_spdm == 0:
                sl_spdm = 1.0
            do_day = self._to_float_or_error(row[6], 'Độ dày', row_idx, errors)
            kho_1  = self._to_float_or_error(row[7], 'Khổ 1', row_idx, errors)
            kho_2  = self._to_float_or_error(row[8], 'Khổ 2', row_idx, errors)
            if None in (sl_dinh_muc, sl_spdm, do_day, kho_1, kho_2):
                continue

            vals_list.append({
                'ma_tp':       ma_tp,
                'ten_tp':      ten_tp or ma_tp,
                'ma_nvl':      ma_nvl,
                'ten_nvl':     ten_nvl or ma_nvl,
                'sl_dinh_muc': sl_dinh_muc,
                'sl_spdm':     sl_spdm if sl_spdm else 1.0,
                'do_day':      do_day,
                'kho_1':       kho_1,
                'kho_2':       kho_2,
            })

        if vals_list:
            nvl_codes = sorted({v['ma_nvl'] for v in vals_list if v.get('ma_nvl')})
            nvl_master = MaHang.search([('ma_sap', 'in', nvl_codes)])
            nvl_map = {r.ma_sap: r for r in nvl_master if r.ma_sap}
            for vals in vals_list:
                master = nvl_map.get(vals['ma_nvl'])
                if master and master.ten_hang:
                    vals['ten_nvl'] = master.ten_hang

            existing = Bom.search([
                ('ma_tp',  'in', [v['ma_tp']  for v in vals_list]),
                ('ma_nvl', 'in', [v['ma_nvl'] for v in vals_list]),
            ])
            existing_map = {(r.ma_tp, r.ma_nvl): r for r in existing}

            create_list = []
            for vals in vals_list:
                key = (vals['ma_tp'], vals['ma_nvl'])
                if key in existing_map:
                    existing_map[key].write({
                        'do_day':      vals['do_day'],
                        'kho_1':       vals['kho_1'],
                        'kho_2':       vals['kho_2'],
                        'sl_dinh_muc': vals['sl_dinh_muc'],
                        'sl_spdm':     vals['sl_spdm'],
                    })
                else:
                    create_list.append(vals)

        if errors:
            raise UserError(_('File import có lỗi, chưa ghi dữ liệu:\n- %s') % '\n- '.join(errors[:100]))

        if create_list:
            Bom.create(create_list)

        return {'type': 'ir.actions.act_window_close'}
