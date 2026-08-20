# -*- coding: utf-8 -*-
import re
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Font

from markupsafe import Markup

from odoo import _, api, fields, models
from odoo.exceptions import UserError

TEMPLATE_WIDTHS = [16, 36, 14, 14, 14, 14]


class ImportVatTuDiDuongWizard(models.TransientModel):
    _name = 'import.vat.tu.di.duong.wizard'
    _description = 'Import vật tư đi đường'
    _inherit = ['vat.tu.excel.mixin']

    TEMPLATE_SHEET_NAME = 'Vat tu di duong'
    META_ROW = 1
    HEADER_ROW = 4
    DATA_START_ROW = 5
    META_FONT_SIZE = 13
    COL_MA_NVL, COL_TEN_NVL = 0, 1
    COL_T0, COL_T1, COL_T2, COL_T3 = 2, 3, 4, 5

    MONTH_RE = re.compile(r'(\d{1,2})\s*[/\-]\s*(\d{4})')

    file_data = fields.Binary(string='File Excel')
    file_name = fields.Char(string='Tên file')
    period_id = fields.Many2one(
        'ke.hoach.vat.tu',
        string='Kỳ kế hoạch',
        readonly=True,
        default=lambda self: self.env.context.get('default_period_id'),
    )

    def _company_sx(self):
        self.ensure_one()
        if self.period_id and self.period_id.company_sx_id:
            return self.period_id.company_sx_id
        return self.env.company

    def _get_month_keys(self):
        self.ensure_one()
        if self.period_id:
            months = self.period_id._get_horizon_months()
            if len(months) != 4:
                raise UserError(_('Kỳ kế hoạch chưa xác định được 4 tháng tính toán.'))
            return months
        return self._horizon_months_from_today()

    @api.model
    def _horizon_months_from_today(self):
        today = fields.Date.context_today(self)
        m, y = today.month, today.year
        res = []
        for i in range(4):
            tm = m + i
            ty = y
            while tm > 12:
                tm -= 12
                ty += 1
            res.append('%02d/%d' % (tm, ty))
        return res

    def _get_month_headers(self):
        return [_('Tháng %s') % month for month in self._get_month_keys()]

    def _month_key_from_header(self, header):
        text = str(header or '').strip()
        thang_prefix = _('Tháng') + ' '
        if text.startswith(thang_prefix):
            text = text[len(thang_prefix):].strip()
        if isinstance(header, (date, datetime)):
            return header.strftime('%m/%Y')
        match = self.MONTH_RE.search(text)
        if not match:
            return False
        month, year = int(match.group(1)), int(match.group(2))
        try:
            return '%02d/%d' % (month, year)
        except ValueError:
            return False

    # ------------------------------------------------------------------
    # Export template — 1 dòng / mã NVL (gom theo đơn vị SX)
    # ------------------------------------------------------------------
    def action_download_template(self):
        self.ensure_one()
        if self.period_id:
            return self._download_vat_tu_di_duong_template_from_period()
        return self._download_vat_tu_di_duong_template_empty()

    def _build_template_workbook(self, row_vals=None):
        self.ensure_one()
        month_headers = self._get_month_headers()
        wb = Workbook()
        ws = wb.active
        ws.title = self.TEMPLATE_SHEET_NAME

        period_code = (self.period_id.code or '').strip() if self.period_id else ''
        if period_code:
            ws.cell(row=self.META_ROW, column=1, value=_('Số chứng từ'))
            ws.cell(row=self.META_ROW, column=2, value=period_code)
            meta_font = Font(name='Times New Roman', size=self.META_FONT_SIZE, bold=True)
            ws.cell(row=self.META_ROW, column=1).font = meta_font
            ws.cell(row=self.META_ROW, column=2).font = meta_font
            ws.row_dimensions[self.META_ROW].height = 24

        headers = [_('Mã NVL'), _('Tên NVL')] + month_headers
        for col_idx, label in enumerate(headers, start=1):
            ws.cell(row=self.HEADER_ROW, column=col_idx, value=label)

        row_idx = self.DATA_START_ROW
        for row in row_vals or []:
            ws.cell(row=row_idx, column=1, value=row.get('ma_nvl') or '')
            ws.cell(row=row_idx, column=2, value=row.get('ten_nvl') or '')
            qtys = row.get('qtys') or [0.0] * 4
            for offset in range(4):
                ws.cell(
                    row=row_idx, column=3 + offset,
                    value=qtys[offset] if offset < len(qtys) else 0.0,
                )
            row_idx += 1

        max_col = len(headers)
        self._style_excel_header(ws, max_col, header_row=self.HEADER_ROW)
        if row_vals:
            self._style_excel_body(ws, max_col, self.DATA_START_ROW, row_idx - 1)
        self._set_excel_widths(ws, TEMPLATE_WIDTHS[:max_col])
        return wb

    def _download_vat_tu_di_duong_template_empty(self):
        wb = self._build_template_workbook()
        return self._xlsx_download_action(wb, 'Template_vat_tu_di_duong.xlsx')

    def _download_vat_tu_di_duong_template_from_period(self):
        period = self.period_id
        if not period.code:
            raise UserError(_('Kỳ kế hoạch chưa có số chứng từ.'))
        if not period.tinh_toan_vat_tu_ids:
            raise UserError(_(
                'Chưa có dữ liệu tính toán vật tư. Vui lòng chạy bước Tính toán trước.'
            ))
        row_vals = period._vat_tu_di_duong_template_rows()
        wb = self._build_template_workbook(row_vals)
        return self._xlsx_download_action(
            wb, 'Template_vat_tu_di_duong_%s.xlsx' % period.code)

    # ------------------------------------------------------------------
    # Import
    # ------------------------------------------------------------------
    def _parse_doc_code_from_sheet(self, ws):
        label = str(ws.cell(row=self.META_ROW, column=1).value or '').strip()
        value = str(ws.cell(row=self.META_ROW, column=2).value or '').strip()
        if label == _('Số chứng từ') and value:
            return value
        return False

    def _parse_month_headers_from_sheet(self, ws):
        month_keys = []
        for col in range(3, 7):
            header = ws.cell(row=self.HEADER_ROW, column=col).value
            month_key = self._month_key_from_header(header)
            if not month_key:
                raise UserError(
                    _('Header cột %d "%s" không đúng định dạng Tháng MM/YYYY.')
                    % (col, header or '')
                )
            month_keys.append(month_key)
        return month_keys

    def _get_b3_ma_set(self):
        self.ensure_one()
        return {
            (line.ma_vat_tu or '').strip()
            for line in self.period_id.tinh_toan_vat_tu_ids
            if (line.ma_vat_tu or '').strip()
        }

    def _parse_rows(self, ws):
        errors = []
        parsed = []
        seen_keys = set()

        if self.period_id and (self.period_id.code or '').strip():
            expected_code = self.period_id.code.strip()
            doc_code = self._parse_doc_code_from_sheet(ws)
            if not doc_code:
                raise UserError(_('File Excel thiếu Số chứng từ ở ô B1.'))
            if doc_code != expected_code:
                raise UserError(
                    _('Số chứng từ "%s" không khớp kỳ "%s".') % (doc_code, expected_code)
                )

        header_a = str(ws.cell(row=self.HEADER_ROW, column=1).value or '').strip()
        if header_a == _('Đơn vị'):
            raise UserError(_(
                'File Excel dùng mẫu cũ (có cột Đơn vị). '
                'Vui lòng tải lại template mới — chỉ còn Mã NVL và 4 cột tháng.'
            ))
        if header_a != _('Mã NVL'):
            raise UserError(_('File Excel không đúng định dạng template vật tư đi đường.'))

        month_keys = self._parse_month_headers_from_sheet(ws)
        if self.period_id:
            expected_months = self.period_id._get_horizon_months()
            if month_keys != expected_months:
                raise UserError(_('Tháng trên file không khớp kỳ kế hoạch.'))

        b3_ma_set = self._get_b3_ma_set() if self.period_id else None
        company = self._company_sx()
        Period = self.env['ke.hoach.vat.tu']
        month_cols = (self.COL_T0, self.COL_T1, self.COL_T2, self.COL_T3)
        rows = list(ws.iter_rows(min_row=self.DATA_START_ROW, values_only=True))

        for row_number, row in enumerate(rows, start=self.DATA_START_ROW):
            if not any(cell not in (None, '') for cell in row):
                continue

            ma_nvl = self._normalize_ma_nvl(self._cell(row, self.COL_MA_NVL))
            ten_nvl = str(self._cell(row, self.COL_TEN_NVL) or '').strip()

            row_errors = []
            if not ma_nvl:
                row_errors.append(_('Dòng %d: thiếu Mã NVL.') % row_number)

            if b3_ma_set is not None and ma_nvl and ma_nvl not in b3_ma_set:
                row_errors.append(
                    _('Dòng %d: Mã NVL "%s" không có trong tính toán vật tư của kỳ này.')
                    % (row_number, ma_nvl)
                )

            try:
                qtys = [
                    self._parse_number(self._cell(row, col), default=0.0)
                    for col in month_cols
                ]
            except UserError as exc:
                row_errors.append(_('Dòng %d: %s') % (row_number, exc.args[0]))
                qtys = []

            if qtys and any(qty < 0 for qty in qtys):
                row_errors.append(_('Dòng %d: Số lượng không được âm.') % row_number)

            if row_errors:
                errors.extend(row_errors)
                continue

            row_parsed = []
            dup_in_row = False
            for month_key, so_luong in zip(month_keys, qtys):
                dup_key = (ma_nvl, month_key)
                if dup_key in seen_keys:
                    errors.append(
                        _('Dòng %d: trùng Mã NVL + Tháng trong file.') % row_number
                    )
                    dup_in_row = True
                    break
                seen_keys.add(dup_key)
                row_parsed.append({
                    'company_id': company.id,
                    'ma_nvl': ma_nvl,
                    'ten_nvl': ten_nvl or False,
                    'month_key': month_key,
                    'month_date': Period._month_key_to_date(month_key),
                    'so_luong': so_luong,
                    'loai': 'don_vi',
                })
            if not dup_in_row:
                parsed.extend(row_parsed)

        self._raise_import_errors(errors)
        return parsed

    def _apply_rows(self, parsed):
        """Ghi vật tư đi đường SX (loai=don_vi, gom theo đơn vị SX)."""
        if not parsed:
            return 0, 0

        loai = 'don_vi'
        uid = self.env.uid
        is_truong = self.env.user.has_group('sonha_vat_tu.group_truong_bo_phan_vat_tu')
        VatTuDiDuong = self.env['vat.tu.di.duong']
        existing = VatTuDiDuong.sudo().search([
            ('company_id', 'in', list({v['company_id'] for v in parsed})),
            ('ma_nvl', 'in', list({v['ma_nvl'] for v in parsed})),
            ('month_key', 'in', list({v['month_key'] for v in parsed})),
            ('loai', '=', loai),
        ])
        existing_map = {
            (line.company_id.id, line.ma_nvl, line.month_key, line.loai): line
            for line in existing
        }

        to_create = []
        update_ids, update_names, update_qtys = [], [], []
        blocked = []
        for vals in parsed:
            vals.setdefault('loai', loai)
            key = (vals['company_id'], vals['ma_nvl'], vals['month_key'], loai)
            line = existing_map.get(key)
            if line:
                if not is_truong and line.create_uid.id != uid:
                    blocked.append(
                        _('%s / %s') % (vals['ma_nvl'], vals['month_key'])
                    )
                    continue
                update_ids.append(line.id)
                update_names.append(vals['ten_nvl'] or '')
                update_qtys.append(vals['so_luong'])
            else:
                to_create.append(vals)

        if blocked:
            raise UserError(_(
                'Không thể ghi đè vật tư đi đường do người khác import:\n%s'
            ) % '\n'.join(blocked[:20]))

        if update_ids:
            self.env.cr.execute("""
                UPDATE vat_tu_di_duong AS v SET
                    ten_nvl = COALESCE(NULLIF(data.ten_nvl, ''), v.ten_nvl),
                    so_luong = data.so_luong,
                    write_uid = %s,
                    write_date = NOW() AT TIME ZONE 'UTC'
                FROM (
                    SELECT unnest(%s::int[]) AS id,
                           unnest(%s::varchar[]) AS ten_nvl,
                           unnest(%s::numeric[]) AS so_luong
                ) AS data
                WHERE v.id = data.id
            """, [uid, update_ids, update_names, update_qtys])
            VatTuDiDuong.browse(update_ids).invalidate_recordset(
                ['ten_nvl', 'so_luong', 'write_uid', 'write_date'])
        if to_create:
            VatTuDiDuong.with_context(
                tracking_disable=True,
                vat_tu_di_duong_loai='don_vi',
                vat_tu_import_bulk=True,
            ).create(to_create)

        return len(to_create), len(update_ids)

    def action_import(self):
        self.ensure_one()
        ws = self._get_import_worksheet()
        parsed = self._parse_rows(ws)

        if not parsed:
            raise UserError(_('File Excel không có dữ liệu hợp lệ.'))

        created, updated = self._apply_rows(parsed)

        if self.period_id and (created or updated):
            self.period_id.with_context(
                mail_create_nosubscribe=True,
                tracking_disable=True,
            ).message_post(
                body=Markup(
                    '<p><b>Đã import file vật tư đi đường %s.</b></p>'
                    % (self.file_name or '-')
                ),
                attachment_ids=[self.env['ir.attachment'].sudo().create({
                    'name': self.file_name or 'import.xlsx',
                    'type': 'binary',
                    'datas': self.file_data,
                    'res_model': 'ke.hoach.vat.tu',
                    'res_id': self.period_id.id,
                    'mimetype': (
                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    ),
                }).id],
            )

        message = (
            _('Hoàn tất nhập hàng đi đường.')
            if created or updated
            else _('Không có dữ liệu hợp lệ để import.')
        )

        next_action = None
        if self.period_id:
            next_action = {
                'type': 'ir.actions.act_window',
                'res_model': 'ke.hoach.vat.tu',
                'res_id': self.period_id.id,
                'view_mode': 'form',
                'views': [(self.env.ref('sonha_vat_tu.view_ke_hoach_vat_tu_form_b3').id, 'form')],
                'target': 'current',
            }

        return self._notify_and_close(
            _('Import vật tư đi đường'), message,
            success=bool(created or updated), next_action=next_action,
        )
