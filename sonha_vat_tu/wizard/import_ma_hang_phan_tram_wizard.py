# -*- coding: utf-8 -*-
import base64
import io
import warnings

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from odoo import _, fields, models
from odoo.exceptions import UserError


class ImportMaHangPhanTramWizard(models.TransientModel):
    _name = 'import.ma.hang.phan.tram.wizard'
    _description = 'Import phần trăm mã hàng'

    TEMPLATE_SHEET_NAME = 'Phan tram ma hang'
    DATA_START_ROW = 2
    COL_COMPANY, COL_MA_NVL, COL_TEN_NVL, COL_PHAN_TRAM = range(4)

    file_data = fields.Binary(string='File Excel')
    file_name = fields.Char(string='Tên file')

    def _cell(self, row, col):
        return row[col] if col < len(row) else None

    def _parse_float(self, value, allow_empty=False):
        if value in (None, ''):
            if allow_empty:
                return 0.0
            raise UserError(_('Thiếu phần trăm.'))
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip().replace(' ', '').replace(',', '.')
        if text.endswith('%'):
            text = text[:-1]
        try:
            return float(text)
        except ValueError:
            raise UserError(_('Không đọc được phần trăm "%s".') % value)

    def _normalize_ma_nvl(self, value):
        if value in (None, ''):
            return ''
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        if isinstance(value, int):
            return str(value)
        text = str(value).strip()
        if text.endswith('.0') and text[:-2].isdigit():
            return text[:-2]
        return text

    def _xlsx_download_action(self, wb, filename):
        output = io.BytesIO()
        wb.save(output)
        attachment = self.env['ir.attachment'].sudo().create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(output.getvalue()),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }

    def _get_company_codes(self):
        return sorted({
            (c.company_code or '').strip()
            for c in self.env['res.company'].sudo().search([])
            if (c.company_code or '').strip()
        })

    def _build_company_lookup(self):
        return {
            (c.company_code or '').strip(): c
            for c in self.env['res.company'].sudo().search([])
            if (c.company_code or '').strip()
        }

    def _apply_company_code_validation(self, wb, ws, company_codes):
        if not company_codes:
            return
        ref_sheet = wb.create_sheet('_company_codes')
        ref_sheet.sheet_state = 'hidden'
        for row_idx, code in enumerate(company_codes, start=1):
            ref_sheet.cell(row=row_idx, column=1, value=code)
        dv = DataValidation(
            type='list',
            formula1=f"='_company_codes'!$A$1:$A${len(company_codes)}",
            allow_blank=False,
        )
        dv.error = _('Chỉ được chọn mã đơn vị có trong danh mục.')
        dv.errorTitle = _('Đơn vị không hợp lệ')
        ws.add_data_validation(dv)
        dv.add('A2:A5000')

    def _apply_template_style(self, ws, max_col=4):
        body_font = Font(name='Times New Roman', size=10)
        header_font = Font(name='Times New Roman', size=10, bold=True, color='FFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='3F6F8F')
        header_side = Side(style='thin', color='2F556D')
        header_border = Border(
            left=header_side, right=header_side, top=header_side, bottom=header_side,
        )
        for cell in ws[1][:max_col]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = 'A2'
        for row in ws.iter_rows(min_row=2, max_row=5001, min_col=1, max_col=max_col):
            for cell in row:
                cell.font = body_font
                cell.alignment = Alignment(vertical='center')

    def _get_import_worksheet(self, workbook):
        if self.TEMPLATE_SHEET_NAME in workbook.sheetnames:
            return workbook[self.TEMPLATE_SHEET_NAME]
        for name in workbook.sheetnames:
            if not name.startswith('_'):
                return workbook[name]
        return workbook.active

    def _read_data_rows(self):
        if not self.file_data:
            raise UserError(_('Vui lòng chọn file Excel.'))
        try:
            data = base64.b64decode(self.file_data)
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', message='Data Validation extension', category=UserWarning)
                workbook = load_workbook(io.BytesIO(data), data_only=True)
        except Exception as exc:
            raise UserError(_('Không đọc được file Excel: %s') % exc)
        ws = self._get_import_worksheet(workbook)
        return list(ws.iter_rows(min_row=self.DATA_START_ROW, values_only=True))

    def action_download_template(self):
        self.ensure_one()
        wb = Workbook()
        ws = wb.active
        ws.title = self.TEMPLATE_SHEET_NAME
        headers = ['Đơn vị', 'Mã NVL', 'Tên NVL', 'Phần trăm']
        for col_idx, label in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=label)
        self._apply_template_style(ws)
        self._apply_company_code_validation(wb, ws, self._get_company_codes())
        for col_idx, width in enumerate([14, 26, 36, 14], start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
        return self._xlsx_download_action(wb, 'Template_phan_tram_ma_hang.xlsx')

    def action_import(self):
        self.ensure_one()
        rows = self._read_data_rows()
        if not rows:
            raise UserError(_('File Excel không có dòng dữ liệu (từ dòng 2).'))

        PhanTram = self.env['ma.hang.phan.tram'].sudo()
        MaHang = self.env['ma.hang'].sudo()
        company_lookup = self._build_company_lookup()

        errors = []
        created_items = []
        updated_items = []
        seen_keys = set()

        for row_number, row in enumerate(rows, start=self.DATA_START_ROW):
            if not any(cell not in (None, '') for cell in row):
                continue

            company_code = str(self._cell(row, self.COL_COMPANY) or '').strip()
            ma_nvl = self._normalize_ma_nvl(self._cell(row, self.COL_MA_NVL))
            phan_tram_raw = self._cell(row, self.COL_PHAN_TRAM)

            row_errors = []
            if not company_code:
                row_errors.append(_('Dòng %d: thiếu Đơn vị.') % row_number)
            if not ma_nvl:
                row_errors.append(_('Dòng %d: thiếu Mã NVL.') % row_number)

            company = company_lookup.get(company_code) if company_code else False
            if company_code and not company:
                row_errors.append(_('Dòng %d: Đơn vị "%s" không tồn tại.') % (row_number, company_code))

            phan_tram = None
            if phan_tram_raw in (None, ''):
                row_errors.append(_('Dòng %d: thiếu Phần trăm.') % row_number)
            else:
                try:
                    phan_tram = self._parse_float(phan_tram_raw)
                    if phan_tram < 0:
                        row_errors.append(_('Dòng %d: Phần trăm không được âm.') % row_number)
                except UserError as exc:
                    row_errors.append(_('Dòng %d: %s') % (row_number, exc.args[0]))

            if company and ma_nvl:
                dup_key = (company.id, ma_nvl)
                if dup_key in seen_keys:
                    row_errors.append(_('Dòng %d: trùng Đơn vị + Mã NVL trong file.') % row_number)
                else:
                    seen_keys.add(dup_key)

            if company and ma_nvl and not row_errors:
                if not MaHang.search_count([
                    ('company_id', '=', company.id),
                    ('ma_sap', '=', ma_nvl),
                ]):
                    row_errors.append(_(
                        'Dòng %d: Mã NVL "%s" không có trong danh mục mã hàng (ĐVCS %s).'
                    ) % (row_number, ma_nvl, company_code))

            if row_errors:
                errors.extend(row_errors)
                continue

            existing = PhanTram.search([
                ('company_id', '=', company.id),
                ('ma_sap', '=', ma_nvl),
            ], limit=1)
            label = '%s (%s, %s%%)' % (ma_nvl, company_code, phan_tram)
            if existing:
                existing.write({'phan_tram': phan_tram})
                updated_items.append(label)
            else:
                PhanTram.create({
                    'company_id': company.id,
                    'ma_sap': ma_nvl,
                    'phan_tram': phan_tram,
                })
                created_items.append(label)

        if errors:
            shown = errors[:80]
            message = '\n'.join('- %s' % err for err in shown)
            if len(errors) > 80:
                message += _('\n... còn %d lỗi khác.') % (len(errors) - 80)
            raise UserError(message)

        parts = []
        if created_items:
            parts.append(_('thêm %s') % ', '.join(created_items))
        if updated_items:
            parts.append(_('cập nhật %s') % ', '.join(updated_items))
        if parts:
            message = _('Import thành công: %s.') % ', '.join(parts)
        else:
            message = _('Không có dữ liệu hợp lệ để import.')

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Import phần trăm mã hàng'),
                'message': message,
                'type': 'success' if created_items or updated_items else 'warning',
                'sticky': False,
                'next': {'type': 'ir.actions.act_window_close'},
            },
        }
