from odoo import models, fields
from odoo.exceptions import UserError, ValidationError
from datetime import datetime
import base64
import io
import tempfile
import os
from openpyxl.styles import Border, Side, Alignment
from openpyxl import load_workbook


class ExportReportRewardDiscipline(models.TransientModel):
    _name = 'export.report.reward.discipline'

    report = fields.Many2one('config.template.report', string="Chọn báo cáo",
                             domain="[('apply', '=', res_model)]")
    template_file = fields.Binary("Tệp Excel")
    res_model = fields.Many2one('ir.model')

    def action_export(self):
        active_ids = self.env.context.get('active_ids', [])
        if not active_ids:
            raise ValidationError("Không có bản ghi nào được chọn.")
        if self.report.key == "tdkl":
            list_record = self.env['discipline.report'].browse(active_ids)
            row = 7
            row_min = 7
        elif self.report.key == "thklbt":
            list_record = self.env['discipline.report'].browse(active_ids)
            row = 5
            row_min = 5
        elif self.report.key == "tdkt":
            list_record = self.env['reward.report'].browse(active_ids)
            row = 6
            row_min = 6
        elif self.report.key == "qdmn":
            list_record = self.env['dismissal.report'].browse(active_ids)
            row = 10
            row_min = 10
        elif self.report.key == "qdnv" or self.report.key == "qddc":
            list_record = self.env['dismissal.report'].browse(active_ids)
            row = 9
            row_min = 9
        else:
            raise ValidationError("Chỉ sử dụng báo cáo khen thưởng kỷ luật.")
        file_stream = io.BytesIO(base64.b64decode(self.report.file))
        wb = load_workbook(filename=file_stream)
        sheet = wb['Sheet1']
        stt = 1

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        for record in list_record:
            # Báo cáo theo dõi kỷ luật
            if self.report.key == "tdkl":
                sheet.cell(row=row, column=1).value = stt
                sheet.cell(row=row, column=2).value = record.employee_code or ''
                sheet.cell(row=row, column=3).value = record.person_discipline or ''
                sheet.cell(row=row, column=4).value = record.address or ''
                sheet.cell(row=row, column=5).value = record.department_id.name or ''
                sheet.cell(row=row, column=6).value = record.person_discipline_job.name or ''
                sheet.cell(row=row, column=7).value = record.discipline_number or ''
                sheet.cell(row=row, column=8).value = str(record.date_sign) if record.date_sign else ''
                sheet.cell(row=row, column=9).value = str(record.date_start) if record.date_start else ''
                sheet.cell(row=row, column=10).value = record.form_discipline.name or ''
                sheet.cell(row=row, column=11).value = str(record.date_end) if record.date_end else ''
                sheet.cell(row=row, column=12).value = str(record.amount) or ''
                sheet.cell(row=row, column=13).value = record.form_discipline_properties or ''
                sheet.cell(row=row, column=14).value = record.reason or ''
                sheet.cell(row=row, column=15).value = ''
            # Báo cáo theo dõi khen thưởng
            elif self.report.key == "tdkt":
                sheet.cell(row=row, column=1).value = stt
                sheet.cell(row=row, column=2).value = record.employee_code or ''
                sheet.cell(row=row, column=3).value = record.person_reward or ''
                sheet.cell(row=row, column=4).value = record.address or ''
                sheet.cell(row=row, column=5).value = record.department_id.name or ''
                sheet.cell(row=row, column=6).value = record.person_reward_job.name or ''
                sheet.cell(row=row, column=7).value = str(record.desision_number) or ''
                sheet.cell(row=row, column=8).value = str(record.sign_date) if record.sign_date else ''
                sheet.cell(row=row, column=9).value = str(record.effective_date) if record.effective_date else ''
                sheet.cell(row=row, column=10).value = record.get_option_label() or ''
                sheet.cell(row=row, column=11).value = record.title_reward.name or ''
                sheet.cell(row=row, column=12).value = record.reason or ''
                sheet.cell(row=row, column=13).value = record.note or ''
                sheet.cell(row=row, column=14).value = record.amount or ''
                sheet.cell(row=row, column=15).value = record.level_reward.name or ''
                sheet.cell(row=row, column=16).value = record.sign_person.name or ''
                sheet.cell(row=row, column=17).value = ''
            # Báo cáo tổng hợp kỷ luật, bồi thường
            elif self.report.key == "thklbt":
                sheet.cell(row=row, column=1).value = stt
                sheet.cell(row=row, column=2).value = record.person_discipline or ''
                sheet.cell(row=row, column=3).value = record.department_id.name or ''
                sheet.cell(row=row, column=4).value = record.amount or ''
                sheet.cell(row=row, column=5).value = record.reason or ''
                sheet.cell(row=row, column=6).value = str(record.date_start) if record.date_start else ''
                sheet.cell(row=row, column=7).value = str(record.discipline_number) or ''
                sheet.cell(row=row, column=8).value = ''
            # Báo cáp quyết định miễn nhiệm nhân sự
            elif self.report.key == "qdmn":
                sheet.cell(row=row, column=1).value = stt
                sheet.cell(row=row, column=2).value = record.employee_code or ''
                sheet.cell(row=row, column=3).value = record.person_discipline or ''
                sheet.cell(row=row, column=4).value = record.address or ''
                sheet.cell(row=row, column=5).value = record.department_id.name or ''
                sheet.cell(row=row, column=6).value = str(record.discipline_number) or ''
                sheet.cell(row=row, column=7).value = str(record.date_sign) if record.date_sign else ''
                sheet.cell(row=row, column=8).value = str(record.date_start) if record.date_start else ''
                sheet.cell(row=row, column=9).value = record.sign_person.name or ''
                sheet.cell(row=row, column=10).value = record.note or ''
            # Báo cáo quyết định nghỉ việc
            elif self.report.key == "qdnv":
                sheet.cell(row=row, column=1).value = stt
                sheet.cell(row=row, column=2).value = record.employee_code or ''
                sheet.cell(row=row, column=3).value = record.person_discipline or ''
                sheet.cell(row=row, column=4).value = record.address or ''
                sheet.cell(row=row, column=5).value = record.department_id.name or ''
                sheet.cell(row=row, column=6).value = str(record.discipline_number) or ''
                sheet.cell(row=row, column=7).value = str(record.date_sign) if record.date_sign else ''
                sheet.cell(row=row, column=8).value = str(record.date_start) if record.date_start else ''
                sheet.cell(row=row, column=9).value = record.sign_person.name or ''
                sheet.cell(row=row, column=10).value = record.reason or ''
                sheet.cell(row=row, column=11).value = record.note or ''
            # Báo cáo quyết định điều chuyển
            elif self.report.key == "qddc":
                sheet.cell(row=row, column=1).value = stt
                sheet.cell(row=row, column=2).value = record.employee_code or ''
                sheet.cell(row=row, column=3).value = record.person_discipline or ''
                sheet.cell(row=row, column=4).value = record.department_id.name or ''
                sheet.cell(row=row, column=5).value = record.job_id.name or ''
                sheet.cell(row=row, column=6).value = str(record.date_sign) if record.date_sign else ''
                sheet.cell(row=row, column=7).value = str(record.discipline_number) or ''
                sheet.cell(row=row, column=8).value = str(record.date_start) if record.date_start else ''
                sheet.cell(row=row, column=9).value = ''
                sheet.cell(row=row, column=10).value = ''
                sheet.cell(row=row, column=11).value = ''
                sheet.cell(row=row, column=12).value = ''
                sheet.cell(row=row, column=13).value = ''
                sheet.cell(row=row, column=14).value = record.sign_person.name or ''
                sheet.cell(row=row, column=15).value = record.note or ''

            row += 1
            stt += 1
        if self.report.key == "thklbt":
            stt = 1
            row = 5
            list_department = list_record.mapped('department_id')
            for department in list_department:
                count_emp_discipline = len(list_record.filtered(lambda x: x.department_id.id == department.id))
                sheet.cell(row=row, column=10).value = stt
                sheet.cell(row=row, column=11).value = department.name
                sheet.cell(row=row, column=12).value = count_emp_discipline
                row += 1
                stt += 1


            # Lưu lại file kết quả
        for row in sheet.iter_rows(min_row=row_min, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_alignment
        output_stream = io.BytesIO()
        wb.save(output_stream)
        output_stream.seek(0)

        # Trả về file để tải về
        export_file = base64.b64encode(output_stream.read())
        download_filename = f"{self.report.name}.xlsx"
        self.template_file = export_file

        # Tạo record ảo chỉ để tải file
        return {
            'type': 'ir.actions.act_url',
            'url': f"/web/content/export.report.reward.discipline/{self.id}/template_file/{download_filename}?download=true",
            'target': 'self',
        }
