from odoo import models, fields
from odoo.exceptions import UserError, ValidationError
from datetime import datetime
from dateutil.relativedelta import relativedelta
import base64
import io
import tempfile
import os
from openpyxl.styles import Border, Side, Alignment
from openpyxl import load_workbook
from collections import defaultdict
from openpyxl.utils import get_column_letter


class ExportReportEmployee(models.TransientModel):
    _name = 'export.report.employee'

    report = fields.Many2one('config.template.report', string="Chọn báo cáo",
                             domain="[('apply', '=', res_model)]")
    template_file = fields.Binary("Tệp Excel")
    res_model = fields.Many2one('ir.model')

    def action_export(self):
        active_ids = self.env.context.get('active_ids', [])
        if not active_ids:
            raise ValidationError("Không có bản ghi nào được chọn.")

        list_record = self.env['employee.report'].browse(active_ids)
        file_stream = io.BytesIO(base64.b64decode(self.report.file))
        wb = load_workbook(filename=file_stream)
        sheet = wb['Sheet1']
        stt = 1
        if self.report.key == "bdns" or self.report.key == "ttc" or self.report.key == "nsthd":
            row = 7
            row_min = 6
        elif self.report.key == "nsct":
            row = 7
            row_min = 6
        elif self.report.key == "nsnv":
            row = 8
            row_min = 7
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        now = datetime.now().date()
        filter_report = self.env['popup.sonha.employee.report'].sudo().search([], order='id desc', limit=1)
        if list_record[0].state == 'contract':
            filter_report = self.env['popup.sonha.contract.report'].sudo().search([], order='id desc', limit=1)
        start_date = filter_report.from_date.strftime('%d/%m/%Y')
        end_date = filter_report.to_date.strftime('%d/%m/%Y')
        report_date = now.day
        report_month = now.month
        report_year = now.year
        for record in list_record:
            if record.tinh_trang_hon_nhan:
                tinh_trang_hon_nhan = 'Độc thân' if record.tinh_trang_hon_nhan == 'single' else 'Đã kết hôn'
            else:
                tinh_trang_hon_nhan = ''
            if record.ton_giao:
                ton_giao = 'Có' if record.ton_giao == 'yes' else 'Không'
            else:
                ton_giao = ''
            if record.gioi_tinh:
                if record.gioi_tinh == 'male':
                    gioi_tinh = 'Nam'
                elif record.gioi_tinh == 'female':
                    gioi_tinh = 'Nữ'
                else:
                    gioi_tinh = 'Khác'
            else:
                gioi_tinh = ''
            # Báo cáo biến động nhân sự
            if self.report.key == "bdns":
                date_birthday = record.ngay_sinh.strftime('%d/%m/%Y') if record.ngay_sinh else ""
                onboard = record.ngay_vao_cong_ty.strftime('%d/%m/%Y') if record.ngay_vao_cong_ty else ""
                contract_date_end = record.id_ns.contract_id.date_end.strftime('%d/%m/%Y') if record.id_ns.contract_id.date_end else ""
                date_quit = record.ngay_nghi_viec.strftime('%d/%m/%Y') if record.ngay_nghi_viec else ""
                sheet.cell(row=row, column=1).value = stt
                sheet.cell(row=row, column=2).value = record.ma_nhan_vien or ''
                sheet.cell(row=row, column=3).value = record.ten_nhan_vien or ''
                sheet.cell(row=row, column=4).value = record.bac_nhan_su or ''
                sheet.cell(row=row, column=5).value = onboard
                sheet.cell(row=row, column=6).value = record.ngay_vao or ''
                sheet.cell(row=row, column=7).value = record.thang_vao or ''
                sheet.cell(row=row, column=8).value = record.nam_vao or ''
                sheet.cell(row=row, column=9).value = record.phong_ban or ''
                sheet.cell(row=row, column=10).value = record.chuc_vu.name or ''
                sheet.cell(row=row, column=11).value = record.nhom_luong or ''
                sheet.cell(row=row, column=12).value = record.khoi_me or ''
                sheet.cell(row=row, column=13).value = record.khoi_con or ''
                sheet.cell(row=row, column=14).value = record.trinh_do_van_hoa or ''
                sheet.cell(row=row, column=15).value = date_birthday
                sheet.cell(row=row, column=16).value = record.ngay_sinh.year or ''
                sheet.cell(row=row, column=17).value = record.do_tuoi or ''
                sheet.cell(row=row, column=18).value = record.so_nam_tham_nien or ''
                sheet.cell(row=row, column=19).value = record.tham_nien or ''
                sheet.cell(row=row, column=20).value = gioi_tinh
                sheet.cell(row=row, column=21).value = record.so_cccd or ''
                sheet.cell(row=row, column=22).value = record.dia_chi_thuong_tru or ''
                sheet.cell(row=row, column=23).value = ''
                sheet.cell(row=row, column=24).value = record.ten_hd or ''
                sheet.cell(row=row, column=25).value = contract_date_end
                sheet.cell(row=row, column=26).value = date_quit
                sheet.cell(row=row, column=27).value = record.ngay_nghi or ''
                sheet.cell(row=row, column=28).value = record.thang_nghi or ''
                sheet.cell(row=row, column=29).value = record.nam_nghi or ''
                sheet.cell(row=row, column=30).value = record.email or ''
                sheet.cell(row=row, column=31).value = record.sdt or ''
            # Báo cáo nhân sự chi tiết
            elif self.report.key == "nsct":
                total_emp = str(len(active_ids))
                sheet.cell(row=5, column=3).value = sheet.cell(row=5, column=3).value.replace('[total_emp]', total_emp)
                reception_date = record.ngay_tiep_nhan.strftime('%d/%m/%Y') if record.ngay_tiep_nhan else ""
                onboard = record.ngay_vao_cong_ty.strftime('%d/%m/%Y') if record.ngay_vao_cong_ty else ""
                date_birthday = record.ngay_sinh.strftime('%d/%m/%Y') if record.ngay_sinh else ""
                date_cccd = record.ngay_cap.strftime('%d/%m/%Y') if record.ngay_cap else ""
                sheet.cell(row=row, column=1).value = stt
                sheet.cell(row=row, column=2).value = record.ma_cham_cong or ''
                sheet.cell(row=row, column=3).value = record.ma_nhan_vien or ''
                sheet.cell(row=row, column=4).value = record.ten_nhan_vien or ''
                sheet.cell(row=row, column=5).value = onboard
                sheet.cell(row=row, column=6).value = reception_date
                sheet.cell(row=row, column=7).value = record.chuc_vu.name or ''
                sheet.cell(row=row, column=8).value = record.phong_ban or ''
                sheet.cell(row=row, column=9).value = gioi_tinh
                sheet.cell(row=row, column=10).value = date_birthday
                sheet.cell(row=row, column=11).value = tinh_trang_hon_nhan
                sheet.cell(row=row, column=12).value = record.dan_toc or ''
                sheet.cell(row=row, column=13).value = ton_giao
                sheet.cell(row=row, column=14).value = record.nguyen_quan or ''
                sheet.cell(row=row, column=15).value = record.dia_chi_thuong_tru or ''
                sheet.cell(row=row, column=16).value = record.noi_o_hien_tai or ''
                sheet.cell(row=row, column=17).value = record.sdt or ''
                sheet.cell(row=row, column=18).value = record.so_cccd or ''
                sheet.cell(row=row, column=19).value = date_cccd
                sheet.cell(row=row, column=20).value = record.noi_cap or ''
                sheet.cell(row=row, column=21).value = record.ma_so_thue or ''
                sheet.cell(row=row, column=22).value = record.bac_nhan_su or ''
                sheet.cell(row=row, column=23).value = record.tham_nien or ''
                sheet.cell(row=row, column=24).value = record.trinh_do_vi_tinh or ''
                sheet.cell(row=row, column=25).value = record.trinh_do_van_hoa or ''
                sheet.cell(row=row, column=26).value = record.ngoai_ngu or ''
                sheet.cell(row=row, column=27).value = record.trinh_do_ngoai_ngu or ''
                sheet.cell(row=row, column=28).value = record.ten_hd or ''
            # Báo cáo nhân sự nghỉ việc
            elif self.report.key == "nsnv":
                reception_date = record.ngay_tiep_nhan.strftime('%d/%m/%Y') if record.ngay_tiep_nhan else ""
                onboard = record.ngay_vao_cong_ty.strftime('%d/%m/%Y') if record.ngay_vao_cong_ty else ""
                date_birthday = record.ngay_sinh.strftime('%d/%m/%Y') if record.ngay_sinh else ""
                date_cccd = record.ngay_cap.strftime('%d/%m/%Y') if record.ngay_cap else ""
                date_quit = record.ngay_nghi_viec.strftime('%d/%m/%Y') if record.ngay_nghi_viec else ""
                sheet.cell(row=row, column=1).value = stt
                sheet.cell(row=row, column=2).value = record.ma_nhan_vien or ''
                sheet.cell(row=row, column=3).value = record.ten_nhan_vien or ''
                sheet.cell(row=row, column=4).value = date_birthday
                sheet.cell(row=row, column=5).value = gioi_tinh
                sheet.cell(row=row, column=6).value = tinh_trang_hon_nhan
                sheet.cell(row=row, column=7).value = record.dan_toc or ''
                sheet.cell(row=row, column=8).value = ton_giao
                sheet.cell(row=row, column=9).value = record.nguyen_quan or ''
                sheet.cell(row=row, column=10).value = record.dia_chi_thuong_tru or ''
                sheet.cell(row=row, column=11).value = record.noi_o_hien_tai or ''
                sheet.cell(row=row, column=12).value = str(record.sdt) if record.sdt else ''
                sheet.cell(row=row, column=13).value = record.so_cccd or ''
                sheet.cell(row=row, column=14).value = date_cccd
                sheet.cell(row=row, column=15).value = record.noi_cap or ''
                sheet.cell(row=row, column=16).value = record.trinh_do_vi_tinh or ''
                sheet.cell(row=row, column=17).value = record.ngoai_ngu or ''
                sheet.cell(row=row, column=18).value = record.trinh_do_ngoai_ngu or ''
                sheet.cell(row=row, column=19).value = record.ten_hd or ''
                sheet.cell(row=row, column=20).value = record.don_vi.name or ''
                sheet.cell(row=row, column=21).value = record.phong_ban or ''
                sheet.cell(row=row, column=22).value = record.chuc_vu.name or ''
                sheet.cell(row=row, column=23).value = onboard
                sheet.cell(row=row, column=24).value = reception_date
                sheet.cell(row=row, column=25).value = ''
                sheet.cell(row=row, column=26).value = date_quit
                sheet.cell(row=row, column=27).value = date_quit
                sheet.cell(row=row, column=28).value = record.id_ns.reason_quit or ''
                sheet.cell(row=row, column=29).value = ''
            # Báo cáo nhân sự theo hợp đồng
            elif self.report.key == "nsthd":
                contract_date_end = record.id_ns.contract_id.date_end.strftime('%d/%m/%Y') if record.id_ns.contract_id.date_end else ""
                if record.state == 'employee':
                    cong_tac_vien = record.cong_tac_vien
                    hd_thoi_vu = record.hd_thoi_vu
                    hd_thu_viec = record.hd_thu_viec
                    co_xd_thoi_han = record.co_xd_thoi_han
                    khong_xd_thoi_han = record.khong_xd_thoi_han
                    khac = record.khac
                else:
                    cong_tac_vien = 'X' if record.hop_dong_id.id == 11 and record.hop_dong_id else ''
                    hd_thoi_vu = 'X' if record.hop_dong_id.id == 12 and record.hop_dong_id else ''
                    hd_thu_viec = 'X' if record.hop_dong_id.id == 6 and record.hop_dong_id else ''
                    co_xd_thoi_han = 'X' if record.hop_dong_id.id == 9 and record.hop_dong_id else ''
                    khong_xd_thoi_han = 'X' if record.hop_dong_id.id == 10 and record.hop_dong_id else ''
                    khac = 'X' if record.hop_dong_id.id not in [11, 12, 6, 9, 10] and record.hop_dong_id else ''
                sheet.cell(row=row, column=1).value = stt
                sheet.cell(row=row, column=2).value = record.ma_nhan_vien or ''
                sheet.cell(row=row, column=3).value = record.ten_nhan_vien or ''
                sheet.cell(row=row, column=4).value = record.don_vi.name or ''
                sheet.cell(row=row, column=5).value = record.phong_ban or ''
                sheet.cell(row=row, column=6).value = record.chuc_vu.name or ''
                sheet.cell(row=row, column=7).value = cong_tac_vien or ''
                sheet.cell(row=row, column=8).value = hd_thoi_vu or ''
                sheet.cell(row=row, column=9).value = hd_thu_viec or ''
                sheet.cell(row=row, column=10).value = co_xd_thoi_han or ''
                sheet.cell(row=row, column=11).value = khong_xd_thoi_han or ''
                sheet.cell(row=row, column=12).value = khac or ''
                sheet.cell(row=row, column=13).value = ''
                sheet.cell(row=row, column=14).value = ''
                sheet.cell(row=row, column=15).value = ''
                sheet.cell(row=row, column=16).value = contract_date_end
                sheet.cell(row=row, column=17).value = ''
            if self.report.key != "ttc":
                row += 1
                sheet.insert_rows(row)
                stt += 1
        # Báo cáo thông tin chung
        if self.report.key == "ttc":
            list_department = list_record.mapped('phong_ban_id')
            for department in list_department:
                department_records = list_record.filtered(lambda x: x.phong_ban_id.id == department.id)
                count_male = len(department_records.filtered(lambda x: x.gioi_tinh == 'male'))
                count_female = len(department_records.filtered(lambda x: x.gioi_tinh == 'female'))
                count_thpt = len(department_records.filtered(lambda x: "trung học phổ thông" in str(x.trinh_do_van_hoa).lower()))
                count_intermediate = len(department_records.filtered(lambda x: "trung cấp" in str(x.trinh_do_van_hoa).lower()))
                count_college = len(department_records.filtered(lambda x: "cao đẳng" in str(x.trinh_do_van_hoa).lower()))
                count_university = len(department_records.filtered(lambda x: "đại học" in str(x.trinh_do_van_hoa).lower()))
                count_master = len(department_records.filtered(lambda x: "thạc sỹ" in str(x.trinh_do_van_hoa).lower()))
                count_deputy_phd = len(department_records.filtered(lambda x: "phó tiến sỹ" in str(x.trinh_do_van_hoa).lower()))
                count_phd = len(department_records.filtered(lambda x: "tiến sỹ" in str(x.trinh_do_van_hoa).lower() and not "phó tiến sỹ" in str(x.trinh_do_van_hoa).lower()))
                count_other_level = len(department_records.filtered(lambda x: str(x.trinh_do_van_hoa) == ''))
                count_deputy_pro = len(department_records.filtered(lambda x: "phó giáo sư" in str(x.trinh_do_van_hoa).lower()))
                count_pro = len(department_records.filtered(lambda x: "giáo sư" in str(x.trinh_do_van_hoa).lower() and not "phó giáo sư" in str(x.trinh_do_van_hoa).lower()))
                count_under_25 = len(department_records.filtered(lambda x: x.do_tuoi and x.do_tuoi < 25))
                count_25_30 = len(department_records.filtered(lambda x: x.do_tuoi and 25 <= x.do_tuoi <= 30))
                count_31_40 = len(department_records.filtered(lambda x: x.do_tuoi and 31 <= x.do_tuoi <= 40))
                count_41_50 = len(department_records.filtered(lambda x: x.do_tuoi and 41 <= x.do_tuoi <= 50))
                count_above_50 = len(department_records.filtered(lambda x: x.do_tuoi and x.do_tuoi > 50))
                count_single = len(department_records.filtered(lambda x: x.tinh_trang_hon_nhan == 'single'))
                count_married = len(department_records.filtered(lambda x: x.tinh_trang_hon_nhan == 'married'))
                if record.state == 'employee':
                    count_tern_contract = len(department_records.filtered(lambda x: "X" in str(x.khong_xd_thoi_han) and x.khong_xd_thoi_han))
                    count_indefinite_contract = len(department_records.filtered(lambda x: "X" in str(x.co_xd_thoi_han) and x.co_xd_thoi_han))
                    count_probation_contract = len(department_records.filtered(lambda x: "X" in str(x.hd_thu_viec) and x.hd_thu_viec))
                    count_temporary_contract = len(department_records.filtered(lambda x: "X" in str(x.hd_thoi_vu) and x.hd_thoi_vu))
                    count_collaborator_contract = len(department_records.filtered(lambda x: "X" in str(x.cong_tac_vien) and x.cong_tac_vien))
                    other_contract = len(department_records.filtered(lambda x: "X" in str(x.khac) and x.khac))
                else:
                    count_tern_contract = len(department_records.filtered(lambda x: x.hop_dong_id.id == 10 and x.hop_dong_id))
                    count_indefinite_contract = len(department_records.filtered(lambda x: x.hop_dong_id.id == 9 and x.hop_dong_id))
                    count_probation_contract = len(department_records.filtered(lambda x: x.hop_dong_id.id == 6 and x.hop_dong_id))
                    count_temporary_contract = len(department_records.filtered(lambda x: x.hop_dong_id.id == 12 and x.hop_dong_id))
                    count_collaborator_contract = len(department_records.filtered(lambda x: x.hop_dong_id.id == 11 and x.hop_dong_id))
                    other_contract = len(department_records.filtered(lambda x: x.hop_dong_id.id not in [10, 9, 6, 12, 11] and x.hop_dong_id))

                sheet.cell(row=row, column=1).value = stt
                sheet.cell(row=row, column=2).value = department.company_id.name or ''
                sheet.cell(row=row, column=3).value = department.name or ''
                sheet.cell(row=row, column=4).value = len(department_records)
                sheet.cell(row=row, column=5).value = count_male
                sheet.cell(row=row, column=6).value = count_female
                sheet.cell(row=row, column=7).value = count_thpt
                sheet.cell(row=row, column=8).value = count_intermediate
                sheet.cell(row=row, column=9).value = count_college
                sheet.cell(row=row, column=10).value = count_university
                sheet.cell(row=row, column=11).value = count_master
                sheet.cell(row=row, column=12).value = count_deputy_phd
                sheet.cell(row=row, column=13).value = count_phd
                sheet.cell(row=row, column=14).value = count_other_level
                sheet.cell(row=row, column=15).value = count_deputy_pro
                sheet.cell(row=row, column=16).value = count_pro
                sheet.cell(row=row, column=17).value = count_under_25
                sheet.cell(row=row, column=18).value = count_25_30
                sheet.cell(row=row, column=19).value = count_31_40
                sheet.cell(row=row, column=20).value = count_41_50
                sheet.cell(row=row, column=21).value = count_above_50
                sheet.cell(row=row, column=22).value = count_single
                sheet.cell(row=row, column=23).value = count_married
                sheet.cell(row=row, column=24).value = count_indefinite_contract
                sheet.cell(row=row, column=25).value = count_tern_contract
                sheet.cell(row=row, column=26).value = count_probation_contract
                sheet.cell(row=row, column=27).value = count_temporary_contract
                sheet.cell(row=row, column=28).value = count_collaborator_contract
                sheet.cell(row=row, column=29).value = other_contract
                sheet.cell(row=row, column=30).value = ''

                row += 1
                stt += 1

            # Lưu lại file kết quả
        max_lengths = defaultdict(int)
        for row in sheet.iter_rows(min_row=row_min, max_row=row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_alignment
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_lengths[cell.column]:
                        max_lengths[cell.column] = length

                # Cập nhật chiều rộng sau khi đã tính xong
        for col_idx, max_length in max_lengths.items():
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = max_length + 2
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                if isinstance(cell.value, str) and '[start_date]' in cell.value:
                    cell.value = cell.value.replace('[start_date]', start_date)
                if isinstance(cell.value, str) and '[end_date]' in cell.value:
                    cell.value = cell.value.replace('[end_date]', end_date)
                if isinstance(cell.value, str) and '[report_date]' in cell.value:
                    cell.value = cell.value.replace('[report_date]', str(report_date))
                    cell.value = cell.value.replace('[report_month]', str(report_month))
                    cell.value = cell.value.replace('[report_year]', str(report_year))
        output_stream = io.BytesIO()
        wb.save(output_stream)
        output_stream.seek(0)

        # Trả về file để tải về
        export_file = base64.b64encode(output_stream.read())
        download_filename = f"{self.report.name}.xlsx"
        self.template_file = export_file

        # Tạo record ảo chỉ để tải file
        return {
            'type': 'ir.actions.act_url',
            'url': f"/web/content/export.report.employee/{self.id}/template_file/{download_filename}?download=true",
            'target': 'self',
        }
