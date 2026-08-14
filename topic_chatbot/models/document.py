# -*- coding: utf-8 -*-
import base64
import io
import logging
import time
import zipfile
import xml.etree.ElementTree as ET
from odoo import api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

# Try importing pypdf / PyPDF2
try:
    from pypdf import PdfReader
except ImportError:
    try:
        from PyPDF2 import PdfReader
    except ImportError:
        PdfReader = None

class TopicChatbotDocument(models.Model):
    _name = 'topic_chatbot.document'
    _description = 'Topic Document'
    _order = 'create_date desc'  # Show newest documents first
    
    # Constants
    STALE_PROCESSING_MINUTES = 60
    DOCUMENT_PROCESS_LOCK_KEY = 830917
    
    # Basic fields with enhanced validation
    name = fields.Char(
        string='Document Name', 
        required=True,
        help="Descriptive name for the document"
    )
    topic_id = fields.Many2one(
        'topic_chatbot.topic', 
        string='Topic', 
        required=True, 
        ondelete='cascade',
        index=True  # Performance improvement
    )
    datas = fields.Binary(
        string='File Content', 
        required=True, 
        attachment=True,
        help="Binary content of the uploaded file"
    )
    filename = fields.Char(
        string='Filename',
        help="Original filename with extension"
    )
    
    # Enhanced content fields
    text_content = fields.Text(
        string='Extracted Text', 
        readonly=True,
        help="Text content extracted from the document"
    )
    content_length = fields.Integer(
        string='Content Length',
        readonly=True,
        help="Number of characters in extracted text"
    )
    word_count = fields.Integer(
        string='Word Count',
        readonly=True,
        help="Approximate number of words in document"
    )
    
    # Enhanced state management
    state = fields.Selection([
        ('draft', 'Draft'),
        ('processing', 'Processing'),
        ('done', 'Done'),
        ('error', 'Error')
    ], string='Status', default='draft', required=True, readonly=True, index=True)
    
    # Metadata fields
    layout_type = fields.Char(
        string='Layout Type', 
        readonly=True,
        help="Detected document layout (prose, table, mixed, etc.)"
    )
    file_size = fields.Integer(
        string='File Size (KB)',
        readonly=True,
        help="Original file size in kilobytes"
    )
    processing_time = fields.Float(
        string='Processing Time (s)',
        readonly=True,
        help="Time taken to process the document in seconds"
    )
    error_message = fields.Text(
        string='Error Details',
        readonly=True,
        help="Detailed error message if processing failed"
    )
    
    # Statistics
    chunks_count = fields.Integer(
        string='Chunks Created',
        compute='_compute_chunks_count',
        store=True,
        help="Number of text chunks created from this document"
    )

    @api.depends('chunk_ids')
    def _compute_chunks_count(self):
        """Compute number of chunks created from this document."""
        for doc in self:
            doc.chunks_count = len(doc.chunk_ids)
    
    chunk_ids = fields.One2many(
        'topic_chatbot.chunk',
        'document_id',
        string='Text Chunks',
        readonly=True,
        help="Text chunks extracted from this document"
    )

    @api.constrains('filename', 'datas')
    def _check_file_extension(self):
        """Validate file extension and size."""
        ALLOWED_EXTENSIONS = ('.pdf', '.docx', '.xlsx', '.xls', '.csv', '.txt')
        MAX_FILE_SIZE_MB = 50  # 50MB limit
        
        for doc in self:
            if doc.filename and doc.datas:
                filename_lower = doc.filename.lower()
                
                # Check file extension
                if not filename_lower.endswith(ALLOWED_EXTENSIONS):
                    raise UserError(
                        f"Tệp '{doc.filename}' không đúng định dạng. "
                        "Hệ thống chỉ chấp nhận tệp định dạng: " + 
                        ", ".join(ALLOWED_EXTENSIONS) + "!"
                    )
                
                # Check file size
                if doc.datas:
                    import base64
                    file_size_mb = len(base64.b64decode(doc.datas)) / (1024 * 1024)
                    if file_size_mb > MAX_FILE_SIZE_MB:
                        raise UserError(
                            f"Tệp '{doc.filename}' có kích thước {file_size_mb:.1f}MB "
                            f"vượt quá giới hạn {MAX_FILE_SIZE_MB}MB cho phép!"
                        )

    @api.model_create_multi
    def create(self, vals_list):
        """Enhanced create with metadata calculation."""
        import base64
        
        for vals in vals_list:
            if 'datas' in vals or 'filename' in vals:
                vals['state'] = 'draft'
                
                # Calculate file size
                if vals.get('datas'):
                    try:
                        file_content = base64.b64decode(vals['datas'])
                        vals['file_size'] = len(file_content) // 1024  # KB
                    except Exception:
                        pass
                        
        records = super().create(vals_list)
        return records

    def write(self, vals):
        """Enhanced write with state management."""
        import base64
        
        if 'datas' in vals or 'filename' in vals:
            vals['state'] = 'draft'
            
            # Recalculate file size if file changed
            if vals.get('datas'):
                try:
                    file_content = base64.b64decode(vals['datas'])
                    vals['file_size'] = len(file_content) // 1024  # KB
                except Exception:
                    pass
                    
        return super().write(vals)

    def action_process_document(self):
        """Action method to trigger background processing on selected document(s)."""
        import threading
        stale_cutoff = fields.Datetime.subtract(
            fields.Datetime.now(),
            minutes=self.STALE_PROCESSING_MINUTES,
        )
        draft_docs = self.filtered(
            lambda d: d.state in ('draft', 'error') or (
                d.state == 'processing' and d.write_date and d.write_date <= stale_cutoff
            )
        )
        if not draft_docs:
            return True

        draft_docs.filtered(lambda d: d.state == 'error').write({'state': 'draft'})

        db_name = self.env.cr.dbname
        uid = self.env.uid
        doc_ids = draft_docs.ids

        thread = threading.Thread(
            target=self._run_process_documents_in_thread,
            args=(db_name, uid, doc_ids),
            daemon=True
        )
        thread.start()

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Xử lý tài liệu',
                'message': f'Đã bắt đầu xử lý {len(draft_docs)} tài liệu trong nền. Trạng thái sẽ cập nhật tự động khi hoàn thành.',
                'type': 'info',
                'sticky': False,
            }
        }

    @classmethod
    def _run_process_documents_in_thread(cls, db_name, uid, doc_ids):
        """Worker function for running document processing in a separate DB cursor thread."""
        import odoo
        with odoo.registry(db_name).cursor() as new_cr:
            env = api.Environment(new_cr, uid, {})
            documents = env['topic_chatbot.document'].browse(doc_ids)
            for doc in documents:
                try:
                    doc._process_document()
                    new_cr.commit()
                except Exception as e:
                    _logger.error("Background document processing error for doc id %s: %s", doc.id, str(e))
                    new_cr.rollback()
                    try:
                        env['topic_chatbot.document'].browse(doc.id).sudo().write({
                            'state': 'error',
                            'text_content': "Lỗi xử lý nền: %s" % str(e),
                        })
                        new_cr.commit()
                    except Exception as write_err:
                        new_cr.rollback()
                        _logger.error(
                            "Failed to mark document %s as error after background failure: %s",
                            doc.id,
                            str(write_err),
                        )

    @api.model
    def _cron_process_documents(self):
        """Cron job to process draft documents and recover stale processing records."""
        stale_cutoff = fields.Datetime.subtract(
            fields.Datetime.now(),
            minutes=self.STALE_PROCESSING_MINUTES,
        )
        documents = self.search([
            '|',
            ('state', '=', 'draft'),
            '&',
            ('state', '=', 'processing'),
            ('write_date', '<=', stale_cutoff),
        ], limit=5)
        for doc in documents:
            if doc.state == 'processing':
                _logger.warning(
                    "Reprocessing stale topic chatbot document %s (id=%s), last update: %s",
                    doc.name,
                    doc.id,
                    doc.write_date,
                )
            doc._process_document()

    def _process_document(self):
        """Enhanced document processing with better error handling and statistics."""
        import time
        import base64
        
        for doc in self:
            start_time = time.time()
            
            # Advisory lock to prevent concurrent processing
            doc.env.cr.execute(
                "SELECT pg_try_advisory_xact_lock(%s, %s)",
                (self.DOCUMENT_PROCESS_LOCK_KEY, doc.id),
            )
            lock_acquired = doc.env.cr.fetchone()[0]
            if not lock_acquired:
                _logger.info(
                    "Skipping document %s (id=%s) because another worker is already processing it.",
                    doc.name, doc.id,
                )
                continue

            # Early return for empty documents
            if not doc.datas:
                doc.write({
                    'state': 'done',
                    'processing_time': time.time() - start_time,
                    'content_length': 0,
                    'word_count': 0
                })
                continue

            try:
                doc.write({'state': 'processing'})
                
                # Decode base64 file content
                file_content = base64.b64decode(doc.datas)
                filename = (doc.filename or '').lower()
                extracted_text = ""

                # Extract text based on file type
                if filename.endswith('.pdf'):
                    extracted_text = doc._extract_pdf_text(file_content)
                elif filename.endswith('.docx'):
                    extracted_text = doc._extract_docx_text(file_content)
                elif filename.endswith(('.xlsx', '.xls')):
                    extracted_text = doc._extract_excel_text(file_content)
                elif filename.endswith(('.txt', '.csv')):
                    extracted_text = doc._extract_txt_or_csv_text(file_content)
                else:
                    raise UserError(
                        f"Tệp '{doc.filename}' không được hỗ trợ. "
                        "Hệ thống chỉ chấp nhận tệp định dạng .pdf, .docx, .xlsx, .xls, .csv hoặc .txt!"
                    )

                # Calculate text statistics
                content_length = len(extracted_text) if extracted_text else 0
                word_count = len(extracted_text.split()) if extracted_text else 0

                # Update document with extracted content
                doc.write({
                    'text_content': extracted_text,
                    'content_length': content_length,
                    'word_count': word_count
                })
                
                # Security check for prompt injection
                doc._warn_prompt_injection_patterns(extracted_text)

                # Remove old chunks and create new ones
                doc.env['topic_chatbot.chunk'].search([('document_id', '=', doc.id)]).unlink()

                # Create new chunks with Vector Embeddings
                chunks_created = 0
                if extracted_text and len(extracted_text.strip()) > 10:
                    params = doc.env['ir.config_parameter'].sudo()
                    api_key = params.get_param('topic_chatbot.gemini_api_key')
                    embedding_model = params.get_param('topic_chatbot.embedding_model', default='gemini-embedding-2')
                    
                    # Model validation
                    if embedding_model in ('text-embedding-004', 'embedding-001') or not embedding_model:
                        embedding_model = 'gemini-embedding-2'

                    chunks = doc._chunk_text(extracted_text)
                    chunk_vals = []
                    
                    for index, chunk in enumerate(chunks, start=1):
                        emb_json = None
                        if api_key and len(chunk.strip()) > 10:
                            try:
                                emb_json = doc.env['topic_chatbot.chunk']._generate_embedding(
                                    chunk, api_key, embedding_model
                                )
                            except Exception as e:
                                _logger.warning(
                                    "Failed to generate embedding for chunk %d of document %s: %s",
                                    index, doc.name, str(e)
                                )

                        chunk_vals.append({
                            'topic_id': doc.topic_id.id,
                            'document_id': doc.id,
                            'sequence': index,
                            'content': chunk,
                            'embedding': emb_json,
                        })

                    if chunk_vals:
                        created_chunks = doc.env['topic_chatbot.chunk'].create(chunk_vals)
                        chunks_created = len(created_chunks)
                        
                        # Update pgvector column using raw SQL
                        for chunk_rec in created_chunks:
                            if chunk_rec.embedding:
                                try:
                                    doc.env.cr.execute(
                                        "UPDATE topic_chatbot_chunk SET embedding_vector = %s WHERE id = %s",
                                        (chunk_rec.embedding, chunk_rec.id)
                                    )
                                except Exception as e:
                                    _logger.warning("Failed to save pgvector for chunk %s: %s", chunk_rec.id, str(e))
                
                # Calculate processing time
                processing_time = time.time() - start_time
                
                # Final state update
                doc.write({
                    'state': 'done',
                    'processing_time': processing_time,
                    'error_message': False  # Clear any previous errors
                })
                
                _logger.info(
                    "Successfully processed document '%s' (id=%s) in %.2fs: "
                    "%d chars, %d words, %d chunks",
                    doc.name, doc.id, processing_time,
                    content_length, word_count, chunks_created
                )
                
            except Exception as e:
                processing_time = time.time() - start_time
                err_msg = str(e)
                
                _logger.error(
                    "Error processing document %s (id=%s) after %.2fs: %s", 
                    doc.name, doc.id, processing_time, err_msg
                )
                
                doc.write({
                    'state': 'error',
                    'processing_time': processing_time,
                    'error_message': err_msg,
                    'text_content': f"LỖI XỬ LÝ: {err_msg}",
                    'content_length': 0,
                    'word_count': 0
                })

    def _clean_extracted_text(self, text):
        """Clean and normalize extracted text from PDF engines.
        
        Removes excessive whitespace, fixes common OCR errors,
        and improves text readability.
        """
        if not text:
            return ""
        
        import re
        
        # Remove excessive whitespace and normalize line breaks
        text = re.sub(r'\n\s*\n\s*\n+', '\n\n', text)  # Multiple empty lines → double line break
        text = re.sub(r'[ \t]+', ' ', text)  # Multiple spaces/tabs → single space
        text = re.sub(r'\n[ \t]+', '\n', text)  # Leading whitespace after newlines
        text = re.sub(r'[ \t]+\n', '\n', text)  # Trailing whitespace before newlines
        
        # Fix common PDF extraction issues
        text = re.sub(r'(\w)-\s*\n\s*(\w)', r'\1\2', text)  # Fix hyphenated words split across lines
        text = re.sub(r'([.!?])\s*\n\s*([A-ZÀÁẢÃẠÂẦẤẨẪẬĂẰẮẲẴẶ])', r'\1\n\n\2', text)  # Sentence breaks
        
        # Clean up common OCR artifacts
        text = re.sub(r'[^\w\s\n\r.,;:!?()\[\]{}"\'+=\-*/<>@#$%^&|\\`~àáảãạâầấẩẫậăằắẳẵặèéẻẽẹêềếểễệìíỉĩịòóỏõọôồốổỗộơờớởỡợùúủũụưừứửữựỳýỷỹỵđĐ]', '', text)
        
        # Remove lines that are likely headers/footers (very short, just numbers, etc.)
        lines = text.split('\n')
        cleaned_lines = []
        for line in lines:
            line = line.strip()
            if line and not (
                len(line) <= 3 and line.isdigit() or  # Page numbers
                len(line) <= 10 and re.match(r'^[^a-zA-ZàáảãạâầấẩẫậăằắẳẵặèéẻẽẹêềếểễệìíỉĩịòóỏõọôồốổỗộơờớởỡợùúủũụưừứửữựỳýỷỹỵđĐ]*$', line)  # Symbol-only lines
            ):
                cleaned_lines.append(line)
        
        result = '\n'.join(cleaned_lines).strip()
        return result if len(result) > 10 else ""  # Ensure meaningful content

    @staticmethod
    def _is_meaningful_pdf_text(text_parts, total_pages=1):
        """Check if extracted PDF text layer is meaningful or just sparse garbage/watermarks.
        
        Enhanced detection considers:
        - Character density per page
        - Presence of actual words vs symbols
        - Text diversity and structure
        """
        if not text_parts:
            return False
            
        full_text = "\n".join(text_parts).strip()
        
        # Basic length check (enhanced thresholds)
        min_chars_overall = 100  # Increased from 50
        min_chars_per_page = 50   # Increased from 35
        
        if len(full_text) < max(min_chars_overall, total_pages * min_chars_per_page):
            return False
        
        # Advanced meaningfulness checks
        import re
        
        # Count actual words (not just symbols/numbers)
        words = re.findall(r'\b[a-zA-ZàáảãạâầấẩẫậăằắẳẵặèéẻẽẹêềếểễệìíỉĩịòóỏõọôồốổỗộơờớởỡợùúủũụưừứửữựỳýỷỹỵđĐ]{3,}\b', full_text)
        
        # Require at least 5 actual words per page on average
        if len(words) < max(10, total_pages * 5):
            return False
        
        # Check text diversity (not just repeated headers/footers)
        unique_words = set(word.lower() for word in words)
        if len(unique_words) < max(5, len(words) * 0.3):  # At least 30% word diversity
            return False
        
        # Check for common watermark/header patterns that shouldn't be considered meaningful
        watermark_patterns = [
            r'confidential',
            r'draft',
            r'page\s+\d+',
            r'copyright',
            r'proprietary',
            r'^[^a-zA-Z]*$',  # Lines with no letters
        ]
        
        meaningful_lines = []
        for line in full_text.split('\n'):
            line = line.strip()
            if len(line) > 10:  # Ignore very short lines
                is_watermark = any(re.search(pattern, line, re.IGNORECASE) for pattern in watermark_patterns)
                if not is_watermark:
                    meaningful_lines.append(line)
        
        # Require at least some meaningful lines
        return len(meaningful_lines) >= max(2, total_pages)

    def _extract_pdf_text(self, file_content):
        """Extract text from PDF using multiple engines with intelligent fallback.
        
        Strategy:
        1. Try native text extraction (fitz → pypdf → PyPDF2)
        2. Check if extracted text is meaningful
        3. Fall back to Gemini Vision OCR for scanned PDFs
        """
        text_parts = []
        total_pages = 1

        # Engine 1: PyMuPDF (fitz) - Best & fastest PDF text extractor
        try:
            import fitz
            doc_fitz = fitz.open(stream=file_content, filetype="pdf")
            total_pages = len(doc_fitz)
            
            # Enhanced text extraction with better formatting
            for page_num, page in enumerate(doc_fitz, 1):
                # Try different text extraction methods
                text_methods = [
                    lambda p: p.get_text(),  # Default method
                    lambda p: p.get_text("text"),  # Plain text
                    lambda p: p.get_text("blocks"),  # Block-based (better structure)
                ]
                
                page_text = ""
                for method in text_methods:
                    try:
                        result = method(page)
                        if isinstance(result, str):
                            page_text = result
                        elif isinstance(result, list):  # blocks method returns list
                            page_text = "\n".join([block[4] for block in result if len(block) > 4 and block[4].strip()])
                        
                        if page_text and page_text.strip():
                            break
                    except Exception:
                        continue
                
                if page_text and page_text.strip():
                    # Clean up the extracted text
                    cleaned_text = self._clean_extracted_text(page_text.strip())
                    if cleaned_text:
                        text_parts.append(f"--- Trang {page_num} ---\n{cleaned_text}")
            
            doc_fitz.close()
            
            if text_parts and self._is_meaningful_pdf_text(text_parts, total_pages):
                return "\n".join(text_parts)
                
        except ImportError:
            _logger.info("PyMuPDF (fitz) not installed. Install with: pip install PyMuPDF")
        except Exception as e:
            _logger.debug("PyMuPDF fitz extraction failed: %s", str(e))

        # Engine 2: pypdf (Modern PyPDF) - Enhanced extraction
        if not text_parts or not self._is_meaningful_pdf_text(text_parts, total_pages):
            text_parts = []
            try:
                from pypdf import PdfReader
                reader = PdfReader(io.BytesIO(file_content))
                total_pages = len(reader.pages)
                
                for page_num, page in enumerate(reader.pages, 1):
                    # Try multiple extraction strategies
                    page_text = ""
                    try:
                        # Method 1: Standard extraction
                        page_text = page.extract_text()
                        
                        # Method 2: Enhanced extraction with layout preservation
                        if not page_text or len(page_text.strip()) < 50:
                            page_text = page.extract_text(extraction_mode="layout")
                            
                    except Exception:
                        try:
                            page_text = page.extract_text()
                        except Exception:
                            continue
                    
                    if page_text and page_text.strip():
                        cleaned_text = self._clean_extracted_text(page_text.strip())
                        if cleaned_text:
                            text_parts.append(f"--- Trang {page_num} ---\n{cleaned_text}")
                
                if text_parts and self._is_meaningful_pdf_text(text_parts, total_pages):
                    return "\n".join(text_parts)
                    
            except ImportError:
                _logger.info("pypdf not installed. Install with: pip install pypdf")
            except Exception as e:
                _logger.debug("pypdf extraction failed: %s", str(e))

        # Engine 3: PyPDF2 (Enhanced legacy support)
        if not text_parts or not self._is_meaningful_pdf_text(text_parts, total_pages):
            text_parts = []
            try:
                import PyPDF2
                pdf_stream = io.BytesIO(file_content)
                
                # Handle both old and new PyPDF2 versions
                if hasattr(PyPDF2, 'PdfReader'):
                    reader = PyPDF2.PdfReader(pdf_stream)
                    total_pages = len(reader.pages)
                    
                    for page_num, page in enumerate(reader.pages, 1):
                        # Try different extraction methods
                        page_text = ""
                        for extract_method in ['extract_text', 'extractText']:
                            if hasattr(page, extract_method):
                                try:
                                    page_text = getattr(page, extract_method)()
                                    if page_text and page_text.strip():
                                        break
                                except Exception:
                                    continue
                        
                        if page_text and page_text.strip():
                            cleaned_text = self._clean_extracted_text(page_text.strip())
                            if cleaned_text:
                                text_parts.append(f"--- Trang {page_num} ---\n{cleaned_text}")
                
                elif hasattr(PyPDF2, 'PdfFileReader'):
                    # Legacy PyPDF2 version
                    reader = PyPDF2.PdfFileReader(pdf_stream)
                    
                    # Handle encrypted PDFs
                    if reader.isEncrypted:
                        try:
                            reader.decrypt('')  # Try empty password
                        except Exception:
                            _logger.warning("PDF is encrypted and cannot be decrypted")
                            pass
                    
                    total_pages = reader.getNumPages()
                    for i in range(total_pages):
                        try:
                            page = reader.getPage(i)
                            page_text = ""
                            
                            # Try different extraction methods
                            for extract_method in ['extract_text', 'extractText']:
                                if hasattr(page, extract_method):
                                    try:
                                        page_text = getattr(page, extract_method)()
                                        if page_text and page_text.strip():
                                            break
                                    except Exception:
                                        continue
                            
                            if page_text and page_text.strip():
                                cleaned_text = self._clean_extracted_text(page_text.strip())
                                if cleaned_text:
                                    text_parts.append(f"--- Trang {i + 1} ---\n{cleaned_text}")
                        except Exception as e:
                            _logger.debug("Error extracting page %d: %s", i + 1, str(e))
                            continue
                
                if text_parts and self._is_meaningful_pdf_text(text_parts, total_pages):
                    return "\n".join(text_parts)
                    
            except ImportError:
                _logger.info("PyPDF2 not installed. Install with: pip install PyPDF2")
            except Exception as e:
                _logger.debug("PyPDF2 extraction failed: %s", str(e))

        # Engine 4: Fallback to Gemini Vision API OCR for scanned image PDFs or PDFs with insufficient text layer
        safe_filename = (self.filename or '').encode('ascii', 'replace').decode('ascii')
        _logger.info(
            "PDF %s text layer is missing or sparse (pages=%s, extracted_parts=%s). Attempting Gemini Vision OCR...", 
            safe_filename, total_pages, len(text_parts)
        )
        
        # Check if Gemini API key is available before attempting OCR
        params = self.env['ir.config_parameter'].sudo()
        api_key = params.get_param('topic_chatbot.gemini_api_key')
        if not api_key:
            # If no API key but we have some sparse text, return it
            if text_parts:
                _logger.warning(
                    "Gemini API key not configured, but some text was extracted from %s. "
                    "Configure Gemini API key for better OCR of scanned PDFs.", safe_filename
                )
                return "\n".join(text_parts)
            else:
                raise UserError(
                    f"Tệp PDF '{self.filename}' không chứa lớp văn bản (Text layer) hoặc là tệp PDF dạng hình ảnh/scan. "
                    "Vui lòng cấu hình Gemini API Key trong Cài đặt → Tham số hệ thống để tự động OCR tệp PDF scan, "
                    "hoặc tải lên tệp định dạng .docx / .xlsx / .txt!"
                )
        
        try:
            ocr_text = self._extract_pdf_ocr_gemini(file_content)
            if ocr_text and len(ocr_text.strip()) > 100:  # Ensure meaningful OCR result
                return ocr_text
        except Exception as e:
            _logger.error("Gemini OCR failed for %s: %s", safe_filename, str(e))

        # If Gemini OCR failed but we had some sparse text, return it as last resort
        if text_parts:
            _logger.warning(
                "Gemini OCR failed for %s, falling back to sparse text layer extraction", 
                safe_filename
            )
            return "\n".join(text_parts)

        # If no text extracted at all
        raise UserError(
            f"Không thể trích xuất văn bản từ tệp PDF '{self.filename}'. "
            "Vui lòng kiểm tra:\n"
            "1. Tệp PDF không bị hỏng\n"
            "2. Tệp PDF không bị mã hóa (password protected)\n"
            "3. Gemini API Key đã được cấu hình đúng\n"
            "4. Hoặc chuyển đổi sang định dạng .docx / .xlsx / .txt"
        )

    # ── Gemini Vision API Helpers ─────────────────────────────────────────

    def _gemini_call_vision(self, img_b64, prompt_text, api_key, response_json=False):
        """Call Gemini Vision API with automatic model fallback and improved error handling.

        Args:
            img_b64: Base64-encoded JPEG image string.
            prompt_text: The text prompt to send alongside the image.
            api_key: Gemini API key.
            response_json: If True, request JSON output and parse the response.

        Returns:
            Parsed JSON (dict or list) when *response_json* is True,
            otherwise a raw text string. Returns ``{}`` / ``""`` on failure.
        """
        import json as json_lib
        import requests
        import time

        # Updated model list with latest Gemini models
        models_to_try = [
            'gemini-2.0-flash-exp',     # Latest experimental model
            'gemini-1.5-flash',         # Fast and reliable  
            'gemini-1.5-pro',          # High quality
            'gemini-pro-vision'         # Fallback
        ]

        for model_idx, model_name in enumerate(models_to_try):
            url = (
                f"https://generativelanguage.googleapis.com/v1beta/models/"
                f"{model_name}:generateContent?key={api_key}"
            )
            
            gen_config = {
                "temperature": 0.1,  # Low temperature for consistent OCR
                "maxOutputTokens": 8192,  # Sufficient for large tables
            }
            
            if response_json:
                gen_config["responseMimeType"] = "application/json"

            payload = {
                "contents": [{
                    "parts": [
                        {"text": prompt_text},
                        {"inline_data": {"mime_type": "image/jpeg", "data": img_b64}},
                    ]
                }],
                "generationConfig": gen_config,
                "safetySettings": [
                    {
                        "category": "HARM_CATEGORY_HARASSMENT",
                        "threshold": "BLOCK_NONE"
                    },
                    {
                        "category": "HARM_CATEGORY_HATE_SPEECH",
                        "threshold": "BLOCK_NONE"
                    },
                    {
                        "category": "HARM_CATEGORY_SEXUALLY_EXPLICIT",
                        "threshold": "BLOCK_NONE"
                    },
                    {
                        "category": "HARM_CATEGORY_DANGEROUS_CONTENT",
                        "threshold": "BLOCK_NONE"
                    }
                ]
            }

            # Adaptive retry strategy
            max_retries = 3 if model_idx == 0 else 2  # More retries for primary model
            
            for attempt in range(max_retries):
                try:
                    # Timeout increases with each retry
                    timeout = 30 + (attempt * 15)  # 30s, 45s, 60s
                    
                    res = requests.post(url, json=payload, timeout=timeout)
                    
                    if res.status_code == 200:
                        res_json = res.json()
                        candidates = res_json.get('candidates', [])
                        
                        if candidates and len(candidates) > 0:
                            content = candidates[0].get('content', {})
                            parts = content.get('parts', [])
                            
                            if parts and len(parts) > 0:
                                raw_text = parts[0].get('text', '')
                                if raw_text:
                                    if response_json:
                                        try:
                                            return json_lib.loads(raw_text)
                                        except json_lib.JSONDecodeError as e:
                                            _logger.warning(
                                                "Failed to parse JSON response from %s: %s", 
                                                model_name, str(e)
                                            )
                                            import re
                                            # Try to extract JSON from response
                                            json_match = re.search(r'\{.*\}|\[.*\]', raw_text, re.DOTALL)
                                            if json_match:
                                                try:
                                                    return json_lib.loads(json_match.group())
                                                except:
                                                    pass
                                            return {}
                                    else:
                                        return raw_text.strip()
                        
                        # Check for safety blocking
                        finish_reason = candidates[0].get('finishReason') if candidates else None
                        if finish_reason == 'SAFETY':
                            _logger.warning("Gemini API blocked content for safety reasons")
                            continue
                    
                    elif res.status_code == 429:
                        # Rate limiting - intelligent backoff
                        retry_after = res.headers.get('Retry-After')
                        try:
                            sleep_seconds = min(max(int(retry_after or 0), 5), 120)
                        except (TypeError, ValueError):
                            # Exponential backoff with jitter
                            base_delay = 5 * (2 ** attempt)
                            jitter = time.time() % 1  # Random component
                            sleep_seconds = min(base_delay + jitter, 60)
                        
                        _logger.warning(
                            "Gemini API rate limited (429) for %s (attempt %s/%s). "
                            "Sleeping %ss before retry...",
                            model_name, attempt + 1, max_retries, sleep_seconds
                        )
                        time.sleep(sleep_seconds)
                        
                        # On rate limit, try next model after first retry
                        if attempt == 0 and model_idx < len(models_to_try) - 1:
                            break
                        continue
                    
                    elif res.status_code in (400, 404):
                        # Model not available or bad request - try next model
                        _logger.warning(
                            "Gemini API error %s for model %s: %s",
                            res.status_code, model_name, res.text[:200]
                        )
                        break
                    
                    else:
                        _logger.warning(
                            "Gemini API HTTP %s (%s, attempt %s/%s): %s",
                            res.status_code, model_name, attempt + 1, max_retries, res.text[:300]
                        )
                        
                        # For server errors, retry with delay
                        if res.status_code >= 500 and attempt < max_retries - 1:
                            time.sleep(2 ** attempt)
                            continue
                        else:
                            break

                except requests.exceptions.Timeout:
                    _logger.warning(
                        "Gemini API timeout for %s (attempt %s/%s, timeout=%ss)",
                        model_name, attempt + 1, max_retries, timeout
                    )
                    if attempt < max_retries - 1:
                        time.sleep(2)  # Short delay before retry
                        continue
                    else:
                        break
                        
                except Exception as err:
                    _logger.warning(
                        "Gemini API call failed for %s (attempt %s/%s): %s",
                        model_name, attempt + 1, max_retries, str(err)
                    )
                    if attempt < max_retries - 1:
                        time.sleep(1)
                        continue
                    else:
                        break

        _logger.error("All Gemini models failed for OCR request")
        return {} if response_json else ""

    def _render_page_image(self, page, dpi=150, preprocess=False):
        """Render a *fitz* page to a PIL Image.

        When *preprocess* is True the image is converted to grayscale and
        enhanced with contrast (1.8×) and sharpness (2.0×) — useful for
        scanned tables with narrow cells.
        """
        from PIL import Image, ImageEnhance

        pix = page.get_pixmap(dpi=dpi)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

        if preprocess:
            img = img.convert('L')
            img = ImageEnhance.Contrast(img).enhance(1.8)
            img = ImageEnhance.Sharpness(img).enhance(2.0)
        return img

    @staticmethod
    def _img_to_b64(pil_img):
        """Convert a PIL Image to a base64-encoded JPEG string."""
        buf = io.BytesIO()
        pil_img.save(buf, format='JPEG', quality=92)
        return base64.b64encode(buf.getvalue()).decode('ascii')

    @staticmethod
    def _split_image_vertical(img, n_parts, overlap_px=20):
        """Split a PIL image into *n_parts* vertical strips with overlap."""
        width, height = img.size
        part_width = width // n_parts
        strips = []
        for i in range(n_parts):
            left = max(0, i * part_width - overlap_px) if i > 0 else 0
            right = min(width, (i + 1) * part_width + overlap_px) if i < n_parts - 1 else width
            strips.append(img.crop((left, 0, right, height)))
        return strips

    @staticmethod
    def _is_similar_page(ref_size, check_size, tolerance=0.05):
        """Return True if two (w, h) tuples are within *tolerance* of each other."""
        w_ratio = abs(ref_size[0] - check_size[0]) / max(ref_size[0], 1)
        h_ratio = abs(ref_size[1] - check_size[1]) / max(ref_size[1], 1)
        return w_ratio <= tolerance and h_ratio <= tolerance

    # ── Page Layout Classification ─────────────────────────────────────────

    def _classify_page_layout(self, page_image_b64, api_key):
        """Classify a page's layout using Gemini Vision at low resolution.

        Returns a dict::

            {
                "layout_type": "wide_table" | "simple_table" | "prose" | "form" | "mixed",
                "estimated_columns": int,
                "has_narrow_cells": bool,
                "column_headers": [str, ...]
            }

        Falls back to ``"prose"`` when classification fails.
        """
        CLASSIFY_PROMPT = (
            "Bạn là hệ thống phân tích bố cục tài liệu. Phân tích hình ảnh trang "
            "tài liệu này và trả về JSON với các field:\n"
            '{\n'
            '  "layout_type": "wide_table" | "simple_table" | "prose" | "form" | "mixed",\n'
            '  "estimated_columns": <int, số cột ước lượng, 0 nếu không phải bảng>,\n'
            '  "has_narrow_cells": <bool, có ô chứa ký hiệu ngắn 1-3 ký tự không>,\n'
            '  "column_headers": [<danh sách tên cột đọc được, rỗng nếu không phải bảng>]\n'
            '}\n\n'
            "Định nghĩa layout_type:\n"
            '- "wide_table": bảng >= 6 cột HOẶC có ô ký hiệu ngắn (1-3 ký tự).\n'
            '- "simple_table": bảng < 6 cột, không có ô ký hiệu ngắn.\n'
            '- "prose": văn bản đoạn văn thuần, không bảng.\n'
            '- "form": biểu mẫu có cặp label + giá trị điền.\n'
            '- "mixed": trang có cả bảng và đoạn văn.\n\n'
            "Trả về ĐÚNG JSON, không markdown, không giải thích."
        )

        fallback = {
            'layout_type': 'prose',
            'estimated_columns': 0,
            'has_narrow_cells': False,
            'column_headers': [],
        }

        try:
            result = self._gemini_call_vision(
                page_image_b64, CLASSIFY_PROMPT, api_key, response_json=True,
            )
            if isinstance(result, dict) and 'layout_type' in result:
                valid_types = ('wide_table', 'simple_table', 'prose', 'form', 'mixed')
                if result.get('layout_type') not in valid_types:
                    result['layout_type'] = 'prose'
                result.setdefault('estimated_columns', 0)
                result.setdefault('has_narrow_cells', False)
                result.setdefault('column_headers', [])
                return result
            _logger.warning("Classify returned unexpected format: %s", result)
            return fallback
        except Exception as e:
            _logger.warning("Page classification failed, falling back to prose: %s", str(e))
            return fallback

    # ── Per-layout OCR Pipeline Branches ───────────────────────────────────

    def _ocr_page_simple(self, page, api_key):
        """OCR pipeline for *prose* / *form* pages — single Gemini call."""
        img = self._render_page_image(page, dpi=150)
        img_b64 = self._img_to_b64(img)
        prompt = (
            "Trích xuất toàn bộ văn bản trong trang hình ảnh này sang định dạng "
            "markdown tiếng Việt. Giữ nguyên bố cục, đoạn văn và thứ tự nội dung."
        )
        return self._gemini_call_vision(img_b64, prompt, api_key)

    def _ocr_page_simple_table(self, page, api_key, column_headers):
        """OCR pipeline for *simple_table* pages (< 6 columns).

        Uses the *column_headers* discovered during classification to hint
        the OCR model, improving accuracy.
        """
        img = self._render_page_image(page, dpi=200)
        img_b64 = self._img_to_b64(img)

        if column_headers:
            cols_hint = ", ".join(column_headers)
            prompt = (
                f"Trích xuất bảng dữ liệu trong trang hình ảnh này sang markdown table. "
                f"Bảng có các cột: {cols_hint}. "
                "Giữ nguyên toàn bộ nội dung và văn bản đi kèm."
            )
        else:
            prompt = (
                "Trích xuất bảng dữ liệu trong trang hình ảnh này sang markdown table. "
                "Giữ nguyên toàn bộ nội dung và văn bản đi kèm."
            )
        return self._gemini_call_vision(img_b64, prompt, api_key)

    def _ocr_page_wide_table(self, page, api_key, classify_info):
        """OCR pipeline for *wide_table* pages — split, structured JSON, merge.

        The image is preprocessed (grayscale + contrast + sharpness), split
        into vertical strips, each strip is OCR'd as structured JSON, and the
        results are merged by ``row_index`` then converted to a Markdown table.
        Column names come entirely from *classify_info* — nothing is hardcoded.
        """
        estimated_cols = classify_info.get('estimated_columns', 6)
        column_headers = classify_info.get('column_headers', [])

        # Dynamic number of vertical splits
        if estimated_cols > 20:
            n_parts = 3
        elif estimated_cols >= 6:
            n_parts = 2
        else:
            n_parts = 1

        img = self._render_page_image(page, dpi=280, preprocess=True)

        if n_parts <= 1:
            # No splitting — single structured-JSON OCR call
            img_b64 = self._img_to_b64(img)
            prompt = self._build_wide_table_prompt(column_headers, 1, 1)
            rows = self._gemini_call_vision(img_b64, prompt, api_key, response_json=True)
            if isinstance(rows, list):
                return self._json_rows_to_markdown(rows, column_headers)
            return ""

        # Split image into N strips
        strips = self._split_image_vertical(img, n_parts, overlap_px=20)

        slices_data = []
        for i, strip in enumerate(strips):
            if i > 0:
                time.sleep(1)
            strip_b64 = self._img_to_b64(strip)
            prompt = self._build_wide_table_prompt(column_headers, i + 1, n_parts)
            result = self._gemini_call_vision(strip_b64, prompt, api_key, response_json=True)
            slices_data.append(result if isinstance(result, list) else [])

        merged = self._merge_sliced_rows(slices_data, column_headers)
        return self._json_rows_to_markdown(merged, column_headers)

    def _ocr_page_mixed(self, page, api_key):
        """OCR pipeline for *mixed* pages (tables + prose interleaved)."""
        img = self._render_page_image(page, dpi=200)
        img_b64 = self._img_to_b64(img)
        prompt = (
            "Trang tài liệu này chứa cả đoạn văn bản lẫn bảng dữ liệu. "
            "Trích xuất toàn bộ nội dung sang markdown, giữ nguyên thứ tự xuất hiện. "
            "Bảng dùng markdown table format. Văn bản giữ nguyên đoạn."
        )
        return self._gemini_call_vision(img_b64, prompt, api_key)

    # ── Wide-table JSON Helpers ────────────────────────────────────────────

    @staticmethod
    def _build_wide_table_prompt(column_headers, slice_idx, total_slices):
        """Build a dynamic OCR prompt for a wide-table image (or slice).

        Column names are injected from *column_headers* discovered at
        classification time — no document-specific names are hardcoded.
        """
        cols_str = ", ".join(column_headers) if column_headers else "(tự nhận diện từ ảnh)"

        slice_desc = ""
        if total_slices > 1:
            slice_desc = (
                f"Hình ảnh này là PHẦN {slice_idx}/{total_slices} (cắt dọc) của bảng. "
                "CHỈ trích xuất các cột BẠN NHÌN THẤY trong phần ảnh này.\n"
            )

        return (
            "Bạn là hệ thống OCR chuyên dụng cho bảng dữ liệu nhiều cột.\n"
            f"{slice_desc}"
            f"Bảng có các cột (theo thứ tự): {cols_str}\n\n"
            "Trích xuất MỌI hàng dữ liệu thành JSON array. Mỗi phần tử có:\n"
            '  - "row_index": số thứ tự hàng bắt đầu từ 1 (int)\n'
            '  - "cot": object chứa giá trị các cột hiện diện. '
            'Key = tên cột chính xác, value = nội dung ô (string, trống thì "").\n\n'
            "Trả về ĐÚNG JSON array, không markdown, không giải thích."
        )

    def _merge_sliced_rows(self, slices_data, column_headers):
        """Merge JSON row data from multiple vertical slices by ``row_index``.

        Each slice may contain a subset of columns.  Rows are aligned by
        their ``row_index`` field.  Missing matches are logged as warnings.
        """
        if not slices_data:
            return []
        if len(slices_data) == 1:
            return slices_data[0] if slices_data[0] else []

        merged = {}  # row_index -> merged cot dict

        for slice_idx, rows in enumerate(slices_data):
            if not isinstance(rows, list):
                continue
            for row in rows:
                idx = row.get('row_index')
                if idx is None:
                    continue
                idx = int(idx)
                if idx not in merged:
                    merged[idx] = {'row_index': idx, 'cot': {}}

                # Merge column values from 'cot' sub-object
                cot = row.get('cot', {})
                if isinstance(cot, dict):
                    for col, val in cot.items():
                        if col not in merged[idx]['cot'] or not merged[idx]['cot'][col]:
                            merged[idx]['cot'][col] = val

                # Also merge top-level keys matching known column headers
                for col in column_headers:
                    if col in row and col not in merged[idx]['cot']:
                        merged[idx]['cot'][col] = row[col]

        # Warn about gaps
        all_indices = sorted(merged.keys())
        if all_indices:
            expected = set(range(1, max(all_indices) + 1))
            missing = expected - set(all_indices)
            if missing:
                _logger.warning("[REVIEW] Missing row indices after merge: %s", sorted(missing))

        return [merged[i] for i in all_indices]

    @staticmethod
    def _json_rows_to_markdown(rows, column_headers):
        """Convert a list of row dicts to a Markdown table string.

        Uses *column_headers* (from classification) as the table header row.
        Each row dict is expected to have ``{"cot": {col: val, ...}}``.
        """
        if not rows or not column_headers:
            return ""

        md_header = '| ' + ' | '.join(column_headers) + ' |'
        md_sep = '| ' + ' | '.join(['---'] * len(column_headers)) + ' |'

        md_rows = []
        for row in rows:
            cot = row.get('cot', {})
            vals = [str(cot.get(col, '')).strip() for col in column_headers]
            md_rows.append('| ' + ' | '.join(vals) + ' |')

        return md_header + '\n' + md_sep + '\n' + '\n'.join(md_rows)

    # ── Main OCR Orchestrator ──────────────────────────────────────────────

    def _extract_pdf_ocr_gemini(self, file_content):
        """Extract text from a scanned PDF using an adaptive OCR pipeline.

        Pipeline overview:
          1. Classify page layout (wide_table / simple_table / prose / form / mixed)
             using a low-res Gemini call on the first 1-2 pages, cached for
             subsequent pages with similar dimensions.
          2. Branch to the appropriate per-layout OCR method.
          3. Combine page results with ``--- Trang N ---`` markers.

        No column names, split ratios, or document-specific keywords are
        hardcoded — all structural details are discovered by Gemini at runtime.
        """
        try:
            params = self.env['ir.config_parameter'].sudo()
            api_key = params.get_param('topic_chatbot.gemini_api_key')
            if not api_key:
                _logger.warning("Gemini API key not set, skipping OCR.")
                return ""

            import fitz

            doc_fitz = fitz.open(stream=file_content, filetype="pdf")
            total_pages = len(doc_fitz)
            extracted_pages = []

            # ── Step 1: Classify layout with caching ─────────────────────────
            classify_cache = {}  # pg_idx -> classify_info
            ref_classify = None
            ref_img_size = None

            pages_to_classify = [0]
            if total_pages > 1:
                pages_to_classify.append(1)

            safe_filename = (self.filename or '').encode('ascii', 'replace').decode('ascii')

            for pg_idx in pages_to_classify:
                page = doc_fitz[pg_idx]
                img = self._render_page_image(page, dpi=100)
                img_b64 = self._img_to_b64(img)

                classify_info = self._classify_page_layout(img_b64, api_key)
                classify_cache[pg_idx] = classify_info

                if ref_classify is None:
                    ref_classify = classify_info
                    ref_img_size = img.size

                _logger.info(
                    "Page %s of %s classified as: %s (cols=%s, narrow=%s, headers=%d)",
                    pg_idx + 1, safe_filename,
                    classify_info.get('layout_type'),
                    classify_info.get('estimated_columns'),
                    classify_info.get('has_narrow_cells'),
                    len(classify_info.get('column_headers', [])),
                )

            # Persist detected layout for the UI
            detected_layout = ref_classify.get('layout_type', 'prose') if ref_classify else 'prose'
            try:
                self.write({'layout_type': detected_layout})
            except Exception:
                pass  # never let metadata write break the OCR flow

            # ── Step 2: Process each page ─────────────────────────────────────
            for page_num, page in enumerate(doc_fitz, start=1):
                pg_idx = page_num - 1

                # Determine classify_info for this page
                if pg_idx in classify_cache:
                    page_classify = classify_cache[pg_idx]
                else:
                    # Check if dimensions match the reference page
                    pix_check = page.get_pixmap(dpi=100)
                    check_size = (pix_check.width, pix_check.height)
                    if ref_img_size and self._is_similar_page(ref_img_size, check_size):
                        page_classify = ref_classify
                    else:
                        img = self._render_page_image(page, dpi=100)
                        img_b64 = self._img_to_b64(img)
                        page_classify = self._classify_page_layout(img_b64, api_key)
                        classify_cache[pg_idx] = page_classify
                        _logger.info(
                            "Page %s reclassified as: %s", page_num,
                            page_classify.get('layout_type'),
                        )

                layout = page_classify.get('layout_type', 'prose')

                # Branch to appropriate pipeline with fallback
                try:
                    if layout in ('prose', 'form'):
                        page_text = self._ocr_page_simple(page, api_key)
                    elif layout == 'simple_table':
                        page_text = self._ocr_page_simple_table(
                            page, api_key, page_classify.get('column_headers', []),
                        )
                    elif layout == 'wide_table':
                        page_text = self._ocr_page_wide_table(page, api_key, page_classify)
                    elif layout == 'mixed':
                        page_text = self._ocr_page_mixed(page, api_key)
                    else:
                        page_text = self._ocr_page_simple(page, api_key)
                except Exception as page_err:
                    _logger.warning(
                        "OCR pipeline '%s' failed for page %s, falling back to simple: %s",
                        layout, page_num, str(page_err),
                    )
                    page_text = self._ocr_page_simple(page, api_key)

                if page_text:
                    extracted_pages.append(f"--- Trang {page_num} ---\n{page_text}")
                else:
                    _logger.warning("Could not OCR page %s of %s", page_num, safe_filename)

            return "\n\n".join(extracted_pages)
        except Exception as e:
            _logger.error("Gemini OCR error for %s: %s", self.filename, str(e))
            return ""

    def _extract_docx_text(self, file_content):
        try:
            docx = zipfile.ZipFile(io.BytesIO(file_content))
            xml_content = docx.read('word/document.xml')
            root = ET.fromstring(xml_content)
            
            # DOCX tags are namespace-prefixed
            namespace = '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}'
            paragraphs = []
            
            for paragraph in root.iter(namespace + 'p'):
                texts = [node.text for node in paragraph.iter(namespace + 't') if node.text]
                if texts:
                    paragraphs.append("".join(texts))
            
            return "\n".join(paragraphs)
        except Exception as e:
            raise UserError(f"DOCX extraction error: {str(e)}")

    def _extract_excel_text(self, file_content):
        """Extract text content from Excel files (.xlsx, .xls)
        
        Features:
        - Support both modern (.xlsx) and legacy (.xls) formats
        - Extract all sheets with proper labeling
        - Handle dates, numbers, and formulas properly
        - Memory-efficient processing for large files
        """
        filename = (self.filename or '').lower()
        text_lines = []

        # Try modern Excel format first (.xlsx)
        if filename.endswith('.xlsx'):
            try:
                import openpyxl
                from openpyxl.utils.datetime import from_excel
                
                wb = openpyxl.load_workbook(
                    filename=io.BytesIO(file_content), 
                    data_only=True,  # Get calculated values instead of formulas
                    read_only=True   # Memory efficient for large files
                )
                
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    text_lines.append(f"--- Sheet: {sheet_name} ---")
                    
                    # Get the actual used range to avoid processing empty cells
                    max_row = sheet.max_row
                    max_col = sheet.max_column
                    
                    if max_row == 1 and max_col == 1:
                        # Empty sheet
                        text_lines.append("(Sheet trống)")
                        continue
                    
                    for row in sheet.iter_rows(min_row=1, max_row=max_row, 
                                             min_col=1, max_col=max_col, values_only=True):
                        row_values = []
                        for val in row:
                            if val is not None:
                                # Handle different data types
                                if isinstance(val, (int, float)):
                                    # Format numbers properly
                                    if val == int(val):
                                        row_values.append(str(int(val)))
                                    else:
                                        row_values.append(f"{val:.2f}".rstrip('0').rstrip('.'))
                                else:
                                    str_val = str(val).strip()
                                    if str_val:
                                        row_values.append(str_val)
                        
                        if row_values:
                            text_lines.append(" | ".join(row_values))
                
                wb.close()
                return "\n".join(text_lines)
                
            except ImportError:
                _logger.warning("openpyxl not installed. Install with: pip install openpyxl")
                # Fall through to xlrd
            except Exception as e:
                _logger.warning("openpyxl extraction failed for %s: %s. Trying xlrd...", 
                              self.filename, str(e))

        # Fallback to xlrd for both .xlsx and .xls
        try:
            import xlrd
            from xlrd import xldate
            
            workbook = xlrd.open_workbook(file_contents=file_content)
            text_lines = []
            
            for sheet_idx in range(workbook.nsheets):
                sheet = workbook.sheet_by_index(sheet_idx)
                text_lines.append(f"--- Sheet: {sheet.name} ---")
                
                if sheet.nrows == 0:
                    text_lines.append("(Sheet trống)")
                    continue
                
                for row_idx in range(sheet.nrows):
                    row_vals = sheet.row_values(row_idx)
                    row_types = sheet.row_types(row_idx)
                    row_str = []
                    
                    for col_idx, (val, cell_type) in enumerate(zip(row_vals, row_types)):
                        if val is not None and str(val).strip():
                            # Handle Excel cell types
                            if cell_type == xlrd.XL_CELL_DATE:
                                # Convert Excel date to readable format
                                try:
                                    date_val = xldate.xldate_as_datetime(val, workbook.datemode)
                                    row_str.append(date_val.strftime('%Y-%m-%d %H:%M:%S'))
                                except:
                                    row_str.append(str(val).strip())
                            elif cell_type == xlrd.XL_CELL_NUMBER:
                                # Format numbers properly
                                if val == int(val):
                                    row_str.append(str(int(val)))
                                else:
                                    row_str.append(f"{val:.2f}".rstrip('0').rstrip('.'))
                            else:
                                str_val = str(val).strip()
                                if str_val:
                                    row_str.append(str_val)
                    
                    if row_str:
                        text_lines.append(" | ".join(row_str))
            
            return "\n".join(text_lines)
            
        except ImportError:
            raise UserError(
                f"Không thể xử lý tệp Excel '{self.filename}'. "
                "Vui lòng cài đặt thư viện cần thiết:\n"
                "pip install openpyxl xlrd"
            )
        except Exception as e:
            raise UserError(f"Lỗi trích xuất tệp Excel '{self.filename}': {str(e)}")

    def _extract_txt_or_csv_text(self, file_content):
        for encoding in ['utf-8', 'utf-8-sig', 'cp1252', 'latin-1', 'gbk']:
            try:
                return file_content.decode(encoding)
            except UnicodeDecodeError:
                continue
        return file_content.decode('utf-8', errors='ignore')

    def _chunk_text(self, text, chunk_size=1500, overlap=200):
        """Split extracted text into RAG-friendly chunks.

        Table-aware: markdown table headers are detected by **pattern** (lines
        starting with ``|`` that contain ``---``, or that immediately follow a
        page/sheet marker) rather than by checking for specific Vietnamese
        keywords.  Detected headers are prepended to every chunk that contains
        rows from that table so that each chunk is self-contained.
        """
        if not text:
            return []

        import re

        # Split by page / sheet markers
        pages = re.split(
            r'(?=(?:\n|\A)--- (?:Trang \d+|Sheet: [^\n]+) ---)', text,
        )
        chunks = []

        for page in pages:
            page_content = page.strip()
            if not page_content:
                continue

            # Keep small pages intact for full RAG context
            if len(page_content) <= 2500:
                chunks.append(page_content)
                continue

            lines = page_content.split('\n')

            # ── Pattern-based header detection (no hardcoded keywords) ───────
            # Pass 1: mark which lines are headers / table lines.
            is_header = [False] * len(lines)
            is_table_line = [False] * len(lines)

            for idx, line in enumerate(lines):
                stripped = line.strip()
                if not stripped.startswith('|'):
                    continue
                is_table_line[idx] = True

                # Rule 1 – separator line (e.g. |---|---|---| )
                if '---' in stripped:
                    is_header[idx] = True
                    # Rule 2 – lines directly *before* the separator are headers
                    for back in range(idx - 1, -1, -1):
                        back_s = lines[back].strip()
                        if back_s.startswith('|') and '---' not in back_s:
                            is_header[back] = True
                        else:
                            break
                    continue

                # Rule 3 – first pipe-line right after a page/sheet marker
                if idx > 0:
                    prev = lines[idx - 1].strip()
                    if re.match(r'^--- (?:Trang \d+|Sheet: .+?) ---$', prev):
                        is_header[idx] = True

            # ── Pass 2: build chunks ─────────────────────────────────────────
            table_hdrs = []    # accumulated header lines for current table
            table_rows = []    # accumulated data rows for current table
            prose_buf = []     # non-table text accumulator
            in_table = False

            def _flush_table():
                """Emit table chunks (header prepended to each)."""
                if not table_rows:
                    return
                hdr_text = "\n".join(table_hdrs)
                n_hdr = len(table_hdrs)
                cur = [hdr_text] if hdr_text else []
                for r in table_rows:
                    cur.append(r)
                    if len("\n".join(cur)) > chunk_size:
                        chunks.append("\n".join(cur))
                        cur = [hdr_text] if hdr_text else []
                if len(cur) > (n_hdr if hdr_text else 0):
                    chunks.append("\n".join(cur))

            def _flush_prose():
                """Emit non-table text as a single chunk."""
                blob = "\n".join(prose_buf).strip()
                if blob:
                    chunks.append(blob)

            for idx, line in enumerate(lines):
                if is_header[idx]:
                    # Entering (or continuing) a table header section
                    if prose_buf:
                        _flush_prose()
                        prose_buf = []
                    table_hdrs.append(line)
                    in_table = True

                elif in_table and is_table_line[idx]:
                    table_rows.append(line)

                else:
                    # Non-table line (or table ended)
                    if in_table:
                        _flush_table()
                        table_hdrs = []
                        table_rows = []
                        in_table = False
                    prose_buf.append(line)

            # Flush anything remaining
            if in_table:
                _flush_table()
            if prose_buf:
                _flush_prose()

        # Fallback to simple sliding-window chunker
        if not chunks:
            start = 0
            text_len = len(text)
            while start < text_len:
                end = min(start + chunk_size, text_len)
                chunk = text[start:end].strip()
                if chunk:
                    chunks.append(chunk)
                start = end - overlap
                if start >= text_len or end >= text_len:
                    break

        return [c.strip() for c in chunks if len(c.strip()) > 10]

    def action_reprocess_document(self):
        """Action to reprocess a single document (useful after fixing errors)."""
        for doc in self:
            if doc.state == 'processing':
                raise UserError(f"Tài liệu '{doc.name}' đang được xử lý. Vui lòng đợi!")
            
            doc.write({
                'state': 'draft',
                'error_message': False,
                'text_content': False,
                'content_length': 0,
                'word_count': 0,
                'processing_time': 0.0
            })
        
        return self.action_process_document()
    
    def action_view_chunks(self):
        """Action to view chunks created from this document."""
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': f'Text Chunks - {self.name}',
            'res_model': 'topic_chatbot.chunk',
            'view_mode': 'tree,form',
            'domain': [('document_id', '=', self.id)],
            'context': {
                'default_document_id': self.id,
                'default_topic_id': self.topic_id.id,
            }
        }
    
    def action_download_extracted_text(self):
        """Action to download extracted text as .txt file."""
        self.ensure_one()
        
        if not self.text_content:
            raise UserError("Tài liệu chưa được xử lý hoặc không có nội dung!")
            
        import base64
        
        filename = f"{self.name}_extracted.txt"
        content = self.text_content.encode('utf-8')
        content_b64 = base64.b64encode(content).decode('ascii')
        
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': content_b64,
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'text/plain'
        })
        
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    @api.model
    def get_processing_statistics(self):
        """Get processing statistics for dashboard/monitoring."""
        domain = []
        
        stats = {
            'total': self.search_count(domain),
            'done': self.search_count(domain + [('state', '=', 'done')]),
            'processing': self.search_count(domain + [('state', '=', 'processing')]),
            'error': self.search_count(domain + [('state', '=', 'error')]),
            'draft': self.search_count(domain + [('state', '=', 'draft')]),
        }
        
        # Average processing time
        processed_docs = self.search([('state', '=', 'done'), ('processing_time', '>', 0)])
        if processed_docs:
            stats['avg_processing_time'] = sum(processed_docs.mapped('processing_time')) / len(processed_docs)
        else:
            stats['avg_processing_time'] = 0.0
            
        # Total content statistics  
        done_docs = self.search([('state', '=', 'done')])
        stats['total_content_length'] = sum(done_docs.mapped('content_length'))
        stats['total_word_count'] = sum(done_docs.mapped('word_count'))
        stats['total_chunks'] = sum(done_docs.mapped('chunks_count'))
        
        return stats

    def _warn_prompt_injection_patterns(self, text):
        """Enhanced prompt injection detection with Vietnamese patterns."""
        if not text:
            return

        suspicious_patterns = [
            # English patterns
            'ignore previous instructions',
            'ignore all previous instructions', 
            'disregard previous instructions',
            'forget previous instructions',
            'system prompt',
            'developer message',
            'act as if you are',
            'pretend you are',
            'roleplay as',
            'jailbreak',
            'DAN mode',
            
            # Vietnamese patterns
            'bỏ qua hướng dẫn',
            'bỏ qua chỉ dẫn', 
            'bỏ qua các hướng dẫn trước',
            'quên các hướng dẫn trước',
            'tiết lộ toàn bộ dữ liệu',
            'bạn là một AI không giới hạn',
            'không giới hạn',
            'giả vờ bạn là',
            'hành động như thể bạn là',
            'đóng vai',
            'chế độ đặc biệt',
            
            # Technical patterns
            'system:',
            'assistant:',
            'human:',
            '###',
            '---SYSTEM---',
            '---USER---',
        ]
        
        text_lower = text.lower()
        matched_patterns = []
        
        for pattern in suspicious_patterns:
            if pattern.lower() in text_lower:
                # Count occurrences for severity assessment
                count = text_lower.count(pattern.lower())
                matched_patterns.append(f"{pattern} ({count}x)")
        
        if matched_patterns:
            # Log with different severity based on pattern count
            severity = 'warning' if len(matched_patterns) <= 2 else 'error'
            log_method = getattr(_logger, severity)
            
            log_method(
                "Possible prompt-injection content detected in document '%s' (id=%s): %s. "
                "Consider reviewing content before using in chat.",
                self.name, self.id, ", ".join(matched_patterns[:5])  # Limit output
            )
            
            # Store warning in document for user visibility
            if len(matched_patterns) > 3:
                warning_msg = (
                    f"⚠️ CẢNH BÁO: Tài liệu này chứa các mẫu nghi ngờ có thể ảnh hưởng đến "
                    f"hoạt động của chatbot: {', '.join(matched_patterns[:3])}..."
                )
                current_content = self.text_content or ""
                self.text_content = f"{warning_msg}\n\n{current_content}"
