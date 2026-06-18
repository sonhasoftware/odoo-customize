from odoo import models, fields, api
from odoo.exceptions import UserError, ValidationError
from datetime import datetime

import base64
from docx import Document
import io
import tempfile
import os
from io import BytesIO
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


class PopupExportDocument(models.TransientModel):
    _name = 'popup.export.document'

    # document = fields.Many2one('exp.export.document', string="Loại tài liệu", required=True,
    #                            domain="[('type', '=', key)]", store=True)
    documents = fields.One2many('detail.document', 'export_id', string="Loại tài liệu", store=True)
    document_file = fields.Binary(string="Tệp tài liệu", store=True)
    key = fields.Char(string="Key", store=True)
    contract_id = fields.Many2one('exp.contract', store=True)
    shipment_id = fields.Many2one('exp.shipment', store=True)
    document_date = fields.Date(string="Ngày tài liệu", default=fields.Date.context_today, store=True)

    def extract_and_format_date(self, text):
        text = text or ""

        # tìm tất cả cụm số 6-8 ký tự
        matches = re.findall(r'\d{6,8}', text)

        for m in matches:
            try:
                # validate date
                if len(m) == 6:
                    datetime.strptime(m, "%d%m%y")
                elif len(m) == 8:
                    datetime.strptime(m, "%d%m%Y")

                # nếu hợp lệ → format luôn
                return self.format_date(m)

            except:
                continue

        return ""

    def format_date(self, date_str):
        if len(date_str) == 6:  # ddmmyy
            dt = datetime.strptime(date_str, "%d%m%y")
        elif len(date_str) == 8:  # ddmmyyyy
            dt = datetime.strptime(date_str, "%d%m%Y")
        else:
            raise ValidationError("Invalid date format")

        day = dt.day

        if 10 <= day <= 20:
            suffix = "th"
        else:
            suffix = {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")

        return dt.strftime(f"%d{suffix} %B %Y")

    def get_value(self, label, text):
        text = (text or "").replace('\r\n', '\n')
        lines = text.split('\n')

        value_lines = []
        found = False

        for line in lines:
            line_strip = line.strip()

            # tìm label
            if not found:
                if line_strip.startswith(f"- {label}:"):
                    found = True
                    # lấy phần sau dấu :
                    after_colon = line.split(":", 1)[1].strip()
                    if after_colon:
                        value_lines.append(after_colon)
                continue

            # nếu gặp field mới → dừng
            if line_strip.startswith("- "):
                break

            # nếu không phải field mới → lấy value
            if line_strip:
                value_lines.append(line_strip)

        return ' '.join(value_lines)

    def action_confirm(self):
        config_data = self.env['exp.config.data'].sudo().search([])
        data_config = {
            f'config_{r.code}': r.data for r in config_data
        }
        string_contract_date = self.extract_and_format_date(self.contract_id.contract_no)
        ship_from = self.contract_id.shipping_port_from.name if self.contract_id.payment_term else ''
        ship_to = self.contract_id.shipping_port_to.name if self.contract_id.payment_term else ''
        term_of_payment = self.contract_id.payment_term.name if self.contract_id.payment_term else ''
        amount_paid = self.contract_id.total_amount - self.contract_id.deposit_amount
        sign_date = self.format_date(self.contract_id.sign_date.strftime("%d%m%Y"))
        contract_data = {
            'contract_no': self.contract_id.contract_no,
            'sign_date': sign_date,
            'customer_name': self.contract_id.customer_id.name if self.contract_id.customer_id else '',
            'customer_address': self.contract_id.customer_id.address if self.contract_id.customer_id else '',
            'customer_phone': self.contract_id.customer_id.phone if self.contract_id.customer_id else '',
            'customer_email': self.contract_id.customer_id.email if self.contract_id.customer_id else '',
            'currency': self.contract_id.currency.TEN if self.contract_id.currency else '',
            'string_contract_date': string_contract_date,
            'ship_from': ship_from,
            'ship_to': ship_to,
            'term_of_payment': term_of_payment,
            'deposit_payment': self.contract_id.deposit_amount,
            'amount_paid': amount_paid,
            'po_number': self.contract_id.po_number if self.contract_id.po_number else '',
            'cont_number': self.shipment_id.so_cont_number if self.shipment_id and self.shipment_id.so_cont_number else '',
            'total_amount': self.contract_id.total_amount,
        }
        data_config.update(contract_data)
        file_document = self.documents.filtered(lambda x: x.radio_tick)
        if not file_document:
            # raise ValidationError("Phải chọn 1 loại tài liệu để tạo!")
            self.env['bus.bus']._sendone(
                (self._cr.dbname, 'res.partner', self.env.user.partner_id.id),
                'simple_notification',
                {
                    'title': "Cảnh báo!",
                    'message': f"Phải chọn 1 loại tài liệu để tạo!",
                    'sticky': False,
                }
            )
            documents = self.env['exp.export.document'].sudo().search([('type', '=', self.key)])
            vals = [(0, 0, {
                'document': doc.id,
                'file_name': doc.name,
                'radio_tick': False,
            }) for doc in documents]
            return {
                'name': 'Chọn mẫu tài liệu',
                'type': 'ir.actions.act_window',
                'res_model': 'popup.export.document',
                'view_mode': 'form',
                'target': 'new',
                'context': {
                    'default_key': self.key,
                    'default_contract_id': self.contract_id.id,
                    'default_documents': vals,
                    'default_shipment_id': self.shipment_id.id if self.shipment_id else None,
                }
            }
        file_stream = io.BytesIO(base64.b64decode(file_document.document.file))
        wb = load_workbook(filename=file_stream)
        sheet = wb['Sheet1']
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                if isinstance(cell.value, str):
                    for key, val in data_config.items():
                        placeholder = f'[{key}]'
                        if placeholder in cell.value:
                            cell.value = cell.value.replace(placeholder, str(val))
        if self.key == 'si':
            row = 23
            for r in self.contract_id.export_detail:
                sheet.cell(row=row, column=1).value = f"CONTAINER NO.: {r.so_cont_number}"
                sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
                sheet.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=row, column=1).font = Font(name='Times New Roman', bold=True, size=11)
                row += 1
        output_stream = io.BytesIO()
        wb.save(output_stream)
        output_stream.seek(0)

        # Trả về file để tải về
        export_file = base64.b64encode(output_stream.read())
        download_filename = f"{file_document.file_name} {self.contract_id.contract_no}.xlsx"
        self.document_file = export_file

        if self.key == 'invoice':
            self.shipment_id.sudo().write({'file_invoice': self.document_file, 'file_invoice_name': download_filename})
        elif self.key == 'si':
            self.contract_id.sudo().write({'si_file': self.document_file, 'si_file_name': download_filename})

        # Tạo record ảo chỉ để tải file
        return {
            'type': 'ir.actions.act_url',
            'url': f"/web/content/popup.export.document/{self.id}/document_file/{download_filename}?download=true",
            'target': 'new',
            'close': True,
        }

