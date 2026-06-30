import base64
import warnings
from io import BytesIO

from openpyxl import load_workbook

from odoo import _, fields, models
from odoo.exceptions import ValidationError


class MDMTongHopImportWizard(models.TransientModel):
    _name = 'mdm.tong.hop.import.wizard'
    _description = 'Import MDM Hàng hóa từ Excel'

    LOOKUP_FIELDS = (
        ('mdm_hh_type_id', 'mdm.hh.type', 3, 'Loại hàng hóa'),
        ('dvt', 'mdm.dvt', 6, 'Đơn vị tính'),
        ('chung_loai1', 'mdm.chung.loai1', 7, 'Chủng loại 1'),
        ('chung_loai2', 'mdm.chung.loai2', 8, 'Chủng loại 2'),
        ('linh_vuc', 'mdm.linh.vuc', 9, 'Lĩnh vực'),
        ('nganh_hang', 'mdm.nganh.hang', 10, 'Ngành hàng'),
        ('nhan_hang', 'mdm.nhan.hang', 11, 'Nhãn hàng'),
        ('chat_lieu', 'mdm.chat.lieu', 12, 'Chất liệu'),
        ('do_bong', 'mdm.do.bong', 13, 'Độ bóng'),
        ('bom_sale', 'bom.sale', 17, 'Loại trong BOM sales'),
    )

    file_data = fields.Binary(string='File Excel', required=True)
    file_name = fields.Char(string='Tên file')

    def _clean_value(self, value):
        if value is None:
            return False
        if isinstance(value, str):
            value = value.strip()
            return value or False
        return str(value).strip() or False

    @staticmethod
    def _merge_non_empty_vals(existing_record, incoming_vals):
        merged_vals = {}
        for field_name, value in incoming_vals.items():
            if value in (False, None, ''):
                continue

            existing_value = existing_record[field_name]
            if hasattr(existing_value, 'id'):
                existing_value = existing_value.id or False
            if existing_value != value:
                merged_vals[field_name] = value

        return merged_vals

    @staticmethod
    def _merge_non_empty_dict(base_vals, incoming_vals):
        merged_vals = dict(base_vals)
        for field_name, value in incoming_vals.items():
            if value not in (False, None, ''):
                merged_vals[field_name] = value
        return merged_vals

    def _read_import_rows(self, sheet):
        import_rows = []
        company_codes = set()
        lookup_codes_by_model = {model_name: set() for _, model_name, _, _ in self.LOOKUP_FIELDS}

        for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_col=18, values_only=True), start=2):
            cleaned = [self._clean_value(value) for value in row]
            if not any(cleaned):
                continue

            company_code = cleaned[0]
            if company_code:
                company_codes.add(company_code)

            for _, model_name, column_index, _ in self.LOOKUP_FIELDS:
                code = cleaned[column_index]
                if code:
                    lookup_codes_by_model[model_name].add(code.lower())

            import_rows.append({
                'row_index': row_index,
                'company_code': company_code,
                'ma_hang_don_vi': cleaned[1],
                'ma_mdm': cleaned[2],
                'values': cleaned,
            })

        return import_rows, company_codes, lookup_codes_by_model

    def _get_lookup_cache(self, lookup_codes_by_model):
        lookup_cache = {}
        for model_name, codes in lookup_codes_by_model.items():
            if not codes:
                lookup_cache[model_name] = {}
                continue

            records = self.env[model_name].sudo().search([('ma', '!=', False)])
            lookup_cache[model_name] = {
                record.ma.strip().lower(): record.id
                for record in records
                if record.ma and record.ma.strip().lower() in codes
            }

        return lookup_cache

    def _get_lookup_id(self, lookup_cache, model_name, code, field_label):
        if not code:
            return False

        lookup_id = lookup_cache.get(model_name, {}).get(code.lower())
        if lookup_id:
            return lookup_id

        raise ValidationError(_('Không tìm thấy dữ liệu ở field %(field)s với mã "%(code)s".', field=field_label, code=code))

    def _prepare_vals(self, row_data, lookup_cache):
        row_values = row_data['values']
        vals = {
            'ma_tg': row_data['ma_hang_don_vi'],
            'ma': row_data['ma_mdm'],
            'ten_ngan': row_values[4],
            'ten': row_values[5],
            'do_day': row_values[14],
            'dung_tich_plus': row_values[15],
            'dvt_dung_tich': row_values[16],
        }

        for field_name, model_name, column_index, field_label in self.LOOKUP_FIELDS:
            vals[field_name] = self._get_lookup_id(lookup_cache, model_name, row_values[column_index], field_label)

        return vals

    def action_import(self):
        self.ensure_one()

        if not self.file_data:
            raise ValidationError(_('Vui lòng chọn file Excel để import.'))

        data = base64.b64decode(self.file_data)
        with warnings.catch_warnings():
            warnings.filterwarnings('ignore', message='Data Validation extension', category=UserWarning)
            workbook = load_workbook(BytesIO(data), data_only=True)
        sheet = workbook.active

        model = self.env['mdm.tong.hop']
        line_model = self.env['mdm.tong.hop.line']

        imported = 0
        updated = 0
        errors = []

        import_rows, company_codes, lookup_codes_by_model = self._read_import_rows(sheet)

        if len(company_codes) > 1:
            raise ValidationError(_('File không cùng 1 mã công ty. Vui lòng kiểm tra lại cột mã công ty trong file import.'))

        if not company_codes:
            raise ValidationError(_('Thiếu mã công ty trong file import.'))

        company_code = next(iter(company_codes))
        company = self.env['res.company'].sudo().search([('company_code', '=', company_code)], limit=1)
        if not company:
            raise ValidationError(_('Không tìm thấy công ty với mã công ty "%(code)s".', code=company_code))

        lookup_cache = self._get_lookup_cache(lookup_codes_by_model)
        ma_mdm_values = {row_data['ma_mdm'] for row_data in import_rows if row_data['ma_mdm']}
        existing_by_ma = {
            record.ma: record
            for record in model.sudo().search([('ma', 'in', list(ma_mdm_values))])
        }

        pending_create_vals_by_ma = {}
        prepared_rows = []
        import_context = {
            'skip_mdm_similarity': True,
            'skip_mdm_api_sync': True,
        }

        for row_data in import_rows:
            ma_mdm = row_data['ma_mdm']
            try:
                if not ma_mdm:
                    raise ValidationError(_('Thiếu Mã MDM ở cột C.'))

                vals = self._prepare_vals(row_data, lookup_cache)
                existing = existing_by_ma.get(ma_mdm)

                if existing:
                    update_vals = self._merge_non_empty_vals(existing, vals)
                    if existing.dvcs.id != company.id:
                        update_vals['dvcs'] = company.id
                    if update_vals:
                        existing.with_context(**import_context).sudo().write(update_vals)
                    updated += 1
                else:
                    current_create_vals = dict(vals, dvcs=company.id)
                    if ma_mdm in pending_create_vals_by_ma:
                        pending_create_vals_by_ma[ma_mdm] = self._merge_non_empty_dict(
                            pending_create_vals_by_ma[ma_mdm],
                            current_create_vals,
                        )
                    else:
                        pending_create_vals_by_ma[ma_mdm] = current_create_vals
                        imported += 1

                prepared_rows.append(row_data)
            except Exception as exc:
                errors.append(_('Dòng %(row)s (Mã MDM: %(ma_mdm)s): %(error)s', row=row_data['row_index'], ma_mdm=ma_mdm or '-', error=str(exc)))

        if errors:
            message = _('Import hoàn tất. Tạo mới: %(new)s, Cập nhật: %(updated)s', new=imported, updated=updated)
            message = message + '\n' + '\n'.join(errors[:20])
            raise ValidationError(message)

        created_by_ma = {}
        if pending_create_vals_by_ma:
            created_records = model.with_context(**import_context).sudo().create(list(pending_create_vals_by_ma.values()))
            created_by_ma = {record.ma: record for record in created_records}

        line_vals_list = []
        for row_data in prepared_rows:
            ma_mdm = row_data['ma_mdm']
            parent_record = existing_by_ma.get(ma_mdm) or created_by_ma.get(ma_mdm)
            if not parent_record:
                continue

            line_vals = {
                'tong_hop_id': parent_record.id,
                'ma_mdm': ma_mdm,
                'dvcs': company.id,
            }
            if row_data['ma_hang_don_vi']:
                line_vals['ma_dv'] = row_data['ma_hang_don_vi']
            line_vals_list.append(line_vals)

        if line_vals_list:
            line_model.sudo().create(line_vals_list)

        return {'type': 'ir.actions.act_window_close'}
