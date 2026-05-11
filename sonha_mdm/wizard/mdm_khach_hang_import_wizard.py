from odoo import fields, models
from odoo.exceptions import ValidationError
import base64
import io
from openpyxl import load_workbook


class MDMTongHopImportWizard(models.TransientModel):
    _name = 'mdm.tong.hop.import.wizard'
    _description = 'Import MDM Hang Hoa'

    file = fields.Binary("File Excel", required=True)
    filename = fields.Char("Tên file")

    def _clean_value(self, value):
        if value is None:
            return False
        if isinstance(value, str):
            value = value.strip()
            return value or False
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _get_m2o_by_code(self, model_name, code, row_idx, field_label):
        if not code:
            return False
        record = self.env[model_name].sudo().search([('ma', '=', code)], limit=1)
        if not record:
            raise ValidationError(f"Dòng {row_idx}: Không tìm thấy mã '{code}' cho cột '{field_label}'.")
        return record.id

    def action_import(self):
        self.ensure_one()
        workbook = load_workbook(filename=io.BytesIO(base64.b64decode(self.file)), data_only=True)
        sheet = workbook.active

        header_map = {
            'mã tg': 'ma_tg',
            'ma tg': 'ma_tg',
            'id': 'ma',
            'loại hàng hóa': 'mdm_hh_type_id',
            'loai hang hoa': 'mdm_hh_type_id',
            'tên ngắn': 'ten_ngan',
            'ten ngan': 'ten_ngan',
            'tên đầy đủ': 'ten',
            'ten day du': 'ten',
            'đvt': 'dvt',
            'dvt': 'dvt',
            'chủng loại 1': 'chung_loai1',
            'chung loai 1': 'chung_loai1',
            'chủng loại 2': 'chung_loai2',
            'chung loai 2': 'chung_loai2',
            'lĩnh vực': 'linh_vuc',
            'linh vuc': 'linh_vuc',
            'ngành hàng': 'nganh_hang',
            'nganh hang': 'nganh_hang',
            'nhãn hàng': 'nhan_hang',
            'nhan hang': 'nhan_hang',
            'chất liệu': 'chat_lieu',
            'chat lieu': 'chat_lieu',
            'độ bóng': 'do_bong',
            'do bong': 'do_bong',
            'độ dày': 'do_day',
            'do day': 'do_day',
            'đvt thể tích': 'dung_tich',
            'dvt the tich': 'dung_tich',
        }

        many2one_model_map = {
            'mdm_hh_type_id': ('mdm.hh.type', 'Loại hàng hóa'),
            'chung_loai1': ('mdm.chung.loai1', 'Chủng loại 1'),
            'chung_loai2': ('mdm.chung.loai2', 'Chủng loại 2'),
            'linh_vuc': ('mdm.linh.vuc', 'Lĩnh vực'),
            'nganh_hang': ('mdm.nganh.hang', 'Ngành hàng'),
            'nhan_hang': ('mdm.nhan.hang', 'Nhãn hàng'),
            'chat_lieu': ('mdm.chat.lieu', 'Chất liệu'),
            'do_bong': ('mdm.do.bong', 'Độ bóng'),
            'do_day': ('mdm.do.day', 'Độ dày'),
            'dung_tich': ('mdm.dung.tich', 'ĐVT thể tích'),
        }

        headers = []
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=col).value
            headers.append((cell or '').strip().lower() if isinstance(cell, str) else '')

        processed = 0
        for row_idx in range(2, sheet.max_row + 1):
            values = {}
            is_empty = True
            for col_idx, header in enumerate(headers, start=1):
                field_name = header_map.get(header)
                if not field_name:
                    continue
                raw_value = self._clean_value(sheet.cell(row=row_idx, column=col_idx).value)
                if raw_value:
                    is_empty = False
                if field_name in many2one_model_map:
                    model_name, field_label = many2one_model_map[field_name]
                    values[field_name] = self._get_m2o_by_code(model_name, raw_value, row_idx, field_label)
                else:
                    values[field_name] = raw_value

            if is_empty:
                continue

            ma_tg = values.get('ma_tg')
            if not ma_tg:
                raise ValidationError(f"Dòng {row_idx} thiếu 'Mã TG'.")

            existed = self.env['mdm.tong.hop'].sudo().search([('ma_tg', '=', ma_tg)], limit=1)
            if existed:
                existed.write(values)
            else:
                if not values.get('ten'):
                    values['ten'] = ma_tg
                self.env['mdm.tong.hop'].sudo().create(values)
            processed += 1

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Import thành công',
                'message': f'Đã xử lý {processed} dòng dữ liệu hàng hóa.',
                'type': 'success',
                'sticky': False,
            }
        }
