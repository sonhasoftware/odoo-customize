from datetime import datetime, time, timedelta, date
from dateutil.relativedelta import relativedelta
from odoo.exceptions import UserError, ValidationError

from odoo import models, fields, api, _
import base64
from docx import Document
import io
import tempfile
import os
from io import BytesIO
from openpyxl import load_workbook


class HrContract(models.Model):
    _inherit = 'hr.contract'

    #
    contract_type_id = fields.Many2one('hr.contract.type', string="Loại hợp đồng", ondelete='cascade')
    employee_code = fields.Char(string="Mã nhân viên", related='employee_id.employee_code')
    mail = fields.Boolean('mail', default=False)
    name = fields.Char(required=False)

    def action_confirm(self):
        for r in self:
            if r.state == 'draft':
                r.state = 'open'

    def action_cancel(self):
        for r in self:
            if r.state == 'draft':
                r.state = 'cancel'

    def cancel_confirm(self):
        for r in self:
            if r.state == 'open':
                r.state = 'draft'

    def action_cron_send_warning_mail_method(self):
        now = datetime.now()
        template = self.env.ref('sonha_employee.warning_mail_template')
        if now.day == 1 or now.day == 15:
            if now.day == 1:
                start = now.replace(day=1) + timedelta(hours=7)
                end_date = start + relativedelta(months=1, days=-1)
            else:
                start = now.replace(day=15) + timedelta(hours=7)
                end_date = start + relativedelta(months=1)
            list_contracts = self.env['hr.contract'].sudo().search([('date_end', '>=', start),
                                                                    ('date_end', '<=', end_date),
                                                                    ('state', '=', 'open')])
            list_expired_contract = []
            for contract in list_contracts:
                if contract.mail == False:
                    contract.mail = True
                    expired_contract = str(contract.employee_id.name) + " - " + str(contract.employee_code) + " có hợp đồng " + str(contract.name) + " sắp hết hạn"
                else:
                    expired_contract = str(contract.employee_id.name) + " - " + str(contract.employee_code) + " có hợp đồng " + str(contract.name) + " sắp hết hạn (Đã gửi mail)"
                list_expired_contract.append(expired_contract)
            body_mail = ', '.join(list_expired_contract)
            custom_body = "<p>Kính gửi HR,<br></br>Dưới đây là danh sách những nhân viên sắp hết hạn:<br></br></p>" + body_mail
            if body_mail:
                template.write({
                    'body_html': custom_body
                })
                template.send_mail(contract.id, force_send=True)
                
    def export_contract(self):
        if not self.contract_type_id.file:
            raise ValidationError("Chưa có mẫu hợp đồng cho loại hợp đồng này!")
        
        dt_now = datetime.now()
        file_data = base64.b64decode(self.contract_type_id.file)

        # Tạo file tạm thời
        with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as tmp_file:
            tmp_file.write(file_data)
            tmp_file_path = tmp_file.name

        try:
            # Mở file Word từ đường dẫn tạm thời
            doc = Document(tmp_file_path)

            # Thay thế các biến với dữ liệu thực tế
            for paragraph in doc.paragraphs:
                if '[name_contract]' in paragraph.text:
                    if self.name:
                        paragraph.text = paragraph.text.replace('[name_contract]', self.name)
                    else:
                        raise ValidationError("Không có tên hợp đồng!")
                if '[date]' in paragraph.text:
                    paragraph.text = paragraph.text.replace('[date]', str(dt_now.day))
                if '[month]' in paragraph.text:
                    paragraph.text = paragraph.text.replace('[month]', str(dt_now.month))
                if '[year]' in paragraph.text:
                    paragraph.text = paragraph.text.replace('[year]', str(dt_now.year))
                if '[name]' in paragraph.text:
                    paragraph.text = paragraph.text.replace('[name]', self.employee_id.name)
                if '[date_of_birth]' in paragraph.text:
                    paragraph.text = paragraph.text.replace('[date_of_birth]', str(self.employee_id.date_birthday.strftime("%d/%m/%Y")) or '')
                if '[permanent_address]' in paragraph.text:
                    paragraph.text = paragraph.text.replace('[permanent_address]', self.employee_id.permanent_address or '')
                if '[cccd]' in paragraph.text:
                    paragraph.text = paragraph.text.replace('[cccd]', str(self.employee_id.number_cccd) or '')
                if '[date_cccd]' in paragraph.text:
                    paragraph.text = paragraph.text.replace('[date_cccd]', str(self.employee_id.date_cccd.strftime("%d/%m/%Y")) or '')
                if '[cccd_nc]' in paragraph.text:
                    paragraph.text = paragraph.text.replace('[cccd_nc]', self.employee_id.place_of_issue or '')
                if '[phone]' in paragraph.text:
                    paragraph.text = paragraph.text.replace('[phone]', str(self.employee_id.sonha_number_phone) or '')
                if '[job_position]' in paragraph.text:
                    paragraph.text = paragraph.text.replace('[job_position]', self.job_id.name or '')
                if '[department]' in paragraph.text:
                    paragraph.text = paragraph.text.replace('[department]', self.department_id.name or '')
                if '[date_contract]' in paragraph.text:
                    paragraph.text = paragraph.text.replace('[date_contract]', str(self.date_start.strftime("%d/%m/%Y")) or '')
                if '[date_sign]' in paragraph.text:
                    paragraph.text = paragraph.text.replace('[date_sign]', str(self.date_start.day) or '')
                if '[month_sign]' in paragraph.text:
                    paragraph.text = paragraph.text.replace('[month_sign]', str(self.date_start.month) or '')
                if '[year_sign]' in paragraph.text:
                    paragraph.text = paragraph.text.replace('[year_sign]', str(self.date_start.year) or '')
                if '[end_date_contract]' in paragraph.text:
                    if self.date_end:
                        paragraph.text = paragraph.text.replace('[end_date_contract]', str(self.date_end.strftime("%d/%m/%Y")) or '')
                    else:
                        raise ValidationError("Không có ngày kết thúc hợp đồng!")
            # Lưu tài liệu vào bộ nhớ
            file_stream = io.BytesIO()
            doc.save(file_stream)
            file_stream.seek(0)

            # Lưu vào attachments
            attachment = self.env['ir.attachment'].create({
                'name': f'Hợp Đồng {self.name}.docx',  # Sử dụng trường name
                'type': 'binary',
                'datas': base64.b64encode(file_stream.read()),
                'res_model': 'hr.contract',
                'res_id': self.id,
            })

            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/{attachment.id}?download=true',
                'target': 'self',
            }
        finally:
            # Xóa file tạm thời
            os.remove(tmp_file_path)

    def notifi_contract_expired(self):
        config_ids = self.env['config.contract'].sudo().search([('sent_mail', '=', True)])
        for config in config_ids:
            today = date.today()
            if today.day == 1:
                start = today.replace(day=15)
                end = start + relativedelta(months=1)
                contracts = self.search([
                    ('company_id', '=', config.company_id.id),
                    ('date_end', '>=', start),
                    ('date_end', '<=', end),
                    ('state', '=', 'open')
                ])

                if not contracts:
                    continue
                attachment = self.export_contracts_to_excel(contracts, start, end, config)
                mail_values = {
                    'subject': f'Danh sách hợp đồng sắp hết hạn!',
                    'body_html': '<p>Kính gửi anh/chị,</p>'
                                 '<p>Dưới đây là danh sách những nhân sự sắp hết hạn:</p>',
                    'email_to': ','.join(config.receiver.mapped('work_email')),
                    'email_cc': ','.join(config.cc_mail.mapped('work_email')),
                    'attachment_ids': [(4, attachment.id)],
                }
                self.env['mail.mail'].create(mail_values).send()
            elif today.day == 15:
                start = (today.replace(day=1)) + relativedelta(month=1)
                end = start + relativedelta(day=-1, month=1)
                contracts = self.search([
                    ('company_id', '=', config.company_id.id),
                    ('date_end', '>=', start),
                    ('date_end', '<=', end)
                ])

                if not contracts:
                    continue
                attachment = self.export_contracts_to_excel(contracts, start, end, config)
                mail_values = {
                    'subject': f'Danh sách hợp đồng sắp hết hạn!',
                    'body_html': '<p>Kính gửi anh/chị,</p>'
                                 '<p>Dưới đây là danh sách những nhân sự sắp hết hạn:</p>',
                    'email_to': ','.join(config.receiver.mapped('work_email')),
                    'email_cc': ','.join(config.cc_mail.mapped('work_email')),
                    'attachment_ids': [(4, attachment.id)],
                }
                self.env['mail.mail'].create(mail_values).send()

    def contract_extend(self):
        today = date.today()
        list_contract = self.env['hr.contract'].sudo().search([('date_start', '=', today),
                                                               ('state', '=', 'draft')])
        for con in list_contract:
            check = self.env['config.contract'].sudo().search([
                ('company_id', '=', con.employee_id.company_id.id)], limit=1)
            if check and check.auto:
                con.state = 'open'

    def export_contracts_to_excel(self, contracts, start_date, end_date, config):
        # 1. Load template

        # template_path = base64.b64encode(open('/sonha_employee/static/src/template/hdld.xlsx', 'rb').read())
        #
        # wb = load_workbook(template_path)
        # ws = wb.active

        file_data = base64.b64decode(config.file)  # config.file là binary field
        wb = load_workbook(filename=io.BytesIO(file_data))  # load workbook từ memory
        ws = wb.active

        current_value = ws['A1'].value
        if current_value:
            current_value = current_value.replace('[start_date]', start_date.strftime('%d/%m/%Y'))
            current_value = current_value.replace('[end_date]', end_date.strftime('%d/%m/%Y'))
            ws['A1'].value = current_value

        # 3. Fill dữ liệu từ dòng 4
        row = 4
        for idx, contract in enumerate(contracts, start=1):
            ws.cell(row=row, column=1, value=idx)  # STT
            ws.cell(row=row, column=2, value=contract.name or '')  # Mã HĐ/Số HĐ
            ws.cell(row=row, column=3, value=contract.employee_id.employee_code or '')  # Mã nhân sự
            ws.cell(row=row, column=4, value=contract.employee_id.name or '')  # Tên nhân sự
            ws.cell(row=row, column=5, value=contract.department_id.name or '')  # Phòng ban
            ws.cell(row=row, column=6, value=contract.job_id.name or '')  # Vị trí
            ws.cell(row=row, column=7,
                    value=contract.date_start.strftime('%d/%m/%Y') if contract.date_start else '')  # Hiệu lực từ ngày
            ws.cell(row=row, column=8,
                    value=contract.date_end.strftime('%d/%m/%Y') if contract.date_end else '')  # Đến ngày
            ws.cell(row=row, column=9, value='Đang chạy')  # Tình trạng
            ws.cell(row=row, column=10, value=contract.contract_type_id.name or '')  # Tên hợp đồng
            row += 1

        # 4. Xuất file Excel ra attachment
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        attachment = self.env['ir.attachment'].create({
            'name': f'{start_date.strftime("%d%m%Y")}_{end_date.strftime("%d%m%Y")} DS cảnh báo HĐLĐ.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(output.read()),
            'res_model': 'hr.contract',
            'res_id': contracts[0].id if contracts else False,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return attachment

    def create(self, vals):
        res = super(HrContract, self).create(vals)
        check = self.env['config.contract'].sudo().search([('company_id', '=', res.company_id.id)])
        if check.auto_code:
            contract = self.sudo().search([('employee_id', '=', res.employee_id.id)])
            res.name = f'{res.employee_code}-0{len(contract)}'
        return res

