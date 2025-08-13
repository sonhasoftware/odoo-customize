from odoo import models, fields, api
from odoo.exceptions import UserError, ValidationError
import base64
import io
from openpyxl import load_workbook
from datetime import datetime


class WizardImportInfor(models.TransientModel):
    _name = 'wizard.import.infor'

    file = fields.Binary("File")
    model = fields.Selection([('customer', "Khách hàng"),
                              ('vendor', "Nhà cung cấp"),
                              ('material', "Vật tư")], string="Model")

    def action_confirm(self):
        file_stream = io.BytesIO(base64.b64decode(self.file))
        wb = load_workbook(filename=file_stream, data_only=True)
        if self.model == 'customer':
            sheet = wb['Sheet1']
            sheet_2 = wb['Sheet2']
            max_row = sheet.max_row
            row = 7
            while row <= max_row:
                search = self.env['declare.md.customer'].sudo().search(['|',
                                                                        ('tax_code', '=', sheet.cell(row=row, column=28).value),
                                                                        ('cccd_number', '=', sheet.cell(row=row, column=12).value)])
                if not search:
                    company_code = self.env['company.code'].sudo().search([('code', '=', sheet_2.cell(row=row, column=3).value)])
                    company_id = company_code.company_id
                    payment_methods = sheet_2.cell(row=row, column=8).value
                    list_method = [t.strip() for t in payment_methods.split(',') if t.strip()]
                    method_ids = self.env['cus.payment.method'].sudo().search([('code', 'in', list_method)]).ids
                    vals = {
                        'customer_name': sheet.cell(row=row, column=7).value,
                        'cust_phone': sheet.cell(row=row, column=22).value,
                        'address': sheet.cell(row=row, column=13).value,
                        'tax_code': sheet.cell(row=row, column=28).value,
                        'bp_type_id': self.env['bp.type'].sudo().search([('type_code', '=', sheet.cell(row=row, column=33).value)]).id,
                        'customer_group_id': self.env['bp.group.account'].sudo().search([('group_code', '=', sheet.cell(row=row, column=3).value)]).id,
                        'country_id': self.env['res.country'].sudo().search([('name', '=', sheet.cell(row=row, column=20).value)]).id,
                        'state_id': self.env['res.country.state'].sudo().search([('name', '=', sheet.cell(row=row, column=17).value)]).id,
                        'district_id': self.env['config.district'].sudo().search([('name', '=', sheet.cell(row=row, column=16).value)]).id,
                        'cccd_number': sheet.cell(row=row, column=12).value,
                        'mail_address': sheet.cell(row=row, column=27).value,
                        'postal_code': sheet.cell(row=row, column=19).value,
                        'language': sheet.cell(row=row, column=21).value,
                        'fax_number': sheet.cell(row=row, column=25).value,
                        'bank_number': sheet.cell(row=row, column=31).value,
                        'bank_owner': sheet.cell(row=row, column=32).value,
                        'vendor_code': sheet.cell(row=row, column=11).value,
                        'bank_country_code': self.env['bank.country'].sudo().search([('code', '=', sheet.cell(row=row, column=29).value)]).id,
                        'bank_key': self.env['bank.data'].sudo().search([('bank_key', '=', sheet.cell(row=row, column=30).value)]).id,
                        'company_code': company_code.id,
                        'company_id': company_id.id,
                        'trading_partner': self.env['trading.partner'].sudo().search([('name', '=', sheet_2.cell(row=row, column=5).value)]).id,
                        'reconciliation_account': self.env['reconciliation.account'].sudo().search([('name', '=', sheet_2.cell(row=row, column=6).value)], limit=1).id,
                        'payment_term': self.env['payment.term'].sudo().search([('name', '=', sheet_2.cell(row=row, column=7).value)]).id,
                        'payment_method': [(6, 0, method_ids)],
                    }
                    self.env['declare.md.customer'].sudo().create(vals)
                row += 1
            sheet_3 = wb['Sheet3']
            max_row = sheet.max_row
            row = 7
            while row <= max_row:
                search = self.env['declare.md.customer'].sudo().search([('customer_name', '=',  sheet_3.cell(row=row, column=2).value),
                                                                        ('cccd_number', '=', sheet_3.cell(row=row, column=3).value),
                                                                        ('tax_code', '=', sheet_3.cell(row=row, column=4).value)])
                sale_organization = self.env['sale.organization'].sudo().search([('code', '=', sheet_3.cell(row=row, column=5).value)]).id
                distribution_channel = self.env['distribution.channel'].sudo().search([('code', '=', sheet_3.cell(row=row, column=6).value)]).id
                division = self.env['x.division'].sudo().search([('code', '=', sheet_3.cell(row=row, column=7).value)]).id
                search_sale = self.env['customer.sale'].sudo().search([('sale_organization', '=', sale_organization),
                                                                       ('distribution_channel', '=', distribution_channel),
                                                                       ('division', '=', division)])
                if search and not search_sale:
                    currency = {
                        'VND': 'vnd',
                        'USD': 'usd',
                        'EUR': 'eur',
                        'CNY': 'cny',
                        'MMK': 'mmk',
                    }
                    currency_label = sheet_3.cell(row=row, column=12).value
                    currency_value = currency.get(currency_label)
                    tax = {
                        '0': 'no',
                        '1': 'yes'
                    }
                    tax_label = sheet_3.cell(row=row, column=24).value
                    tax_value = tax.get(tax_label)
                    raw_input = sheet_3.cell(row=row, column=16).value
                    order_combination = str(raw_input).strip().lower() == 'x'
                    vals = {
                        'sale_organization': sale_organization,
                        'distribution_channel': distribution_channel,
                        'division': division,
                        'sale_district': self.env['sale.district'].sudo().search([('code', '=', sheet_3.cell(row=row, column=8).value)]).id,
                        'cus_group_1': self.env['cus.group.1'].sudo().search([('name', '=', sheet_3.cell(row=row, column=9).value)], limit=1).id,
                        'cus_group': self.env['cus.group'].sudo().search([('sap_code', '=', sheet_3.cell(row=row, column=10).value)], limit=1).id,
                        'sale_office': self.env['sale.office'].sudo().search([('branch_code', '=', sheet_3.cell(row=row, column=11).value)]).id,
                        'customer_price_group': self.env['customer.price.group'].sudo().search([('code', '=', sheet_3.cell(row=row, column=13).value)]).id,
                        'plant': self.env['stock.warehouse'].sudo().search([('plant', '=', sheet_3.cell(row=row, column=15).value)]).id,
                        'shipping_condition': self.env['shipping.condition'].sudo().search([('code', '=', sheet_3.cell(row=row, column=17).value)]).id,
                        'incoterm': self.env['cus.incoterm'].sudo().search([('code', 'ilike', sheet_3.cell(row=row, column=19).value)]).id,
                        'incoterm_more': sheet_3.cell(row=row, column=20).value,
                        'payment_term': self.env['payment.term'].sudo().search([('name', '=', sheet_3.cell(row=row, column=22).value)]).id,
                        'customer_assignment_group': self.env['customer.assignment.group'].sudo().search([('code', '=', sheet_3.cell(row=row, column=23).value)]).id,
                        'declare_md_customer': search.id,
                        'currency': currency_value,
                        'tax_classification': tax_value,
                        'order_combination': order_combination,
                    }
                    self.env['customer.sale'].sudo().create(vals)
                row += 1
            sheet_4 = wb['Sheet4']
            max_row = sheet.max_row
            row = 7
            while row <= max_row:
                search = self.env['declare.md.customer'].sudo().search([('customer_name', '=', sheet_4.cell(row=row, column=2).value),
                                                                        ('cccd_number', '=', sheet_4.cell(row=row, column=3).value),
                                                                        ('tax_code', '=', sheet_4.cell(row=row, column=4).value)])
                sale_organization = self.env['sale.organization'].sudo().search([('code', '=', sheet_4.cell(row=row, column=5).value)]).id
                distribution_channel = self.env['distribution.channel'].sudo().search([('code', '=', sheet_4.cell(row=row, column=6).value)]).id
                division = self.env['x.division'].sudo().search([('code', '=', sheet_4.cell(row=row, column=7).value)]).id
                search_sale = self.env['customer.sale'].sudo().search([('sale_organization', '=', sale_organization),
                                                                       ('distribution_channel', '=', distribution_channel),
                                                                       ('division', '=', division),
                                                                       ('declare_md_customer', '=', search.id)])
                if search_sale:
                    vals = {
                        'sold_to': self.env['declare.md.customer'].sudo().search([('id', '=', sheet_4.cell(row=row, column=8).value)]).id,
                        'ship_to': self.env['declare.md.customer'].sudo().search([('id', '=', sheet_4.cell(row=row, column=9).value)]).id,
                        'bill_to': self.env['declare.md.customer'].sudo().search([('id', '=', sheet_4.cell(row=row, column=10).value)]).id,
                        'payer': self.env['declare.md.customer'].sudo().search([('id', '=', sheet_4.cell(row=row, column=11).value)]).id,
                        'contract_person': sheet_4.cell(row=row, column=12).value,
                        'sale_employee': self.env['declare.md.saleman'].sudo().search([('code', '=', sheet_4.cell(row=row, column=13).value)], limit=1).id,
                    }
                    search_sale.sudo().write(vals)
                row += 1
            sheet_5 = wb['Sheet5']
            max_row = sheet.max_row
            row = 6
            while row <= max_row:
                search = self.env['declare.md.customer'].sudo().search([('customer_name', '=', sheet_5.cell(row=row, column=2).value),
                                                                        ('cccd_number', '=', sheet_5.cell(row=row, column=3).value),
                                                                        ('tax_code', '=', sheet_5.cell(row=row, column=4).value)])
                limit_area = self.env['limit.area'].sudo().search([('area', '=', sheet_5.cell(row=row, column=5).value)]).id
                search_sale = self.env['credit.limit'].sudo().search([('credit_segment', '=', limit_area),
                                                                      ('declare_customer', '=', search.id)])
                if search and not search_sale:
                    raw_date = sheet_5.cell(row=row, column=7).value
                    date = datetime.strptime(raw_date.strip(), '%d.%m.%Y')
                    date_str = date.strftime('%Y-%m-%d')
                    vals = {
                        'credit_segment': limit_area,
                        'credit_limit': sheet_5.cell(row=row, column=6).value,
                        'valid_date': date_str,
                        'declare_customer': search.id,
                    }
                    self.env['credit.limit'].sudo().create(vals)
                row += 1
        if self.model == 'vendor':
            sheet = wb['Sheet1']
            sheet_2 = wb['Sheet2']
            max_row = sheet.max_row
            row = 14
            row_2 = 12
            while row <= max_row:
                search = self.env['declare.md.supplier'].sudo().search([('tax_code', '=', sheet.cell(row=row, column=27).value)])
                company_code = self.env['company.code'].sudo().search([('code', '=', sheet_2.cell(row=row_2, column=4).value)])
                company_id = company_code.company_id
                payment_methods = sheet_2.cell(row=row_2, column=13).value
                list_method = [t.strip() for t in payment_methods.split(',') if t.strip()]
                method_ids = self.env['cus.payment.method'].sudo().search([('code', 'in', list_method)]).ids
                double_invoice = sheet_2.cell(row=row_2, column=12).value
                double_invoice_bool = str(double_invoice).strip().lower() == 'x'
                clear_debt = sheet_2.cell(row=row_2, column=14).value
                clear_debt_bool = str(clear_debt).strip().lower() == 'x'
                vendor_lock = sheet_2.cell(row=row_2, column=15).value
                vendor_lock_bool = str(vendor_lock).strip().lower() == 'x'
                if not search:
                    vals = {
                        'name': sheet.cell(row=row, column=7).value,
                        'supp_phone': sheet.cell(row=row, column=21).value,
                        'tax_code': sheet.cell(row=row, column=27).value,
                        'address': sheet.cell(row=row, column=12).value,
                        'supp_type': self.env['bp.type'].sudo().search([('type_code', '=', sheet.cell(row=row, column=33).value)]).id,
                        'supp_group': self.env['bp.group.account'].sudo().search([('group_code', '=', sheet.cell(row=row, column=3).value)]).id,
                        'country_id': self.env['res.country'].sudo().search([('name', '=', sheet.cell(row=row, column=19).value)]).id,
                        'state_id': self.env['res.country.state'].sudo().search([('name', '=', sheet.cell(row=row, column=16).value)]).id,
                        'district_id': self.env['config.district'].sudo().search([('name', '=', sheet.cell(row=row, column=15).value)]).id,
                        'postal_code': sheet.cell(row=row, column=18).value,
                        'mail_address': sheet.cell(row=row, column=26).value,
                        'language': sheet.cell(row=row, column=20).value,
                        'fax_number': sheet.cell(row=row, column=24).value,
                        'bank_number': sheet.cell(row=row, column=30).value,
                        'bank_owner': sheet.cell(row=row, column=31).value,
                        'bank_country_code': self.env['bank.country'].sudo().search([('code', '=', sheet.cell(row=row, column=28).value)]).id,
                        'bank_key': self.env['bank.data'].sudo().search([('bank_key', '=', sheet.cell(row=row, column=29).value)]).id,
                        'industry': self.env['vendor.industry'].sudo().search([('vendor_type', '=', sheet.cell(row=row, column=34).value)]).id,
                        'form_of_org': self.env['form.organization'].sudo().search([('form_of_org', '=', sheet.cell(row=row, column=35).value)]).id,
                        'deposit_proportion': sheet.cell(row=row, column=36).value,
                        'company_id': company_id.id,
                        'company_code': company_code.id,
                        'payment_method': method_ids,
                        'trading_partner': self.env['trading.partner'].sudo().search([('name', '=', sheet_2.cell(row=row_2, column=6).value)]).id,
                        'reconciliation_account': self.env['reconciliation.account'].sudo().search([('name', '=', sheet_2.cell(row=row_2, column=7).value)], limit=1).id,
                        'planning_group': self.env['planning.group'].sudo().search([('name', '=', sheet_2.cell(row=row_2, column=8).value)]).id,
                        'payment_term': self.env['payment.term'].sudo().search([('name', '=', sheet_2.cell(row=row_2, column=9).value)]).id,
                        'customer_code': sheet_2.cell(row=row_2, column=5).value,
                        'previous_num': sheet_2.cell(row=row_2, column=9).value,
                        'double_invoice': double_invoice_bool,
                        'clear_debt': clear_debt_bool,
                        'vendor_block': vendor_lock_bool,
                    }
                    self.env['declare.md.supplier'].sudo().create(vals)
                row += 1
                row_2 += 1
            sheet_3 = wb['Sheet3']
            max_row = sheet_3.max_row
            row = 12
            while row <= max_row:
                search = self.env['declare.md.supplier'].sudo().search([('tax_code', '=', sheet_3.cell(row=row, column=3).value)])
                purchasing_org = self.env['sale.organization'].sudo().search([('code', '=', sheet_3.cell(row=row, column=4).value)])
                vendor_sale = self.env['vendor.purchase'].sudo().search([('purchasing_org', '=', purchasing_org.id),
                                                                         ('declare_vendor', '=', search.id)])
                if search and not vendor_sale:
                    minimum_order = sheet_3.cell(row=row, column=11).value
                    if isinstance(minimum_order, str):
                        if ',' in minimum_order:
                            minimum_order = minimum_order.replace(',', '.')
                        elif not minimum_order.isdigit() and '.' not in minimum_order:
                            minimum_order = 0
                    minimum_order = float(minimum_order)
                    indicator = sheet_3.cell(row=row, column=7).value
                    indicator_bool = str(indicator).strip().lower() == 'x'
                    vals = {
                        'purchasing_org': purchasing_org.id,
                        'incoterm': self.env['cus.incoterm'].sudo().search([('code', '=', sheet_3.cell(row=row, column=5).value)]).id,
                        'order_currency': self.env['order.currency'].sudo().search([('name', '=', sheet_3.cell(row=row, column=10).value)]).id,
                        'purchasing_group': self.env['purchasing.group'].sudo().search([('purchasing_group_code', '=', sheet_3.cell(row=row, column=8).value)]).id,
                        'plan_del_time': sheet_3.cell(row=row, column=9).value,
                        'representative': sheet_3.cell(row=row, column=13).value,
                        'declare_vendor': search.id,
                        'minimum_order': minimum_order,
                        'indicator': indicator_bool,
                    }
                    self.env['vendor.purchase'].sudo().create(vals)
                row += 1
        return {
            'type': 'ir.actions.client',
            'tag': 'reload',
        }
